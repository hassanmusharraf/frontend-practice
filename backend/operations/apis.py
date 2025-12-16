from django.conf import settings
from rest_framework.views import APIView
from django.core.exceptions import ValidationError
from django.db.models import ProtectedError, Count, Q, F, Value, Sum
from django.db.models.functions import TruncMonth
from core.response import StandardResponse, ServiceError
from django.db import transaction, IntegrityError
from .models import (
    PurchaseOrder, PurchaseOrderLine, Consignment, PurchaseOrderUpload, ConsignmentPackaging,
    AWBFile, PackagingAllocation, ConsignmentPOLine
    )
from .serializers import PurchaseOrderLineSerializer, GetPurchaseOrderSerializer
from entities.models import Supplier, StorerKey, Client
from portal.utils import get_all_fields, convert_to_decimal, empty_directory
from portal.choices import Role, PurchaseOrderStatusChoices, ConsignmentStatusChoices, PackageStatusChoices, Role, OperationUserRole, POImportFormatsChoices
from uuid import uuid4
from datetime import timedelta
from portal.mixins import SearchAndFilterMixin, PaginationMixin
from django.utils import timezone
import calendar
from .mixins import ConsignmentStatusSummary, PurchaseOrderMixin, FilterMixin
from portal.service import ExcelService
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from decimal import Decimal
import os
from rest_framework.parsers import MultiPartParser, FormParser
from core.decorators import role_required
from .services import POLineService, ConsignmentServices, PurchaseOrderService
from operations.utils import validate_file_size, addresses_and_pickup
from .notifications import NotificationService
from .other_services.po_import import POImportValidationService

  
from .signals import awb_file_added_audit_trail, awb_file_deleted_audit_trail



def allow_fields_to_update(model):
    exclude_fields = ['created_at', 'updated_at', 'is_active', 'is_deleted', 'deleted_at', 'reference_number', 'customer_reference_number']
    fields = [field.name for field in model._meta.get_fields() if field.concrete and not field.is_relation and not field.primary_key and field.name not in exclude_fields]   
    return fields
    

def seller_details(seller_data):
        data_to_sent = {}
        fields = ["address_line_1", "address_line_2", "city", "state", "country", "postal_code", "phone_number", "tax_number", "email"]
        for f in fields:
            data_to_sent["seller_" + f] = seller_data.get(f, "") 
        return data_to_sent

# def update_po_quantity(po):
#     # print("po",po)
#     sum = PurchaseOrderLine.objects.filter(purchase_order=po).aggregate(sum=Sum('quantity'))
#     po.open_quantity = sum["sum"]    
#     # print("po.open_quantity",po.open_quantity)
#     po.save()


    

    
class PurchaseOrderCreateAPI(APIView):  
    
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        po_data = request.data.copy()
        supplier_data = po_data.get("seller_details")
        client_data = po_data.get("buyer_details")
        po_lines = po_data.get("pieces_detail", [])
            
        storerkey_object = StorerKey.objects.filter(storerkey_code=po_data.get("storerkey")).first()
        if not storerkey_object:
            return StandardResponse(success=False, errors=["Storer Key not found"], status=400)
            
        seller_code = supplier_data.get("seller_code")
        supplier_object = Supplier.objects.filter(supplier_code=seller_code).first()
        if not supplier_object:
            return StandardResponse(success=False, errors=[f"Supplier with code '{seller_code}' not found"], status=400)
            
        client_object = supplier_object.client
        buyer_code = client_data.get("buyer_code")
        
        if client_object.client_code != buyer_code:
            return StandardResponse(success=False, errors=["Seller Code and Buyer codes do not linked"], status=400)
        
        if storerkey_object.client.id != client_object.id:
            return StandardResponse(success=False, errors=["Storer Key is not linked to the client"], status=400)

        if not supplier_object.storerkeys.filter(id=storerkey_object.id).exists():
            return StandardResponse(success=False, errors=["Storer Key is not linked to the supplier"], status=400)
            
        if not po_lines:
            return StandardResponse(success=False, errors=["Purchase Order lines required"], status=400)
        
        reference_number = po_data.get("reference_number")
        if PurchaseOrder.objects.filter(reference_number=reference_number).exists():
            return StandardResponse(success=False, errors=[f"Purchase Order with reference number {reference_number} already exists"], status=400)
        
        customer_reference_number = po_data.get("customer_reference_number")
        if PurchaseOrder.objects.filter(customer_reference_number=customer_reference_number).exists():
            return StandardResponse(success=False, errors=[f"Purchase Order with customer reference number {customer_reference_number} already exists"], status=400)
            
        po_line_ref_numbers = [i.get("reference_number") for i in po_lines]
        po_line_cust_ref_numbers = [i.get("customer_reference_number") for i in po_lines]
        
        existing_po_line_refs = set(
            PurchaseOrderLine.objects.filter(reference_number__in=po_line_ref_numbers)
            .values_list("reference_number", flat=True)
        )

        existing_po_line_cust_refs = set(
            PurchaseOrderLine.objects.filter(customer_reference_number__in=po_line_cust_ref_numbers)
            .values_list("customer_reference_number", flat=True)
        )

        errors = []
        if existing_po_line_refs:
            errors.append(f"Purchase Order Line(s) with reference number(s) {', '.join(existing_po_line_refs)} already exist.")

        if existing_po_line_cust_refs:
            errors.append(f"Purchase Order Line(s) with customer reference number(s) {', '.join(existing_po_line_cust_refs)} already exist.")

        if errors:
            return StandardResponse(success=False, errors=errors, status=400)
        
        
        po_data["client"] = client_object
        po_data["supplier"] = supplier_object
        po_data["storerkey"] = storerkey_object
        po_data.update(seller_details(supplier_data))        
        try:
            # po_serailizer = PurchaseOrderSerializer(data=po_data)
            # if not po_serailizer.is_valid():
            #     transaction.set_rollback(True)
            #     return StandardResponse(
            #         success=False,
            #         errors=po_serailizer.errors,
            #         status=400
            #     )
            # po_object = po_serailizer.save()
            del po_data["seller_details"]
            del po_data["pieces_detail"]
                        
            po_object = PurchaseOrder.objects.create(**po_data)
            po_lines_to_create = []
            for po_line in po_lines:
                po_line["purchase_order"] = po_object.id
                po_line["open_quantity"] = po_line.get("quantity")               
                po_line["purchase_order"] = po_object
                po_lines_to_create.append(PurchaseOrderLine(**po_line))
            
            if len(po_lines_to_create) > 0:
                obj = PurchaseOrderLine.objects.bulk_create(po_lines_to_create)

            POLineService.update_line_quantities(obj)

            po_object.update_quantity()
        
            return StandardResponse(
                success=True,
                message="Purchase Order created successfully",
                status=201
            )
        
        except IntegrityError as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, message="Database error.", errors=[str(e)])
    
        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(success=False, errors=[str(e)], status=400)


class PurchaseOrderBulkCreateAPI(APIView,PurchaseOrderMixin):
    
    @role_required(Role.ADMIN)
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            data = request.data.get("data", [])
            if not data:
                return StandardResponse(success=False, errors=["Data not found"], status=400)

            validation_errors = self.validate_po_entries(data)
            if validation_errors:
                return StandardResponse(success=False, errors=validation_errors, status=400)
            
            _, errors = self.create_purchase_orders(data)
            if errors:
                return StandardResponse(success=False, errors=errors.values(), status=400)

            return StandardResponse(success=True, message="Purchase Order created successfully", status=201)

        except Exception as e:
            return StandardResponse(errors=[str(e)], status=400)
        # if request.data.get("data",[]):

        #     po_to_create = []
        #     po_lines_to_create = []
        #     data = request.data.get("data", [])
 
        #     validation_errors = self.validate_po_entries(data)
           
        #     if validation_errors:
        #         return StandardResponse(success=False, errors=validation_errors, status=400)

            # for po_index, po_data in enumerate(data):
 
            #     supplier_data = po_data.get("seller_details")
            #     client_data = po_data.get("buyer_details")
            #     po_lines = po_data.get("pieces_detail", [])

            #     try:
            #         storerkey_object = StorerKey.objects.get(storerkey_code=po_data.get("storer_key"))
            #     except StorerKey.DoesNotExist:
            #         return StandardResponse(success=False, errors=[f"Storer Key '{po_data.get('storer_key')}' not found"], status=400)

            #     # try:
            #     #     supplier_object = Supplier.objects.get(supplier_code=supplier_data.get("seller_code"))
            #     # except Supplier.DoesNotExist:
            #     #     return StandardResponse(success=False, errors=[f"Supplier with code '{supplier_data.get('seller_code')}' not found"], status=400)

            #     # client_object = supplier_object.client
            #     buyer_code = client_data.get("buyer_code")
                
            #     # if client_object.client_code != buyer_code:
            #     #     return StandardResponse(success=False, errors=["Seller Code and Buyer codes do not linked"], status=400)
            #     try:
            #         supplier_object, client_object = self.enrich_party_details(supplier_data, buyer_code)
            #     except ValueError as e:
            #         raise Exception(str(e)) 
            #     if storerkey_object.client.id != client_object.id:
            #         return StandardResponse(success=False, errors=["Storer Key is not linked to the client"], status=400)

            #     if not supplier_object.storerkeys.filter(id=storerkey_object.id).exists():
            #         return StandardResponse(success=False, errors=["Storer Key is not linked to the supplier"], status=400)   
            #     customer_reference_number = po_data.get("customer_reference_number")
            #     if PurchaseOrder.objects.filter(customer_reference_number=customer_reference_number).exists():
            #         return StandardResponse(success=False, errors=[f"Purchase Order with customer reference number {customer_reference_number} already exists"], status=400)
                
            #     po_line_cust_ref_numbers = [i.get("customer_reference_number") for i in po_lines]
                
            #     unique_po_line_Cust_ref_numbers = set(po_line_cust_ref_numbers)
            #     errors = []
            #     for crm in unique_po_line_Cust_ref_numbers:
            #         if po_line_cust_ref_numbers.count(crm) > 1:
            #             errors.append(f"Duplicate customer reference number '{crm}' in the '{customer_reference_number}'.")  

            #     if errors:
            #         return StandardResponse(success=False, errors=errors, status=400)
                
            #     po_id = uuid4()
            #     po_data.update({
            #         "id": po_id,
            #         "client": client_object,
            #         "supplier": supplier_object,
            #         "storerkey": storerkey_object
            #     })
            #     po_data.update(seller_details(supplier_data))
                
            #     del po_data["storer_key"]
            #     del po_data["seller_details"]
            #     del po_data["pieces_detail"]
            #     po_to_create.append(PurchaseOrder(**po_data))
            #     for po_line in po_lines:
            #         po_line["purchase_order_id"] = po_id
            #         po_line["open_quantity"] = po_line.get("quantity")               
            #         po_lines_to_create.append(PurchaseOrderLine(**po_line))
            
            # # try:
            # pos = PurchaseOrder.objects.bulk_create(po_to_create)
            # PurchaseOrderLine.objects.bulk_create(po_lines_to_create)

            # for po in pos:
            #     update_po_quantity(po)
            
            # return StandardResponse(
            #     success=True,
            #     message="Purchase Order created successfully",
            #     status=201
            # )
            
            # except IntegrityError as e:
            #     transaction.set_rollback(True)
            #     return StandardResponse(status=400, success=False, message="Database error.", errors=[str(e)])
            # except ValidationError as e:
            #     transaction.set_rollback(True)
            #     return StandardResponse(status=400, success=False, message="Database error.", errors=[str(e)])
            # except Exception as e:
            #     transaction.set_rollback(True)
            #     return StandardResponse(success=False, errors=[str(e)], status=400)
        # else:
        #     return StandardResponse(success=False, errors=["Data not found"], status=400)
    
    @role_required(Role.ADMIN)
    def delete(self, request, *args, **kwargs):
        po_crn = request.query_params.get("po_crn")
        if not po_crn:
            return StandardResponse(success=False, errors=["Data not found"], status=400)

        po = PurchaseOrder.objects.prefetch_related("lines").get(customer_reference_number=po_crn)
        
        lines = po.lines.all()
        allowed_statuses = [PurchaseOrderStatusChoices.OPEN, PurchaseOrderStatusChoices.CANCELLED]
        
        lines_in_process = lines.filter(processed_quantity__gt=0)
        if lines_in_process.exists():
            return StandardResponse(success=False, errors=["Some lines are already processed and cannot be deleted."], status=400)
        
        allowed_statuses = [PurchaseOrderStatusChoices.OPEN, PurchaseOrderStatusChoices.CANCELLED]

        # Check if all lines are in allowed statuses
        if not lines.exclude(status__in=allowed_statuses).exists():
            # Only update lines that are not already CLOSED
            lines_to_update = lines.exclude(status=PurchaseOrderStatusChoices.CANCELLED)
            lines_to_update.update(status=PurchaseOrderStatusChoices.CANCELLED)

            po.status = PurchaseOrderStatusChoices.CANCELLED
            po.save(update_fields=["status"])

            return StandardResponse(success=True, message="Purchase Order deleted successfully", status=200)

        return StandardResponse(success=False, errors=["This item is part of a consignment and cannot be removed."], status=400)

        
            
class PurchaseOrderBulkUpdateAPI(APIView):
    
    def create_po(self, data, po_to_create, po_lines_to_create):
        supplier_data = data.get("seller_details")
        client_data = data.get("buyer_details")
        po_lines = data.get("pieces_detail", [])
        try:
            storerkey_object = StorerKey.objects.get(storerkey_code=data.get("storerkey"))
        except StorerKey.DoesNotExist:
            return StandardResponse(success=False, errors=["Storer Key not found"], status=400)

        try:
            supplier_object = Supplier.objects.get(supplier_code=supplier_data.get("seller_code"))
        except Supplier.DoesNotExist:
            return StandardResponse(success=False, errors=[f"Supplier with code '{supplier_data.get('seller_code')}' not found"], status=400)

        client_object = supplier_object.client
        buyer_code = client_data.get("buyer_code")
        
        if client_object.client_code != buyer_code:
            return StandardResponse(success=False, errors=["Seller Code and Buyer codes do not linked"], status=400)
        
        if storerkey_object.client.id != client_object.id:
            return StandardResponse(success=False, errors=["Storer Key is not linked to the client"], status=400)

        if not supplier_object.storerkeys.filter(id=storerkey_object.id).exists():
            return StandardResponse(success=False, errors=["Storer Key is not linked to the supplier"], status=400)
            
        if not po_lines:
            return StandardResponse(success=False, errors=["Purchase Order lines required"], status=400)
        
        customer_reference_number = data.get("customer_reference_number")
        if PurchaseOrder.objects.filter(customer_reference_number=customer_reference_number).exists():
            return StandardResponse(success=False, errors=[f"Purchase Order with customer reference number {customer_reference_number} already exists"], status=400)
        
        customer_reference_number = data.get("customer_reference_number")
        if PurchaseOrder.objects.filter(customer_reference_number=customer_reference_number).exists():
            return StandardResponse(success=False, errors=[f"Purchase Order with customer reference number {customer_reference_number} already exists"], status=400)
            
        po_line_ref_numbers = [i.get("reference_number") for i in po_lines]
        po_line_cust_ref_numbers = [i.get("customer_reference_number") for i in po_lines]
        
        existing_po_line_refs = set(
            PurchaseOrderLine.objects.filter(reference_number__in=po_line_ref_numbers)
            .values_list("reference_number", flat=True)
        )

        existing_po_line_cust_refs = set(
            PurchaseOrderLine.objects.filter(customer_reference_number__in=po_line_cust_ref_numbers)
            .values_list("customer_reference_number", flat=True)
        )
        
        errors = []
        if existing_po_line_refs:
            errors.append(f"Purchase Order Line(s) with reference number(s) {', '.join(existing_po_line_refs)} already exist.")

        if existing_po_line_cust_refs:
            errors.append(f"Purchase Order Line(s) with customer reference number(s) {', '.join(existing_po_line_cust_refs)} already exist.")

        if errors:
            return StandardResponse(success=False, errors=errors, status=400)
        
        po_id = uuid4()
        data.update({
            "id": po_id,
            "client": client_object,
            "supplier": supplier_object,
            "storerkey": storerkey_object
        })
        data.update(seller_details(supplier_data))
        
        del data["seller_details"]
        del data["pieces_detail"]
                    
        po_to_create.append(PurchaseOrder(**data))
        for po_line in po_lines:
            po_line["purchase_order_id"] = po_id
            po_line["open_quantity"] = po_line.get("quantity")               
            po_lines_to_create.append(PurchaseOrderLine(**po_line))
    
    @role_required(Role.ADMIN)    
    @transaction.atomic
    def put(self, request, *args, **kwargs):
        closed_po = []
        cannot_update_sku = []
        quantity_not_matched = []
        po_to_create = []
        po_lines_to_create = []
        po_to_update = []
        po_lines_to_update = []
        data = request.data.get("data")
        existing_pos = PurchaseOrder.objects.filter(customer_reference_number__in=[i.get("customer_reference_number") for i in data])
        existing_all_po_lines = PurchaseOrderLine.objects.filter(purchase_order_id__in=existing_pos.values_list("id", flat=True))
        
        for po_data in data:
            customer_reference_number = str(po_data.get("customer_reference_number"))
            existing_po = existing_pos.filter(customer_reference_number=customer_reference_number).first()
            if existing_po:
                if existing_po.status == PurchaseOrderStatusChoices.CLOSED:
                    closed_po.append(customer_reference_number)
                    continue
                
                if "client" in po_data:
                    del po_data["client"]
                if "supplier" in po_data:
                    del po_data["supplier"]
                if "storerkey" in po_data:
                    del po_data["storerkey"]
                if "reference_number" in po_data:
                    del po_data["reference_number"]
                if "customer_reference_number" in po_data:
                    del po_data["customer_reference_number"]
                    
                for key,value in po_data.items():
                    setattr(existing_po, key, value)
                    
                po_to_update.append(existing_po)
                    
                po_lines = po_data.get("pieces_detail", [])
                existing_po_lines = existing_all_po_lines.filter(purchase_order=existing_po)
                existing_po_lines_dict = {obj.customer_reference_number : obj for obj in existing_po_lines}  
                
                incoming_line_refs = set()  
                
                for index, line in enumerate(po_lines):
                    line_customer_reference_number = line.get("customer_reference_number")
                    if not line_customer_reference_number:
                        return StandardResponse(status=400, success=False, errors=[f"Customer reference number is required for line item at index {index}"])
                    
                    incoming_line_refs.add(line_customer_reference_number)
                    
                    if line_customer_reference_number in existing_po_lines_dict:
                        
                        obj = existing_po_lines_dict[line_customer_reference_number]
                        
                        incoming_sku = line.get("sku")
                        if incoming_sku and incoming_sku != obj.sku:
                            if obj.quantity != obj.open_quantity:
                                cannot_update_sku.append(line_customer_reference_number)
                                continue
                        
                        incoming_quantity = convert_to_decimal(line.get("quantity"),2)
                        
                        if incoming_quantity != convert_to_decimal(obj.quantity):
                            if incoming_quantity < (obj.processed_quantity + obj.fulfilled_quantity):
                                quantity_not_matched.append(line_customer_reference_number)
                                continue
                            
                        del line["reference_number"]
                        del line["customer_reference_number"]
                        
                        for key, value in line.items():
                            setattr(obj, key, value)
                        po_lines_to_update.append(obj)
                    else:
                        line["purchase_order"] = existing_po
                        line["open_quantity"] = line.get("quantity")
                        po_lines_to_create.append(PurchaseOrderLine(**line))  
                        
                
            else:
                error = self.create_po(po_data, po_to_create, po_lines_to_create)
                if error:
                    return error
                
            for cust_ref, obj in existing_po_lines_dict.items():
                if cust_ref not in incoming_line_refs:
                    obj.status = PurchaseOrderStatusChoices.CANCELLED 
                    po_lines_to_update.append(obj)  
                        
            
        for cust_ref, obj in existing_po_lines_dict.items():
            if cust_ref not in incoming_line_refs:
                obj.status = PurchaseOrderStatusChoices.CANCELLED 
                po_lines_to_update.append(obj) 
        
        errors = []

        if closed_po:
            errors.append(
                f"Purchase Orders with customer reference numbers [{', '.join(closed_po)}] are closed and cannot be updated."
            )

        if cannot_update_sku:
            errors.append(
                f"SKU cannot be updated for Purchase Order Lines with customer reference numbers [{', '.join(cannot_update_sku)}] because some quantity is already processed or shipped."
            )

        if quantity_not_matched:
            errors.append(
                f"Quantity cannot be reduced for Purchase Order Lines with customer reference numbers [{', '.join(quantity_not_matched)}] as it is less than the already processed or shipped quantity."
            )
                       
        try:                
            created_pos = PurchaseOrder.objects.bulk_create(po_to_create, batch_size=1000)
            PurchaseOrder.objects.bulk_update(po_to_update, allow_fields_to_update(PurchaseOrder), batch_size=1000)
            created_po_lines = PurchaseOrderLine.objects.bulk_create(po_lines_to_create, batch_size=1000)
            PurchaseOrderLine.objects.bulk_update(po_lines_to_update, allow_fields_to_update(PurchaseOrderLine), batch_size=1000)

            for obj in created_pos+created_po_lines+po_to_update+po_lines_to_update:
                obj.update_status()
                
            PurchaseOrderService.update_open_quantity(created_pos)

            # for po in created_pos:
            #     po.update_quantity()

            POLineService.update_line_quantities(created_po_lines + po_lines_to_update)

            return StandardResponse(
                success=True,
                message="Purchase Order updated successfully",
                errors=errors,
                status=201
            )
        
        except IntegrityError as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, message="Database error.", errors=[str(e)])
    
        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(success=False, errors=[str(e)], status=400)


class PurchaseOrderUploadHistoryAPI(PaginationMixin,APIView):
    @role_required(Role.ADMIN)
    def get(self,request,my_uploads=None):
        try:
            pg = request.GET.get("pg") or 0
            limit = request.GET.get("limit") or 25
            my_uploads = str(request.GET.get("show_my_uploads")).lower() == "true"
            
            fields = ["status","uploaded_file","error_file","created_at","name","document_type"]

            queryset = PurchaseOrderUpload.objects.select_related("uploaded_by").annotate(name=F("uploaded_by__name"),document_type = Value("Bulk PO Upload"))
            
            if my_uploads == True:
                user = request.this_user
                queryset = queryset.filter(uploaded_by = user)
  
            queryset = queryset.values(*fields).order_by("-updated_at")
            count = queryset.count()
            paginate_result = self.paginate_results(queryset, pg, limit)

            for row in paginate_result:

                if row["uploaded_file"]:
                    row["uploaded_file"] = settings.MEDIA_URL + row["uploaded_file"]

                if row["error_file"]:
                    row["error_file"] = settings.MEDIA_URL + row["error_file"]

            return StandardResponse(status=200,data=paginate_result,count=count)
        except Exception as e:
            return StandardResponse(status=400,success=False,message=e)
    


class PurchaseOrderHeaderLineView(APIView):
     
    def seller_details(self, seller_data):
        data_to_sent = {}
        fields = ["address_line_1", "address_line_2", "city", "state", "country", "postal_code", "phone_number", "tax_number", "email"]
        for f in fields:
            data_to_sent["seller_" + f] = seller_data.get(f, "") 
        return data_to_sent
    
    def allow_fields(self):
        exclude_fields = ['created_at', 'updated_at', 'is_active', 'is_deleted', 'deleted_at', 'reference_number', 'customer_reference_number']
        fields = [field.name for field in PurchaseOrderLine._meta.get_fields() if field.concrete and not field.is_relation and not field.primary_key and field.name not in exclude_fields]   
        return fields
    
    @role_required(Role.SUPPLIER_USER)
    @transaction.atomic
    def put(self, request, *args, **kwargs):
        po_data = request.data.copy()
        purchase_order_id = request.data.get("purchase_order")
        supplier_data = po_data.get("seller_details")
        po_lines_data = po_data.get("pieces_detail",[])

        if not purchase_order_id:
            return StandardResponse(status=400, success=False, errors=["Purchase order ID is required"])
        try:
            purchase_order = PurchaseOrder.objects.get(customer_reference_number=purchase_order_id)
        except (PurchaseOrder.DoesNotExist, ValidationError):
                return StandardResponse(status=400, success=False, errors=["Purchase Order not found"])

        if "client" in po_data:
            del po_data["client"]
        if "supplier" in po_data:
            del po_data["supplier"]
        if "storerkey" in po_data:
            del po_data["storerkey"]
        if "reference_number" in po_data:
            del po_data["reference_number"]
        if "customer_reference_number" in po_data:
            del po_data["customer_reference_number"]


        try:
            # Po Header Update
            if supplier_data:
                po_data.update(self.seller_details(supplier_data))
                del po_data["seller_details"]
            for key,value in po_data.items():
                setattr(purchase_order, key, value)
            purchase_order.save()
            
            # Po Lines Update
            existing_po_lines = PurchaseOrderLine.objects.filter(purchase_order=purchase_order)
            existing_po_lines_dict = {obj.customer_reference_number : obj for obj in existing_po_lines}
            po_lines_to_update = []
            po_lines_to_create = []

            for index,line in enumerate(po_lines_data):
                line_customer_reference_number = line.get("customer_reference_number")
                if not line_customer_reference_number:
                    return StandardResponse(status=400, success=False, errors=[f"Customer reference number is required for line item at index {index}"])
                if line_customer_reference_number and line_customer_reference_number in existing_po_lines_dict:
                    del line["reference_number"]
                    obj = existing_po_lines_dict[line_customer_reference_number]
                    for key, value in line.items():
                        setattr(obj, key, value)
                    po_lines_to_update.append(obj)
                else:
                    line["purchase_order"] = purchase_order
                    line["open_quantity"] =  line.get("quantity") 
                    line_create = PurchaseOrderLine(**line)
                    po_lines_to_create.append(line_create)

            if len(po_lines_to_update) > 0:
                PurchaseOrderLine.objects.bulk_update(po_lines_to_update,self.allow_fields())

            if len(po_lines_to_create) > 0:
                PurchaseOrderLine.objects.bulk_create(po_lines_to_create)

                
            purchase_order.update_quantity()   
        except (IntegrityError) as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, message="Database error.", errors=[str(e)])

        except (Exception) as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=[str(e)])
        
        return StandardResponse(success=True, message="Purchase Order Header And Line(s) updated successfully", status=201)
            
class SupplierPurchaseOrders(APIView,PaginationMixin):

    def get(self, request, *args, **kwargs):
        supplier_code = request.GET.get("supplier_code")
        q = request.GET.get("q","").strip()
        if not supplier_code:
            return StandardResponse(data=[], status=200)

        pg = request.GET.get("pg", 0)
        limit = request.GET.get("limit", 50)
        created = request.GET.get("created", "false").lower() == "true"

        purchase_orders = (
            PurchaseOrder.objects
            .filter(
                supplier__supplier_code=supplier_code,
            )
            .exclude(
                status__in=[PurchaseOrderStatusChoices.CANCELLED, PurchaseOrderStatusChoices.CLOSED]
            )
            .annotate(
                total_open_qty=Sum('lines__open_quantity'),
                lines_count=Count('lines', filter=Q(lines__open_quantity__gt=0))
            )
            .filter(
                # total_open_qty__gt=0,  # Only include POs having at least one open qty
                customer_reference_number__icontains=q
            )
            .values("id", "reference_number", "customer_reference_number", "lines_count")
        )

        if not created:
            purchase_orders = purchase_orders.filter(total_open_qty__gt=0) # Only include POs having at least one open qty


        count = purchase_orders.count()
        paginated_pos = self.paginate_results(purchase_orders, pg, limit)

        return StandardResponse(data=paginated_pos, status=200,count =count)
    


class StorerKeyPurchaseOrders(APIView,PaginationMixin):

    def get(self, request, *args, **kwargs):
        storerkey_code = request.GET.get("storerkey_code")
        supplier_code = request.GET.get("supplier_code")

        if not storerkey_code:
            return StandardResponse(success=False, status=400, errors=["Storerkey code required"])
        
        if not supplier_code:
            return StandardResponse(success=False, status=400, errors=["Supplier code required"])
        
        q = request.GET.get("q","").strip()
        pg = request.GET.get("pg", 0)
        limit = request.GET.get("limit", 50)
        created = request.GET.get("created", "false").lower() == "true"

        filters = Q(storerkey__storerkey_code=storerkey_code, supplier__supplier_code=supplier_code)
        if q:
            filters &= Q(customer_reference_number__icontains=q)

        purchase_orders = (
            PurchaseOrder.objects
            .filter(filters)
            .exclude(
                status__in=[PurchaseOrderStatusChoices.CANCELLED, PurchaseOrderStatusChoices.CLOSED]
            )
            .annotate(
                total_open_qty=Sum('lines__open_quantity'),
                lines_count=Count('lines', filter=Q(lines__open_quantity__gt=0))
            )
            # .filter(
            #     # total_open_qty__gt=0,  # Only include POs having at least one open qty
            #     customer_reference_number__icontains=q
            # )
            .values("id", "reference_number", "customer_reference_number", "lines_count")
        )

        if not created:
            purchase_orders = purchase_orders.filter(total_open_qty__gt=0) # Only include POs having at least one open qty


        count = purchase_orders.count()
        paginated_pos = self.paginate_results(purchase_orders, pg, limit)

        return StandardResponse(data=paginated_pos, status=200,count =count)
class PurchaseOrderHeaderView(FilterMixin, SearchAndFilterMixin, PaginationMixin, APIView):
    
    po_fields = ["id", "customer_reference_number","reference_number","description","open_quantity","type","order_due_date","expected_delivery_date","status"];    
    supplier_fields = ["id","supplier_code","name"]

    def seller_details(self, seller_data):
        data_to_sent = {}
        fields = ["address_line_1", "address_line_2", "city", "state", "country", "postal_code", "phone_number", "tax_number", "email"]
        for f in fields:
            data_to_sent["seller_" + f] = seller_data.get(f, "") 
        return data_to_sent
    

    def get(self, request, ref_no=None, *args, **kwargs):
        
        filters = {}
        filters = self.build_filter(user=request.this_user,model="PO",filters = filters)

        if ref_no != "list":
            try:
                filters["customer_reference_number"] = ref_no
                obj = PurchaseOrder.objects.get(**filters)
                obj.update_quantity()
            except (PurchaseOrder.DoesNotExist, ValidationError):
                return StandardResponse(status=400, success=False, errors=["Object not found"])
            return StandardResponse(status=200, data=GetPurchaseOrderSerializer(obj).data)
        
        
        search_fields = ["supplier__name"]

            
        pg = int(request.GET.get("pg", 0))
        limit = int(request.GET.get("limit", 25))
        search = request.GET.get("q", "").strip()
        # fields = get_all_fields(PurchaseOrder, ignore_fields=[], include_relational_fields=False)
        # fields.extend(search_fields)
        search_fields.extend(self.po_fields)
        search_fields = set(search_fields)
        # filters = self.build_filter(user=request.this_user,model="PO",filters = filters)
        queryset = PurchaseOrder.objects.filter(**filters).select_related("supplier").only(*self.po_fields,"supplier").order_by("-created_at")
        
        if search:
            queryset = self.apply_search(search_fields, queryset, search)
        
        filters = self.make_filters_list(request)
        if filters:
            apply_filters = self.appy_dynamic_filter(filters)  
            queryset = queryset.filter(apply_filters)
            
        count = queryset.count()
        paginate_result = self.paginate_results(queryset, pg, limit)

        data = []
        for result in paginate_result:
            item = {field: getattr(result, field, None) for field in self.po_fields}
            supplier = {field: getattr(result.supplier, field, None) for field in self.supplier_fields}
            item["supplier"] = supplier
            data.append(item)

        return StandardResponse(
            success=True,
            # data=GetPurchaseOrderSerializer(paginate_result, many=True).data,
            data=data,
            count=count,
            status=200
        )  


    @role_required(Role.ADMIN)
    @transaction.atomic
    def put(self, request, ref_no=None, *args, **kwargs):
        
        allowed_fields_to_update = [
            "description","destination_country","type","order_due_date","expected_delivery_date","origin_country", 
            "payment_terms" ,"status", "group_code", "client", "storerkey", "supplier", "inco_terms", "center_code", "plant_id","notes"
        ]
        
        po_data = request.data.copy()
        # supplier_data = po_data.get("seller_details")

        po_object = PurchaseOrder.objects.select_for_update().prefetch_related(
            "lines"
        ).filter(customer_reference_number=ref_no)
        
        if not po_object.exists():
            return StandardResponse(status=400, success=False, errors=["Purchase Order not found"])
        
        purchase_order = po_object.first()
        lines = purchase_order.lines.select_for_update().all()

        if (
            ConsignmentPOLine.objects.filter(
                purchase_order_line_id__in=lines.values_list("id", flat=True),
                consignment__consignment_status = ConsignmentStatusChoices.DRAFT
            ).exists()
        ):
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=["Lines for this PO is currently being processed in a consignment."])

    

        try:
            # if supplier_data:
            #     po_data.update(self.seller_details(supplier_data))
            #     del po_data["seller_details"]'
            
            data_to_update = {}
            for key in po_data.keys():
                if key in allowed_fields_to_update:
                    data_to_update[key] = po_data[key]

            po_object.update(**data_to_update)
        
            NotificationService.po_update(purchase_order, request.this_user)

        except (IntegrityError) as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, message="Database error.", errors=[str(e)])

        except (Exception) as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=[str(e)])
        
        return StandardResponse(success=True, message="Purchase Order updated successfully", status=200)
            
    
class PurchaseOrderLineView(SearchAndFilterMixin, FilterMixin, PaginationMixin,PurchaseOrderMixin, APIView):
    
    
    line_fields = [
        "id",
        "created_at",
        "updated_at",
        "is_active",
        "is_deleted",
        "deleted_at",
        "reference_number",
        "customer_reference_number",
        "notes",
        "alternate_unit",
        "order_due_date",
        "expected_delivery_date",
        "stock_number",
        "is_chemical",
        "is_dangerous_good",
        "product_code",
        "description",
        "quantity",
        "sku",
        "unit_price",
        "unit_cost",
        "weight",
        "volume",
        "length",
        "width",
        "height",
        "source_location",
        "hs_code",
        "initial_promise_date",
        "new_promise_date",
        "open_quantity",
        "fulfilled_quantity",
        "processed_quantity",
        "status",
        "inco_terms",
        "batch",
        "lot",
        "expiry_date",
        "manufacturing_date",
        "origin_country",
        "purchase_order"
    ]

    def po_line_quantity_validations(self, po_data, po_line_obj):

        """
        Validates and updates the quantity for a PO line.
        Automatically adjusts open_quantity to maintain consistency:
        quantity = delivered + processing + open

        Args:
            po_data (dict): Input data containing 'quantity'.
            po_line_obj (PurchaseOrderLine): Existing line with original values.

        Returns:
            tuple: (updated_po_data, error_message)
        """
        try:
            raw_quantity = po_data.get("quantity")
            if raw_quantity is None:
                return po_data, "'quantity' must be provided."
            
            """
            Uncomment this, while inconsistency coming in the quantities
            """
            
            # delivered_qty = (
            # PackagingAllocation.objects
            # .filter(
            #     purchase_order_line_id=po_line_obj.id,
            #     consignment_packaging__status__in=[
            #         PackageStatusChoices.DELIVERED,
            #         PackageStatusChoices.RECEIVED
            #     ]
            # )
            # .aggregate(total=Sum("allocated_qty"))
            # .get("total") or Decimal(0)
            # )

            # processing_qty = (
            #     PackagingAllocation.objects
            #     .filter(
            #         purchase_order_line_id=po_line_obj.id,
            #         consignment_packaging__status=PackageStatusChoices.NOT_RECEIVED
            #     )
            #     .aggregate(total=Sum("allocated_qty"))
            #     .get("total") or Decimal(0)
            # )


            # Convert all to Decimal for accuracy
            updated_quantity = Decimal(str(raw_quantity))
            delivered_qty = Decimal(po_line_obj.fulfilled_quantity)
            processing_qty = Decimal(po_line_obj.processed_quantity)
            total_allocated = delivered_qty + processing_qty

            # Rule: original_quantity must be >= delivered + processing
            if updated_quantity < total_allocated:
                return po_data, (
                    "Quantity must be greater than or equal to delivered + processing quantity "
                    f"({total_allocated})."
                )

            # Calculate new open_quantity
            calculated_open_quantity = updated_quantity - total_allocated

            # Rule: open_quantity must be >= 0
            if calculated_open_quantity < 0:
                return po_data, "Open quantity cannot be negative."

            # All good, update both in po_data
            po_data["quantity"] = updated_quantity
            po_data["open_quantity"] = calculated_open_quantity
            po_data["status"] = PurchaseOrderStatusChoices.OPEN if updated_quantity == calculated_open_quantity else PurchaseOrderStatusChoices.PARTIALLY_FULFILLED

            return po_data, ""

        except Exception as e:
            return po_data, f"Error during quantity validation: {str(e)}"
            
    def allow_fields(self):
        exclude_fields = ['created_at', 'updated_at', 'is_active', 'is_deleted', 'deleted_at', 'reference_number', 'customer_reference_number']
        fields = [field.name for field in PurchaseOrderLine._meta.get_fields() if field.concrete and not field.is_relation and not field.primary_key and field.name not in exclude_fields]   
        return fields
    
    def _queryset_to_custom_json(self, queryset):
        """
        Converts a queryset into a list of dictionaries with custom fields.
        Uses getattr on each object to populate the dict.
        """
        data = []

        for qs in queryset:
            # Build dict for each object (you can add/remove keys as needed)
            obj_dict = {
                "id": qs.id,
                "created_at": qs.created_at,
                "updated_at": qs.updated_at,
                "is_active": qs.is_active,
                "is_deleted": qs.is_deleted,
                "deleted_at": qs.deleted_at,
                "reference_number": qs.reference_number,
                "customer_reference_number": qs.customer_reference_number,
                "notes": qs.notes,
                "alternate_unit": qs.alternate_unit,
                "order_due_date": qs.order_due_date,
                "expected_delivery_date": qs.expected_delivery_date,
                "stock_number": qs.stock_number,
                "is_chemical": qs.is_chemical,
                "is_dangerous_good": qs.is_dangerous_good,
                "product_code": qs.product_code,
                "description": qs.description,
                "quantity": qs.quantity,
                "sku": qs.sku,
                "unit_price": qs.unit_price,
                "unit_cost": qs.unit_cost,
                "weight": qs.weight,
                "volume": qs.volume,
                "length": qs.length,
                "width": qs.width,
                "height": qs.height,
                "source_location": qs.source_location,
                "hs_code": qs.hs_code,
                "initial_promise_date": qs.initial_promise_date,
                "new_promise_date": qs.new_promise_date,
                "open_quantity": qs.open_quantity,
                "fulfilled_quantity": qs.fulfilled_quantity,
                "processed_quantity": qs.processed_quantity,
                "status": qs.status,
                "inco_terms": qs.inco_terms,
                "batch": qs.batch,
                "lot": qs.lot,
                "expiry_date": qs.expiry_date,
                "manufacturing_date": "",
                "origin_country": qs.origin_country,
                "manufacturing_country": "",
                "purchase_order": getattr(qs.purchase_order, "id", None),
                "consignment_count": getattr(qs, "consignment_count", 0),
            }

            data.append(obj_dict)

        return data


    def get(self, request, ref_no=None, *args, **kwargs):

        filters = {}
        filters = self.build_filter(user=request.this_user,model="PO-Line",filters = filters)
        if ref_no != "list":

            try:
                filters["customer_reference_number"] = ref_no
                obj = PurchaseOrderLine.objects.get(**filters)
            except (PurchaseOrderLine.DoesNotExist, ValidationError):
                return StandardResponse(status=400, success=False, errors=["Object not found"])
            return StandardResponse(status=200, data=PurchaseOrderLineSerializer(obj).data)

        # resp = None
        pg = request.GET.get("pg") or 0
        limit = request.GET.get("limit") or 25
        search = request.GET.get("q", "")
        status = request.GET.get("status","")
        created = (request.GET.get("created","") or "true") == "true"
        po_number = request.GET.get("po")

        filters["purchase_order__customer_reference_number"] = po_number
        if status and status != "all":
            filters["status"] = status 

        queryset =(
            PurchaseOrderLine.objects
            .filter(**filters)
            .only(*self.line_fields)
            .select_related("purchase_order")
            .annotate(
                consignment_count=Count(
                    'consignments_po_lines',
                    filter=~Q(consignments_po_lines__consignment__consignment_status=ConsignmentStatusChoices.DRAFT),
                    distinct=True
                )
            )
        )

        if not queryset:
            return StandardResponse(status=200,data=[], success=False, errors=["Purchase Order Lines not found"])
        
        purchase_order = queryset.first().purchase_order

        if not created:
            queryset = queryset.filter(open_quantity__gt = 0).exclude(status=PurchaseOrderStatusChoices.CANCELLED)
            count = queryset.count()
            paginated_result = self.paginate_results(queryset, pg, limit)
            serialized_data = self._queryset_to_custom_json(paginated_result)
            serialized_data = POLineService.add_additional_info(serialized_data, purchase_order)


            return StandardResponse(
                success=True,
                data=serialized_data,
                count=count,
                status=200
            )

        queryset = queryset.order_by("reference_number")
        
        if search:
            queryset = self.apply_search(self.line_fields, queryset, search)
        
        filters = self.make_filters_list(request)
        if filters:
            apply_filters = self.appy_dynamic_filter(filters)  
            queryset = queryset.filter(apply_filters)


        count = queryset.count()
        paginated_result = self.paginate_results(queryset, pg, limit)
        serialized_data = self._queryset_to_custom_json(paginated_result)
        
        serialized_data = POLineService.add_additional_info(serialized_data, purchase_order)
       
        return StandardResponse(
            success=True,
            data=serialized_data,
            count=count,
            status=200
        )  


    @role_required(Role.ADMIN,Role.SUPPLIER_USER)
    @transaction.atomic
    def put(self, request, ref_no=None, *args, **kwargs):

        allowed_field_to_updated = [
            "quantity","open_quantity","status", "expected_delivery_date", "order_due_date", "new_promise_date", "description", "unit_price","unit_cost",
            "weight","length", "width", "height","hs_code", "is_dangerous_good", "alternate_unit", "stock_number", "is_chemical"
        ]
        po_data = request.data.copy()
        
        try:
            # ------------ SINGLE LINE UPDATE (NON BULK) ------------
            if ref_no != "bulk-update":
                purchase_order_id = request.data.get("purchase_order")
                po_line_qs = (
                    PurchaseOrderLine.objects
                    .select_for_update()
                    .select_related("purchase_order")
                    .filter(customer_reference_number=ref_no, purchase_order_id=purchase_order_id)
                )                
                
                if not po_line_qs.exists():
                    return StandardResponse(status=400, success=False, errors=["Object not found"]) 
                
                po_line = po_line_qs.first()

                if (
                    ConsignmentPOLine.objects.filter(
                        purchase_order_line_id=po_line.id,
                        consignment__consignment_status = ConsignmentStatusChoices.DRAFT
                    ).exists()
                ):
                    transaction.set_rollback(True)
                    return StandardResponse(status=400, success=False, errors=["This PO line is currently being processed in a consignment."])
                
                po_data , message = self.po_line_quantity_validations(po_data,po_line)
                if message:
                    return StandardResponse(status=400, success=False, message=message)
                
                # data_to_update = {}
                # for key in po_data.keys():
                #     if key in allowed_field_to_updated:
                #         data_to_update[key] = po_data[key]

                update_data = {k: v for k, v in po_data.items() if k in allowed_field_to_updated}
                po_line_qs.update(**update_data)
                po_line.update_status()
                
                purchase_order = po_line.purchase_order
                purchase_order.update_quantity()
                purchase_order.update_status()

                NotificationService.po_update(purchase_order, request.this_user)

                return StandardResponse(success=True, message="Purchase Order Line updated successfully", data={"id": po_line.id}, status=201)
            
            lines = request.data.get("lines", [])
            purchase_order_id = request.data.get("purchase_order")
            if not purchase_order_id:
                return StandardResponse(status=400, success=False, errors=["Purchase order ID is required"])

            purchase_order = PurchaseOrder.objects.select_for_update().get(customer_reference_number=purchase_order_id)
                
            existing_po_lines = PurchaseOrderLine.objects.select_for_update().filter(purchase_order=purchase_order)

            if (
                ConsignmentPOLine.objects.filter(
                    purchase_order_line_id__in=existing_po_lines.values_list("id", flat=True),
                    consignment__consignment_status = ConsignmentStatusChoices.DRAFT
                ).exists()
            ):
                transaction.set_rollback(True)
                return StandardResponse(status=400, success=False, errors=["Lines for this PO is currently being processed in a consignment."])
            
            
            existing_po_lines_dict = {obj.customer_reference_number : obj for obj in existing_po_lines}

            po_lines_to_update = []
            po_lines_to_create = []

            for index,line in enumerate(lines):
                line_customer_reference_number = line.get("customer_reference_number")
                if not line_customer_reference_number:
                    return StandardResponse(status=400, success=False, errors=["Customer reference number is required for line item at index {index}"])
                if line_customer_reference_number and line_customer_reference_number in existing_po_lines_dict:
                    line.pop("purchase_order", None)
                    line.pop("reference_number", None)

                    obj = existing_po_lines_dict[line_customer_reference_number]
                    for key, value in line.items():
                        setattr(obj, key, value)
                    po_lines_to_update.append(obj)
                else:
                    line_create = PurchaseOrderLine(**line)
                    po_lines_to_create.append(line_create)

                if len(po_lines_to_update) > 0:
                    PurchaseOrderLine.objects.bulk_update(po_lines_to_update,self.allow_fields())

                if len(po_lines_to_create) > 0:
                    PurchaseOrderLine.objects.bulk_create(po_lines_to_create)

            return StandardResponse(status=201, success=True, message="Purchase Order Lines updated successfully.")

        except PurchaseOrder.DoesNotExist:
            return StandardResponse(status=400, success=False, errors=["Purchase order not found"])
        
        except (IntegrityError) as e:
            return StandardResponse(status=400, success=False, message="Database error.", errors=[str(e)])
        
        except (Exception) as e:
            return StandardResponse(status=400, success=False, errors=[str(e)])
                                
    # def put(self, request, ref_no=None, *args, **kwargs):
            
    #     if ref_no != "bulk-update":
    #         try:
    #             obj = PurchaseOrderLine.objects.get(reference_number=ref_no)
    #         except (PurchaseOrderLine.DoesNotExist, ValidationError):
    #             return StandardResponse(status=400, success=False, errors=["Object not found"]) 
                        

    #         serializer = PurchaseOrderLineSerializer(obj, data=request.data, partial=True)
            
    #         if not serializer.is_valid():
    #             return StandardResponse(status=400, success=False, errors=serializer.errors)
    #         obj = serializer.save()
    #         return StandardResponse(status=201, message="Purchase Order Line updated successfully.", data={"id":obj.id})
    #     else:
    #         lines = request.data.get("lines", [])
    #         purchase_order_id = request.data.get("purchase_order")

    #         try:
    #             purchase_order = PurchaseOrder.objects.get(reference_number=purchase_order_id)
    #         except:
    #             return StandardResponse(status=400, success=False, errors=["Purchase order not found"])
            
    #         if not purchase_order_id:
    #             return StandardResponse(status=400, success=False, errors=["Purchase order ID is required"])

    #         existing_po_lines = PurchaseOrderLine.objects.filter(purchase_order=purchase_order)
    #         existing_po_lines_dict = {obj.reference_number: obj for obj in existing_po_lines}
    #         try:
    #             with transaction.atomic():
    #                 for line in lines:
    #                     line_reference_number = line.get("reference_number")

    #                     if line_reference_number and line_reference_number in existing_po_lines_dict:
    #                         obj = existing_po_lines_dict[line_reference_number]
    #                         serializer = PurchaseOrderLineSerializer(obj, data=line, partial=True)
    #                         if not serializer.is_valid():
    #                             transaction.set_rollback(True)
    #                             return StandardResponse(status=400, success=False, errors=serializer.errors)                    
    #                     else:
    #                         line["purchase_order"] = purchase_order.id
    #                         serializer = PurchaseOrderLineSerializer(data=line)
    #                         if not serializer.is_valid():
    #                             transaction.set_rollback(True)
    #                             return StandardResponse(status=400, success=False, errors=serializer.errors)
    #                     serializer.save()

    #                 return StandardResponse(status=201, message="Purchase Order Lines updated successfully.")
            
    #         except IntegrityError as e:
    #             transaction.set_rollback(True)
    #             return StandardResponse(status=400, success=False, message="Database error.", errors=[str(e)])
            
    #         except Exception as e:  
    #             transaction.set_rollback(True)
    #             return StandardResponse(status=400, success=False, message="An unexpected error occurred.", errors=[str(e)])
    
                
    
    def delete(self, request, ref_no=None, *args, **kwargs):
        try:
            if ref_no != "bulk-delete":
                po_crn = request.query_params.get("po_crn")
                obj = PurchaseOrderLine.objects.get(customer_reference_number=ref_no, purchase_order__customer_reference_number = po_crn)
                if obj.fulfilled_quantity > 0 or obj.processed_quantity > 0:
                    return StandardResponse(success=False, errors=["This item is part of a consignment and cannot be deleted."], status=400)
                obj.status = PurchaseOrderStatusChoices.CANCELLED
                obj.save() ## dont use update here. We are capturing the logs using signals, update dont trigger the signals. 
                return StandardResponse(status=200, message="Purchase Order Line deleted successfully.")
                
            reference_numbers = request.data.get("reference_numbers", [])

            if not reference_numbers:
                return StandardResponse(status=400, success=False, errors=["No reference numbers provided for deletion."])

            po_lines = PurchaseOrderLine.objects.filter(reference_number__in=reference_numbers)
            if po_lines == 0:
                return StandardResponse(status=404, success=False, errors=["No matching Purchase Order Lines found."])
            po_lines.update(status = PurchaseOrderStatusChoices.CANCELLED)
        except (PurchaseOrderLine.DoesNotExist, ValidationError):
            return StandardResponse(status=400, success=False, errors=["Object not found"])
        
        except ProtectedError as e:
            related_models = {str(related._meta.verbose_name_plural).capitalize() for related in e.protected_objects}
            return StandardResponse(
                status=400,
                success=False,
                errors=[f"Cannot delete Purchase Order Line as it is protected by related {', '.join(related_models)}"],
            )

        return StandardResponse(status=200, message="Purchase Order Line(s) deleted successfully.")
# class PurchaseorderConsignmentsAPI(APIView,PurchaseOrderMixin):
#     def get(self, request, poId=None, *args, **kwargs):

#         print("into new api")
#         resp , message = self.get_all_po_related_consignments(poId)
#         if message:
#             return StandardResponse(status=404, success=False, errors=message)
#         return StandardResponse(status=200, data=resp)
class PurchaseOrderBulkUploadAPI(APIView,PurchaseOrderMixin):
    
    po_header_fields = ["reference_number", "customer_reference_number", "supplier_code", "storer_key", "inco_terms", "quantity", "order_due_date", "expected_delivery_date", "plant_id", "center_code", "description"]
    
    po_lines_fields = ["purchase_order_crn", "reference_number", "customer_reference_number", "sku", "quantity", "order_due_date", "expected_delivery_date", "description", "product_code", "hs_code"]

    @role_required(Role.ADMIN)
    def get(self, request, id=None, *args, **kwargs):
        
        sheet_data = {
            "PO Header": {
                "fields": {
                    "REFERENCE NUMBER": {"type": "text"},
                    "CUSTOMER REFERENCE NUMBER": {"type": "text"},
                    "SUPPLIER CODE": {"type": "list", "choices": list(Supplier.objects.filter().values_list("supplier_code", flat=True))},
                    "STORER KEY": {"type": "list", "choies": list(StorerKey.objects.filter().values_list("storerkey_code", flat=True))},
                    "INCO TERMS": {"type": "text"},
                    "QUANTITY": {"type": "text"},
                    "BUYER CODE": {"type": "list", "choices": list(Client.objects.filter().values_list("client_code", flat=True))},
                    "PLANT ID": {"type": "text"},
                    "CENTER CODE": {"type": "text"},
                    "DESCRIPTION": {"type": "text"},
                    "GROUP CODE": {"type": "text"},
                    "TYPE": {"type": "text"},
                    "NOTES": {"type": "text"},
                    "PAYMENT TERMS": {"type": "text"},
                    "DESTINATION COUNTRY": {"type": "text"},
                    "ORIGIN COUNTRY": {"type": "text"},
                    "ORDER DATE": {"type": "text"},
                    "EXPECTED DELIVERY DATE": {"type": "text"},
                    "ORDER DUE DATE": {"type": "text"},
                }
            },
            "PO Lines": {
                "fields": {
                    "PURCHASE ORDER CRN": {"type": "text"},
                    "REFERENCE NUMBER": {"type": "text"},
                    "CUSTOMER REFERENCE NUMBER": {"type": "text"},
                    "SKU": {"type": "text"},
                    "QUANTITY": {"type": "text"},
                    "DESCRIPTION": {"type": "text"},
                    "PRODUCT CODE": {"type": "text"},
                    "HS CODE": {"type": "text"},
                    "ALTERNATE UNIT": {"type": "text"},
                    "STOCK NUMBER": {"type": "text"},
                    "CHEMICAL": {"type": "text"},
                    "DANGEROUS GOOD": {"type": "text"},
                    "UNIT PRICE": {"type": "text"},
                    "UNIT COST": {"type": "text"},
                    "WEIGHT": {"type": "text"},
                    "VOLUME": {"type": "text"},
                    "LENGTH": {"type": "text"},
                    "WIDTH": {"type": "text"},
                    "HEIGHT": {"type": "text"},
                    "SOURCE LOCATION": {"type": "text"},
                    "NOTES": {"type": "text"},
                    "EXPECTED DELIVERY DATE": {"type": "text"},
                    "ORDER DUE DATE": {"type": "text"},

                }
            }
        }
        # sample_dir = "media/po/sample_file"
        # empty_directory(str(sample_dir))
        # filename = f"{sample_dir}/Purchase_Order.xlsx"
        # excel_service = ExcelService()
        # filename = excel_service.download_formatted_file(filename, fields={}, sheets_data=sheet_data)

        
        sample_dir = os.path.join(settings.MEDIA_ROOT, "po/sample_file")
        empty_directory(sample_dir)
        filename = os.path.join(sample_dir, "Purchase_Order.xlsx")
        excel_service = ExcelService()
        filename = excel_service.download_formatted_file(filename, fields={}, sheets_data=sheet_data)
        
        relative_path = os.path.relpath(filename, settings.MEDIA_ROOT)
        file_url = f"{settings.MEDIA_URL}{relative_path.replace(os.sep, '/')}"  # ensure forward slashes
        return StandardResponse(status=200,data=file_url)

        # return StandardResponse(status=200, data=filename)


    def check_validations(self, wb):
        po_header_sheet = wb["PO Header"]
        poh_headers = [str(cell.value).strip().lower().replace(" ", "_") for cell in next(po_header_sheet.iter_rows(min_row=1, max_row=1))]
        missing_poh_headers = [field for field in self.po_header_fields if field not in poh_headers]

        po_lines_sheet = wb["PO Lines"]
        pol_headers = [str(cell.value).strip().lower().replace(" ", "_") for cell in next(po_lines_sheet.iter_rows(min_row=1, max_row=1))]
        missing_pol_headers = [field for field in self.po_lines_fields if field not in pol_headers]

        header_errors = []
        if missing_poh_headers:
            header_errors.append(f"Missing in 'PO Header': {', '.join(missing_poh_headers)}")
        if missing_pol_headers:
            header_errors.append(f"Missing in 'PO Lines': {', '.join(missing_pol_headers)}")

        customer_ref_idx = poh_headers.index("customer_reference_number")
        po_row_two = next(po_header_sheet.iter_rows(min_row=2, max_row=2))
        po_customer_ref_value = po_row_two[customer_ref_idx].value
        has_po_customer_ref = bool(po_customer_ref_value and str(po_customer_ref_value).strip())

        customer_ref_idx_line = pol_headers.index("customer_reference_number")
        po_line_row_two = next(po_lines_sheet.iter_rows(min_row=2, max_row=2))
        po_line_customer_ref_value = po_line_row_two[customer_ref_idx_line].value
        has_pol_customer_ref = bool(po_line_customer_ref_value and str(po_line_customer_ref_value).strip())

        if not has_po_customer_ref:
            header_errors.append("Missing value for customer reference number in 'PO Header'")
        if not has_pol_customer_ref:
            header_errors.append(f"Missing value for customer reference number in 'PO Lines'")

        if header_errors:
            return True, header_errors

        return False, None
    
    
    @role_required(Role.ADMIN)
    @transaction.atomic
    def post(self, request, id=None):
        
        try:
            if not id:
                return StandardResponse(errors=["Id required"], status=400, success=False)
            
            file = request.FILES.get("file")

            ## Bulk Upload for SLB
            if id == POImportFormatsChoices.SLB:

                POImportValidationService.file_validations(file,format=POImportFormatsChoices.SLB)
                
                # po_entries = SLBPOImportService.parse_excel_to_pos_standard_format(file)
                # if not po_entries:
                #     return StandardResponse(success=False, status=400, errors=["File is empty"])
                
                # validation_errors = POImportValidationService.validate_slb_pos(po_entries)
                # if validation_errors:
                #     return StandardResponse(errors=validation_errors, status=400, success=False)

            

            ## Bulk Upload for Standard Format
            elif id == POImportFormatsChoices.PO:

                POImportValidationService.file_validations(file,format=POImportFormatsChoices.PO)
                
                wb = load_workbook(file, data_only=True)
                po_entries = self.parse_excel_to_pos(wb)

                if not po_entries:
                    return StandardResponse(success=False, status=400, errors=["File is empty"])
                validation_errors = self.validate_po_entries(po_entries)
        
                # validation_error, validation_msg = self.check_validations(wb)
                
                if validation_errors:
                    return StandardResponse(success=False, status=400, errors=validation_errors)
                
            else:
                return StandardResponse(success=False, status=400, errors=["Invalid Id"])
            
            obj = PurchaseOrderUpload.objects.create(
                uploaded_file=file,
                uploaded_by_id=request.this_user.id,
                file_format = id
            )

            NotificationService.notify_po_file_upload(
                user=request.this_user,
                header="Bulk PO Upload",
                message="File Uploaded Successfully",
                po_upload = obj
            )

            # SLBPOImportService.process_slb_po_file(obj)
            # process_purchase_orders.delay(obj)
            # process_purchase_orders.apply_async(args=[obj.id], countdown=30)
            return StandardResponse(status=200, message="Data uploaded successfully. Processing in background.")    
        

        except InvalidFileException:
            transaction.set_rollback(True)
            return StandardResponse(errors=["Uploaded file is not a valid Excel file"], status=400, success=False)
        except ServiceError as e:
            transaction.set_rollback(True)
            return StandardResponse(errors=[e.error],success=e.success,status=e.status)
        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(errors=[f"Could not read Excel file: {str(e)}"], status=400, success=False)
        
             
             
                
    # def check_validations(self, data):
    #     required_fields = ["Order Due Date", "Expected Delivery Date", "Purchase Order Number", "Buyer Code", "Supplier Code", "Storer Key", "Order Quantity", "PO Line Number"]
    #     existing_pos = set(PurchaseOrder.objects.values_list("customer_reference_number", flat=True))
    #     existing_po_line = set(PurchaseOrderLine.objects.values_list("customer_reference_number", "purchase_order__customer_reference_number"))
    #     duplicate_po_line = set()

    #     filtered_data = []  # New list to store valid rows

    #     for index, item in enumerate(data):
    #         missing_fields = [field for field in required_fields if not item.get(field)]
    #         if missing_fields:
    #             return f"{', '.join(missing_fields)} can't be empty"

    #         purchase_order = str(item.get("Purchase Order Number"))
            
    #         if purchase_order in existing_pos:
    #             print(f"Purchase Order {purchase_order} already exists, removing row {index+2}.")
    #             continue  # Skip this row

    #         po_line_key = (item.get("PO Line Number"), purchase_order)

    #         if po_line_key in existing_po_line:
    #             print(f"Purchase Order Line already exists, removing row {index+2}.")
    #             continue  # Skip this row

    #         if po_line_key in duplicate_po_line:
    #             print(f"Duplicate Purchase Order Line found, removing row {index+2}.")
    #             continue  # Skip this row

    #         duplicate_po_line.add(po_line_key)
            
    #         order_quantity = convert_to_decimal(item.get("Order Quantity"), 2)
            
    #         if not order_quantity:
    #             return f"Order Quantity must be decimal, check row {index+2}"
            
    #         try:
    #             datetime.strptime(item.get("Order Due Date"), "%Y-%m-%d")
    #         except:
    #             return f"Invalid date format for Order Due Date, Please add date in YYYY-MM-DD, check row {index+2}"
            
    #         try:
    #             datetime.strptime(item.get("Expected Delivery Date"), "%Y-%m-%d")
    #         except:
    #             return f"Invalid date format for Expected Delivery Date, Please add date in YYYY-MM-DD, check row {index+2}"
            
    #         filtered_data.append(item)  # Add valid item to the new list

    #     return filtered_data  # Return cleaned data    
           
    # @transaction.atomic
    # def post(self, request):
    #     file = request.FILES.get("file")
    #     if not file:
    #         return StandardResponse(status=400, success=False, errors=["File Not Found"])

    #     file_name = file.name
    #     engine = "openpyxl" if file_name.endswith(".xlsx") else "xlrd" if file_name.endswith(".xls") else None
    #     if not engine:
    #         return StandardResponse(status=400, success=False, errors=["Unsupported file format"])

    #     df = pd.read_excel(file, engine=engine)
    #     data = json.loads(df.to_json(orient="records"))

    #     if not data:
    #         return StandardResponse(status=400, success=False, errors=["The Excel file you uploaded is empty. Please upload an Excel file with at least one row of data."])

    #     validation_error = self.check_validations(data)
    #     if validation_error:
    #         return StandardResponse(status=400, success=False, errors=[validation_error])

    #     po_ref_and_id = {}
    #     po_to_create = []
    #     po_lines_to_create = []

    #     clients = {client.client_code: client for client in Client.objects.filter(client_code__in=[item.get("Buyer Code") for item in data])}
    #     suppliers = {supplier.supplier_code: supplier for supplier in Supplier.objects.filter(supplier_code__in=[item.get("Supplier Code") for item in data])}
    #     storerkeys = {storerkey.storerkey_code: storerkey for storerkey in StorerKey.objects.filter(storerkey_code__in=[item.get("Storer Key") for item in data])}

    #     for item in data:
    #         po_number = item.get("Purchase Order Number")
    #         if po_number in po_ref_and_id:
    #             po_id = po_ref_and_id[po_number]
    #         else:
    #             po_id = uuid4()
    #             po_ref_and_id[po_number] = po_id
    #             po_data = {
    #                 "id": po_id,
    #                 "customer_reference_number": po_number,
    #                 "inco_terms": item.get("Header Incoterms"),
    #                 "client": clients[item.get("Buyer Code")],
    #                 "supplier": suppliers[item.get("Supplier Code")],
    #                 "storerkey": storerkeys[item.get("Storer Key")],
    #                 "open_quantity": item.get("Order Quantity"),
    #                 "seller_address_line_1": item.get("Supplier Address1"),
    #                 "seller_address_line_2": item.get("Supplier Address2"),
    #                 "seller_city": item.get("Supplier City"),
    #                 "seller_country": item.get("Supplier Country"),
    #                 "seller_postal_code": item.get("Supplier Postal Code"),
    #                 "origin_country": item.get("Supplier Country"),
    #                 "destination_country": item.get("Destination Country"),
    #                 "buyer_details": {
    #                     "city": item.get("Buyer City"),
    #                     "country": item.get("Destination Country"),
    #                 }
    #             }
    #             po_to_create.append(PurchaseOrder(**po_data))

    #         po_line_data = {
    #             "purchase_order_id": po_id,
    #             "customer_reference_number": item.get("PO Line Number"),
    #             "reference_number": item.get("PO Line Number"),
    #             "sku": item.get("SKU"),
    #             "quantity": item.get("Order Quantity"),
    #             "open_quantity": item.get("Order Quantity"),
    #             "description": item.get("Po Line Description"),
    #             "expected_delivery_date": datetime.strptime(item.get("Expected Delivery Date"), "%m/%d/%Y"),
    #             "order_due_date": datetime.strptime(item.get("Order Due Date"), "%m/%d/%Y %H:%M"),
    #         }
    #         po_lines_to_create.append(PurchaseOrderLine(**po_line_data))

    #     try:
    #         PurchaseOrder.objects.bulk_create(po_to_create, batch_size=1000)
    #         PurchaseOrderLine.objects.bulk_create(po_lines_to_create, batch_size=1000)
    #         return StandardResponse(status=200, message="Purchase Order Uploaded successfully.")
    #     except IntegrityError as e:
    #         transaction.set_rollback(True)
    #         return StandardResponse(status=400, success=False, message="Database error.", errors=[str(e)])
    #     except Exception as e:
    #         transaction.set_rollback(True)
    #         return StandardResponse(status=400, success=False, message="An unexpected error occurred.", errors=[str(e)])            
                
    
class ReceivePackages(APIView):

    @role_required(Role.CLIENT_USER,Role.SUPPLIER_USER)
    @transaction.atomic
    def get(self, request, *args, **kwargs):
        package_id = request.query_params.get("id")
        if not package_id:
            return StandardResponse(status=400, success=False, errors="Package ID is required")

        try:
            # Fetch package with all necessary relations
            con_packaging = (
                ConsignmentPackaging.objects
                .select_related(
                    "packaging_type",
                    "consignment__supplier",
                    "consignment__client",
                    "consignment__console__freight_forwarder"
                )
                .prefetch_related("allocations__purchase_order_line__purchase_order")
                .get(id=package_id)
            )

            if con_packaging.status == PackageStatusChoices.RECEIVED:
                return StandardResponse(status=400, success=False, errors="Package already received")

            consignment = con_packaging.consignment
            console = consignment.console
            ff = console.freight_forwarder if console else None

            # MOT types and modes
            mot_qs = ff.mot.all() if ff else []
            mot_types = list(mot_qs.values_list("mot_type", flat=True))
            modes = list(mot_qs.values_list("mode", flat=True))

            # Purchase orders
            allocations = con_packaging.allocations.select_related("purchase_order_line")
            po_numbers = allocations.values_list("purchase_order_line__purchase_order__customer_reference_number", flat=True).distinct()

            # Package dimension string
            pkg = con_packaging.packaging_type
            dimension_str = (
                f"{pkg.length} x {pkg.width} x {pkg.height} {dict(Millimeter='mm', Centimeter='cm', Inch='In', Foot='ft', Yard='yd').get(pkg.dimension_unit, '')}"
                if pkg else ""
            )

            # SKU items
            sku_items = [
                {
                    "sku": alloc.purchase_order_line.product_code,
                    "quantity": alloc.allocated_qty,
                    "description": alloc.purchase_order_line.description,
                    "type": (
                        "dangerous good" if alloc.purchase_order_line.is_dangerous_good
                        else "chemical" if alloc.purchase_order_line.is_chemical
                        else ""
                    )
                }
                for alloc in allocations
            ]

            # DG items
            po_line_ids = allocations.values_list("purchase_order_line_id", flat=True)
            dg_items = list(
                ConsignmentPOLine.objects.filter(
                    consignment=consignment,
                    purchase_order_line_id__in=po_line_ids,
                    dg_class__isnull=False, dg_category__isnull=False
                )
                .select_related("dg_class", "dg_category")
                .values(
                    dg_class_name=F("dg_class__name"),
                    dg_category_name=F("dg_category__name")
                )
            )
            address, _ = addresses_and_pickup(consignment.consignment_id)
            # Final response
            data = {
                "id": str(package_id),
                "address": address,
                "PO_number": list(po_numbers),
                "consignment_id": consignment.consignment_id,
                "console_id": console.console_id if console else "",
                "freight_forwarder": ff.name if ff else "",
                "MOT": modes,
                "service_type": mot_types,
                "date_of_pickup": consignment.actual_pickup_datetime,
                "package_weight": f"{con_packaging.weight} {con_packaging.weight_unit}" if con_packaging.weight else None,
                "package_type": pkg.package_type if pkg else "",
                "package_dimension": dimension_str,
                "sku_items": sku_items,
                "dg_items": dg_items
            }

            return StandardResponse(status=200, success=True, message="Package details retrieved", data=data)

        except ConsignmentPackaging.DoesNotExist:
            return StandardResponse(status=404, success=False, errors="Package not found")
        except Exception as e:
            return StandardResponse(status=500, success=False, errors=str(e))

    
    @role_required(Role.CLIENT_USER)
    @transaction.atomic
    def post(self, request, *args, **kwargs):

        package_id = request.data.get("id")
        date_and_time = request.data.get("date_time")
        time_zone = request.data.get("time_zone")
        if not package_id or not date_and_time or not time_zone:
            return StandardResponse(status=400, success=False, errors="Id, Date and Time and Time Zone required")

        try:
            con_packaging = ConsignmentPackaging.objects.select_related("consignment").get(id=package_id)

            if con_packaging.status == PackageStatusChoices.RECEIVED:
                return StandardResponse(status=400, success=False, errors="Package already received")

            new_status = self._get_updated_consignment_status(con_packaging)

            if not new_status:
                return StandardResponse(status=400, success=False, errors="Unable to determine consignment status")

            con_packaging.status = PackageStatusChoices.RECEIVED
            con_packaging.received_date_time = date_and_time
            con_packaging.time_zone = time_zone
            con_packaging.save(update_fields=["status","received_date_time","time_zone"]) 

            consignment = con_packaging.consignment
            consignment.consignment_status = new_status
            consignment.save(update_fields=["consignment_status"])

            consignment.update_console_status()
            if consignment.consignment_status == ConsignmentStatusChoices.RECEIVED_AT_DESTINATION:
                ConsignmentServices.notify_consignment_update(request.this_user, instance=consignment)

            return StandardResponse(status=200, message="Package received successfully")

        except ConsignmentPackaging.DoesNotExist:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors="Invalid id")
        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=str(e))

    def _get_updated_consignment_status(self, con_packaging):

        existing_statuses = list(
            ConsignmentPackaging.objects.filter(
                consignment_id=con_packaging.consignment_id
            ).exclude(id=con_packaging.id).values_list("status", flat=True)
        )

        existing_statuses.append(PackageStatusChoices.RECEIVED)

        if all(status == PackageStatusChoices.RECEIVED for status in existing_statuses):
            return ConsignmentStatusChoices.RECEIVED_AT_DESTINATION
        return ConsignmentStatusChoices.PARTIALLY_RECEIVED

class ConsignmentEditCheckAPI(APIView):
    def get(self,request,id=None):
        
        if not id :
            return StandardResponse(status=400, success=False, errors=["Id is required"])
        
        try:
            consignment = Consignment.objects.select_related("supplier").filter(consignment_id=id).first()
            
            if not consignment:
                return StandardResponse(status=400, success=False, errors=["Pickup not exists"])

            if consignment.consignment_status != ConsignmentStatusChoices.PENDING_FOR_APPROVAL:
                return StandardResponse(status=400, success=False, errors=[f"Update not allowed for this status {consignment.consignment_status}"])
            
            user = request.this_user
            
            if user.role == Role.CLIENT_USER:
                return StandardResponse(status=400, success=False, errors=["Client user dont have access to edit pickup"])
            
            if user.role == Role.ADMIN:
                return StandardResponse(status=200, message="User is able to edit")
            
            if user.role == Role.OPERATIONS:
                operation = user.profile()

                if operation.access_level in OperationUserRole.L3:
                    return StandardResponse(status=400, success=False, errors=["L3 user dont have access to edit pickup"])

                if operation.access_level in [OperationUserRole.L1 and OperationUserRole.L2]:
                    return StandardResponse(status=200, message="User is able to edit")
                
            supplier = user.profile()
            if supplier != consignment.supplier:
                return StandardResponse(status=400, success=False, errors=["Another supplier can't edit this pickup"])

            return StandardResponse(status=200, message="User is able to edit")
        
            # user_ids = consignment.values_list("user__id",flat=True)
            
            # if user.id in user_ids:
            #     return StandardResponse(status=200, message="User is able to edit")

            # return StandardResponse(status=403, success=False, errors=f"This consignment is already being edited by '{stage_consignment.first().user.name}'")


        except Exception as e:
            return StandardResponse(status=500, success=False, errors=str(e))
        
class PurchaseOrderLineCountAPI(APIView):
    def get(self, request, po_no=None, *args, **kwargs):

        all_statuses = [choice[0] for choice in PurchaseOrderStatusChoices.choices]
        
        status_counts = PurchaseOrderLine.objects.filter(purchase_order__customer_reference_number=po_no).values('status').annotate(count=Count('status'))
        status_dict = {status: 0 for status in all_statuses}
        for status in status_counts:
            status_dict[status['status']] = status['count']
        
        return StandardResponse(status=200, data=status_dict)
       
    
class ConsignmentStatusMonthlyCountAPI(APIView):
    def get(self, request, id=None, *args, **kwargs):
        now = timezone.now()

        year = request.GET.get("year", now.year)
        # if year is number only then convert it to int
        # else return error
        try:
            year = int(year) 
         
            monthly_data = (
                Consignment.objects
                .filter(is_completed=True,created_at__year=year)
                .annotate(month=TruncMonth('created_at'))  
                .values('month')
                .annotate(count=Count('id'))
                .order_by('month')
            )

            months = [f"{calendar.month_abbr[month]} {year}" for month in range(1, 13)]
            
            results = {}
            
            for month in range(1, 13):
                month_data = next((item for item in monthly_data if item['month'].month == month), None)
                count = month_data['count'] if month_data else 0
                results[months[month-1]] = count
        except ValueError:
            return StandardResponse(status=400, success=False, errors=["Input must be an year (number)."])
        
        return StandardResponse(status=200, data=results, count=len(results))

class ConsignmentStatusBetweenDaysCountAPI(SearchAndFilterMixin,PaginationMixin, APIView):

    
    
    def get(self, request, *args, **kwargs):
        fields = ["id", "consignment_id", "purchase_order__customer_reference_number", "supplier__name", "client__name", 
                  "packages", "actual_pickup_datetime", "requested_pickup_datetime", "consignment_status", "created_at",
                  "console__console_id", "freight_forwarder__name", "is_completed", "type", "adhoc__customer_reference_number",
                  "delivery_address__address_name", "consignor_address__address_name"] 

        now = timezone.now()    
        #after this day is inclusive
        before_this_day = request.GET.get("before_this_day", None)
        after_this_day = request.GET.get("after_this_day", None)

        pg = request.GET.get("pg") or 0
        limit = request.GET.get("limit") or 25
        search = request.GET.get("q", "")
        status = request.GET.get("status","")
        print(status, "::status")
        try:
            queryset = Consignment.objects.filter(is_completed=True).select_related("supplier", "client", "purchase_order", "adhoc", "console__console").values(*fields).order_by("-consignment_id")

            if after_this_day:
                # after_this_day = now - timedelta(days = int(after_this_day) or 0, hours = 23, minutes = 59, seconds = 59)
                after_this_day = now - timedelta(days=int(after_this_day) or 0, hours = 0, minutes = 0, seconds = 0)
                queryset = queryset.filter(created_at__date__lte=after_this_day)

            if before_this_day: 
                # before_this_day  = now - timedelta(days = int(before_this_day) or 0, hours = 23, minutes = 59, seconds = 59)
                before_this_day  = now - timedelta(days = int(before_this_day) or 0, hours = 0, minutes = 0, seconds = 0)

                print("before_this_day",before_this_day)
                queryset = queryset.filter(created_at__date__gt=before_this_day)

            if status and status != "all":
                queryset = queryset.filter(consignment_status=status)

            apply_filters = self.make_filters_list(request)
            
            if apply_filters:
                for f in apply_filters:
                    f['column'] = self.transform_fields.get(f['column'], f['column'])
                apply_filters = self.appy_dynamic_filter(apply_filters)  
                queryset = queryset.filter(apply_filters)

            if search:
                queryset = self.apply_search(fields, queryset, search)

            count = queryset.count()
            paginate_result = self.paginate_results(queryset, pg, limit)

            return StandardResponse(success=True, data=paginate_result, count=count, status=200)

        except Exception as e:
            return StandardResponse(status=400, success=False, errors=[str(e)])
        
        

            
 
class ConsignmentStatusDonutChartAPI(APIView):
    def get(self, request, id=None, *args, **kwargs):
        current_date = timezone.now()
        current_month = current_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        if current_date.month == 1:
            previous_month = current_date.replace(year=current_date.year - 1, month=12, day=1, hour=0, minute=0, second=0, microsecond=0) 
        else:
            previous_month = current_date.replace(month=current_date.month - 1, day=1, hour=0, minute=0, second=0, microsecond=0)

        # Fetch consignment status counts for the current month
        current_consignment_status = (
            Consignment.objects
            .filter(is_completed=True,created_at__gte=current_month)
            .annotate(month=TruncMonth('created_at'))
            .values('consignment_status')
            .annotate(count=Count('id'))
        )

        # Fetch consignment status counts for the previous month
        previous_consignment_status = (
            Consignment.objects
            .filter(is_completed=True,created_at__gte=previous_month, created_at__lt=current_month)
            .annotate(month=TruncMonth('created_at'))
            .values('consignment_status')
            .annotate(count=Count('id'))
        )

        # Create dictionaries for current and previous counts
        current_data_map = {
            item['consignment_status']: item['count'] for item in current_consignment_status
        }
        previous_data_map = {
            item['consignment_status']: item['count'] for item in previous_consignment_status
        }

        # Get all possible statuses from ConsignmentStatusChoices
        all_statuses = [status[0] for status in ConsignmentStatusChoices.choices]

        results = []

        # Iterate through all possible statuses
        for status in all_statuses:
            current_count = current_data_map.get(status, 0)  # Default to 0 if status not found
            previous_count = previous_data_map.get(status, 0)  # Default to 0 if status not found
            
            # Get the total count of consignments
            total_consignment_count  = Consignment.objects.filter(consignment_status = status).count()

            if previous_count == 0:
                change = current_count
                unit = 'unit'
                sign = '+' if current_count > 0 else '0'
            else:
                diff_percent = round(((current_count - previous_count) / previous_count) * 100, 2)
                change = abs(diff_percent)
                unit = 'percentage'
                if current_count > previous_count:
                    sign = '+'
                elif current_count < previous_count:
                    sign = '-'
                else:
                    sign = '0'

            results.append({
                'status': status,
                'count': current_count,
                'change': change,
                'unit': unit,
                'sign': sign,
                'total_consignment_count': total_consignment_count ,
            })

        return StandardResponse(status=200, data=results, count=len(results))



class ConsignmentStatusSummaryCountAPI(ConsignmentStatusSummary,APIView):

    def get_queryset_for_range(self, min_days=None, max_days=None):
        now = timezone.now()

        if min_days is not None and max_days is not None:
            start_date = now - timedelta(days=max_days)  
            end_date = now - timedelta(days=min_days)
            return Consignment.objects.filter(
                created_at__gte=start_date, 
                created_at__lt=end_date
            )

        elif max_days is not None:
            start_date = now - timedelta(days=max_days)
            return Consignment.objects.filter(created_at__gte=start_date)

        elif min_days is not None:
            end_date = now - timedelta(days=min_days)
            return Consignment.objects.filter(created_at__lt=end_date)

        return Consignment.objects.none()

    def get_status_summary_dict(self, queryset):
        all_statuses = dict(ConsignmentStatusChoices.choices)
        summary = {label: 0 for label in all_statuses.values()}

        results = queryset.values('consignment_status').annotate(count=Count('id'))

        for entry in results:
            label = all_statuses.get(entry['consignment_status'], entry['consignment_status'])
            summary[label] = entry['count']

        return summary

    def get(self, request, *args, **kwargs):
        ranges = [
            {"label": "<2", "max_days": 2, "min_days": None},  
            {"label": "2-7", "max_days": 7, "min_days": 2},    
            {"label": "7-15", "max_days": 15, "min_days": 7},  
            {"label": "15-30", "max_days": 30, "min_days": 15},
            {"label": "30-45", "max_days": 45, "min_days": 30},
            {"label": "45-60", "max_days": 60, "min_days": 45},
            {"label": "90-120", "max_days": 120, "min_days": 90},
            {"label": "120-150", "max_days": 150, "min_days": 120},
            {"label": "150-180", "max_days": 180, "min_days": 150},
            {"label": ">180", "max_days": None, "min_days": 180},
        ]

        summary = []

        for r in ranges:
            queryset = self.get_queryset_for_range(min_days=r.get("min_days"), max_days=r.get("max_days"))
            status_counts = self.get_status_summary_dict(queryset)
            data = {"day": r["label"]}
            data.update(status_counts)
            summary.append(data)

        return StandardResponse(status=200, data=summary, count=len(summary))

 
class PurchaseOrderStatusCountAPI(FilterMixin,APIView):
    def get(self, request, id=None, *args, **kwargs):
        all_statuses = [choice[0] for choice in PurchaseOrderStatusChoices.choices]

        filters = self.build_filter(request.this_user,"PO",{})
        
        status_counts = PurchaseOrder.objects.filter(**filters).values('status').annotate(count=Count('status'))
        status_dict = {status: 0 for status in all_statuses}
        for status in status_counts:
            status_dict[status['status']] = status['count']
        
        return StandardResponse(status=200, data=status_dict, count=len(status_dict))
       
class ConsignmentDashboardThisMonth(APIView):
    def get(self, request, *args, **kwargs):
        now = timezone.now()
        
        start_of_month = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

        results = Consignment.objects.filter(created_at__gte=start_of_month).values('consignment_status').annotate(count=Count('consignment_status'))
        status_dict = {status[0]: 0 for status in ConsignmentStatusChoices.choices}
        for status in results:
            status_dict[status['consignment_status']] = status['count']
        
        return StandardResponse(status=200, data=status_dict, count=len(status_dict))
    
class AWBAPI(APIView):

    parser_classes = [MultiPartParser, FormParser]

    @role_required(OperationUserRole.L1,OperationUserRole.L2,Role.SUPPLIER_USER,Role.CLIENT_USER)
    def get(self, request, id = None):

        ## id == consignment_id
        try:
            files = AWBFile.objects.filter(consignment__consignment_id = id)
        except AWBFile.DoesNotExist:
            return StandardResponse(status=400,success=False, errors=["Invalid Id"])
        
        if not files:
            return StandardResponse(
                status=200,
                data=[],
                message="No files found : Please upload the file or check your consignment Id"
            )
        
        file_urls = [{"id":file.id, "url":file.file.url} for file in files]
        return StandardResponse(status=200,data=file_urls) 

    @role_required(OperationUserRole.L1,OperationUserRole.L2)
    @transaction.atomic
    def post(self, request, id = None, *args, **kwargs):
        
        try:
            consignment = Consignment.objects.get(consignment_id = id)
        except Consignment.DoesNotExist:
            return StandardResponse(status=400,success=False, errors=["Invalid Id"])
        
        files = request.FILES.getlist("files")
        if not files:
            return StandardResponse(status=400,success=False, errors=["Files required"])
        
        # error = validate_file_size(files)
        # if error:
        #     return None, error
    
        try:
            bulk_upload = [
                AWBFile(file=f, consignment=consignment)
                for f in request.FILES.getlist("files")
            ]

            created_files = AWBFile.objects.bulk_create(bulk_upload)

            for file in created_files:
                awb_file_added_audit_trail(file)


            return StandardResponse(success=True, message ="File uploaded successfully", status=201)
        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400,success=False, errors=[e])
        

    @role_required(OperationUserRole.L1,OperationUserRole.L2)
    @transaction.atomic
    def delete(self, request, id = None, *args, **kwargs):
        
        try:

            files = AWBFile.objects.filter(id=id)
            for f in files:
                awb_file_deleted_audit_trail(f)

            files.delete()
            return StandardResponse(success=True, message ="File deleted successfully", status=200)
        
        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400,success=False, errors=[e])

