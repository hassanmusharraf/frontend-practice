from rest_framework.views import APIView
from django.core.exceptions import ValidationError
from django.http import JsonResponse
from core.response import StandardResponse, ServiceError
from django.db import transaction
from .mixins import PurchaseOrderLineQuantityMixin, ConsignmentMixin, AdhocPurchaseOrderLineMixin
from django.db.models.functions import JSONObject
from django.contrib.postgres.aggregates import JSONBAgg
from django.db.models import OuterRef, Subquery, F, JSONField, Count, Sum, Max, Q
from decimal import Decimal
from rest_framework.parsers import MultiPartParser, FormParser
from .models import (
    PurchaseOrder,
    PurchaseOrderLine,
    # ConsignmentDocumentAttachmentStaging,
    # ConsignmentStaging,
    # ConsignmentPOLineStaging,
    # ConsignmentPackagingStaging,
    # PackagingAllocationStaging,
    # ConsignmentDocumentStaging,
    ConsignmentDocumentAttachment,
    Consignment,
    ConsignmentPackaging,
    PackagingAllocation,
    ConsignmentDocument,
    ConsignmentAuditTrail,
    ConsignmentFFDocument,
    UserGridPreferences
)
from .serializers import (
    GetConsignmentSerializer,
    # GetConsignmentDocumentStagingSerializer,
    # ConsignmentStagingSerializer,
    # GetConsignmentStagingSerializer,
    AdhocConsignmentDocumentSerializer,
    ConsignmentSerializer,
    ConsignmentAdhocSerializer,
    UserGridPreferecesSerializer
)
from portal.serializers import PackagingTypeSerializer
from entities.models import Client, Supplier
from portal.mixins import SearchAndFilterMixin, PaginationMixin
from portal.choices import MeasurementTypeChoices, ConsignmentStatusChoices, Role, ConsignmentTypeChoices, ConsignmentDocumentTypeChoices, PackagingTypeChoices, NotificationChoices, OperationUserRole, PackageStatusChoices
from django.core.files.storage import default_storage
from django.utils import timezone
import pytz
from datetime import datetime
from adhoc.models import AdhocPurchaseOrder, AdhocPurchaseOrderLine
from portal.models import DropDownValues, PackagingType, Notification, UserNotification
from portal.constants import IMPERIAL_SYSTEM_DIMENSION_UNIT, IMPERIAL_SYSTEM_WEIGHT_UNIT, METRIC_SYSTEM_DIMENSION_UNIT, METRIC_SYSTEM_WEIGHT_UNIT
from portal.service import ExcelService
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.styles import Font, PatternFill
from uuid import uuid4 , UUID
from operations.mixins import ConsignmentStatusSummary, PurchaseOrderMixin, FilterMixin
from django.core.exceptions import ObjectDoesNotExist
from portal.utils import empty_directory
import os
from django.conf import settings
from core.decorators import role_required
from .services import ConsignmentWorkflowServices, ConsignmentStatusService, ConsignmentServices

def get_consignments_by_timezone(request, timezone_name):
    # Get the user's timezone
    user_timezone = pytz.timezone(timezone_name)
    
    # Get the current time in the user's timezone
    now = timezone.now().astimezone(user_timezone)
    
    # Filter consignments based on the pickup timezone and convert datetimes
    consignments = Consignment.objects.filter(pickup_timezone=timezone_name)
    for consignment in consignments:
        consignment.requested_pickup_datetime = consignment.requested_pickup_datetime.astimezone(user_timezone)
        if consignment.actual_pickup_datetime:
            consignment.actual_pickup_datetime = consignment.actual_pickup_datetime.astimezone(user_timezone)
    
    return consignments

def update_po_lines_quantities():
    PurchaseOrderLine.objects.all().update(
        fulfilled_quantity=0,
        processed_quantity=0,
        open_quantity=F("quantity")
    )

    return PurchaseOrderLine.objects.all().count() 
    

  
# class StageConsignmentCreateAPI(APIView):
    
#     @role_required(OperationUserRole.L1,OperationUserRole.L2,Role.SUPPLIER_USER)
#     @transaction.atomic
#     def post(self, request, *args, **kwargs):
#         po_cust_ref_no = request.data.get("purchase_order")
#         user = request.this_user.id
        
#         # user = request.this_user
#         try:
#             po = PurchaseOrder.objects.select_related("client", "supplier").get(customer_reference_number=po_cust_ref_no)
#         except (PurchaseOrder.DoesNotExist, ValidationError):
#             return StandardResponse(status=400, success=False, errors=["Pucrhase order does not exist"])
        
#         obj, created = ConsignmentStaging.objects.get_or_create(
#             purchase_order=po,
#             is_update=False,
#             defaults={"user_id": user}
#         )
#         client_fields = ["id","client_code","name","timezone","service_type","measurement_method"]
#         client_data = Client.objects.filter(id=po.client.id).values(*client_fields).first()
#         supplier = Supplier.objects.filter(id=po.supplier.id).select_related("client").first()

#         supplier_data = {
#             "id": supplier.id,
#             "is_active": supplier.is_active,
#             "supplier_code": supplier.supplier_code,
#             "name": supplier.name,
#             "address": supplier.address,
#             "client_id": supplier.client.id
#         }
        
#         if not created:
#             if str(obj.user_id) == str(user):
#                 po_lines = ConsignmentPOLineStaging.objects.filter(consignment=obj).values()
#                 return StandardResponse(status=200, data={"data":po_lines, "client":client_data, "supplier": supplier_data}, count=po_lines.count())
#             return StandardResponse(status=400, success=False, errors=["Purchase Order already picked for consignment"])

#         return StandardResponse(status=200, data={"data": [], "client":client_data, "supplier": supplier_data}, message="Saved Successfully")


# class StageConsignmentGetAPI(ConsignmentStagingMixin, APIView):
#     def get(self, request, id=None, *args, **kwargs):
#         user = request.this_user.id
#         po_cust_ref_no = request.GET.get("po")
#         is_update = request.GET.get("is_update") == "true"
#         existing_consignment_id = None
#         if is_update:
#             existing_consignment_id = request.GET.get("consignment_id")
#             if existing_consignment_id in ["", "undefined", "null", None]:
#                 return StandardResponse(status=400, success=False, errors=["consignment id is required"])
        
#         error, consignment, error_msg, _ = self.check_stage_po_consignment_exists(po_cust_ref_no, user, is_update, existing_consignment_id)
#         if error:
#             return StandardResponse(status=400, success=False, errors=[error_msg])
        
#         return StandardResponse(status=200, data=GetConsignmentStagingSerializer(consignment).data)
        
        
# class StageConsignmentPOLineAPI(ConsignmentStagingMixin, APIView):
    
#     def get(self, request, id=None, *args, **kwargs):
        
#         if id == "list":
#             po_cust_ref_no = request.GET.get("po")
#             user = request.this_user.id
#             is_update = request.GET.get("is_update") == "true"
#             existing_consignment_id = None
#             if is_update:
#                 existing_consignment_id = request.GET.get("consignment_id")
#                 if existing_consignment_id in ["", "undefined", "null", None]:
#                     return StandardResponse(status=400, success=False, errors=["consignment id is required"])
        
#             error, consignment, error_msg, _ = self.check_stage_po_consignment_exists(po_cust_ref_no, user, is_update, existing_consignment_id)
#             if error:
#                 return StandardResponse(status=400, success=False, errors=[error_msg])
                
#             po_lines = ConsignmentPOLineStaging.objects.filter(consignment=consignment).values().order_by("po_line_ref")
            
#             files = ConsignmentDocumentAttachmentStaging.objects.filter(document__consignment=consignment).annotate(
#                 file_document_type = F("document__document_type") 
#             ).values("file_document_type")

#             # file_count = ConsignmentDocumentAttachmentStaging.objects.filter(
#             #     document__consignment=consignment
#             # ).aggregate(file_count=Count("id"))["file_count"]
            
#             packaging_count = ConsignmentPackagingStaging.objects.filter(consignment=consignment).count()
            
#             return StandardResponse(status=200, data={"data": po_lines, "document_attached": files, "packages_count": packaging_count}, count=po_lines.count())

#         else:
#             packages = PackagingAllocationStaging.objects.filter(po_line=id).exclude(allocated_qty=0).select_related("consignment_packaging__packaging_type").annotate(
#                 package_id=F("consignment_packaging__package_id"),
#                 package_name=F("consignment_packaging__packaging_type__package_name"),
#                 package_type=F("consignment_packaging__packaging_type__package_type"),
#                 description=F("consignment_packaging__packaging_type__description"),
#                 is_stackable=F("consignment_packaging__packaging_type__is_stackable"),
#                 measurement_method=F("consignment_packaging__packaging_type__measurement_method"),
#                 length=F("consignment_packaging__packaging_type__length"),
#                 height=F("consignment_packaging__packaging_type__height"),
#                 width=F("consignment_packaging__packaging_type__width"),
#                 dimension_unit=F("consignment_packaging__packaging_type__dimension_unit"),
#                 weight=F("consignment_packaging__weight"),
#                 weight_unit=F("consignment_packaging__weight_unit"),
#             ).values("allocated_qty", "package_id", "package_name", "package_type", "description", "is_stackable", "measurement_method", "length", "height", "width", "dimension_unit", "weight", "weight_unit",)
            
#             return StandardResponse(status=200, data=packages)
            
#     def post(self, request, *args, **kwargs):
#         po_cust_ref_no = request.data.get("purchase_order")
#         user = request.this_user.id
#         is_update = request.data.get("is_update") or False
#         existing_consignment_id = None
#         if is_update:
#             existing_consignment_id = request.data.get("consignment_id")
#             if existing_consignment_id in ["", "undefined", "null", None]:
#                 return StandardResponse(status=400, success=False, errors=["consignment id is required"])
        
#         error, consignment, error_msg, po = self.check_stage_po_consignment_exists(po_cust_ref_no, user, is_update, existing_consignment_id)
#         if error:
#             return StandardResponse(status=400, success=False, errors=[error_msg])
        
#         lines = request.data.get("lines", [])
#         data_to_create = [
#             ConsignmentPOLineStaging(
#                 consignment=consignment,
#                 purchase_order_line_id=line.get("id"),
#                 po_ref=po.customer_reference_number,
#                 po_line_ref=line.get("customer_reference_number"),
#                 qty_to_fulfill=line.get("confirm_quantity", 0),
#                 qty_packed=0,
#                 qty_remaining=line.get("confirm_quantity", 0),
#                 packages=[],
#                 hs_code=line.get("hs_code"),
#                 sku=line.get("sku"),
#                 is_dangerous_good=line.get("is_dangerous_good"),
#                 manufacturing_country=line.get("manufacturing_country", {}).get("label", None),
#                 eccn=line.get("eccn") or False
#             ) for line in lines
#         ]
        
#         try:
#             if data_to_create:
#                 existing_lines = ConsignmentPOLineStaging.objects.filter(consignment=consignment)
#                 if existing_lines.exists():
#                     existing_lines.delete()
#                 existing_packages = ConsignmentPackagingStaging.objects.filter(consignment=consignment)
#                 if existing_packages.exists():
#                     existing_packages.delete()
#                 documents = ConsignmentDocumentStaging.objects.filter(consignment = consignment)
#                 if documents.exists():
#                     documents.delete()
                
#                 ConsignmentPOLineStaging.objects.bulk_create(data_to_create)
#             return StandardResponse(status=201, success=True, message="Consignment PO Lines created successfully", data=ConsignmentPOLineStaging.objects.filter(consignment=consignment).values())
#         except Exception as e:
#             return StandardResponse(status=400, success=False, errors=[str(e)])
    
# class StageConsignmentPackagingPOLineAPI(ConsignmentStagingMixin, APIView):

#     def get(self, request, id=None, *args, **kwargs):

#         po_cust_ref_no = request.GET.get("po")
#         po_line_id = request.GET.get("po_line")
#         user = request.this_user.id
#         is_update = request.GET.get("is_update") == "true"
#         existing_consignment_id = None
#         if is_update:
#             existing_consignment_id = request.GET.get("consignment_id")
#             if existing_consignment_id in ["", "undefined", "null", None]:
#                 return StandardResponse(status=400, success=False, errors=["consignment id is required"])
        
#         error, consignment, error_msg, _ = self.check_stage_po_consignment_exists(po_cust_ref_no, user, is_update, existing_consignment_id)
#         if error:
#             return StandardResponse(status=400, success=False, errors=[error_msg])
#         fields = [
#             "id",
#             "po_lines",
#             "package_name",
#             "package_type",
#             "description",
#             "is_stackable",
#             "measurement_method",
#             "length",
#             "height",
#             "width",
#             "dimension_unit",
#             "weight",
#             "weight_unit",
#             "package_id"
#         ]        

#         result = self.get_packaging_data(consignment, fields, po_line_id=po_line_id)

#         return StandardResponse(status=200, data=result, count=len(result["data"]))


# class StageConsignmentPackagingAPI(ConsignmentStagingMixin, APIView):
    
#     def get(self, request, id=None, *args, **kwargs):
#         po_cust_ref_no = request.GET.get("po")
#         user = request.this_user.id
#         is_update = request.GET.get("is_update") == "true"
#         existing_consignment_id = None
#         if is_update:
#             existing_consignment_id = request.GET.get("consignment_id")
#             if existing_consignment_id in ["", "undefined", "null", None]:
#                 return StandardResponse(status=400, success=False, errors=["consignment id is required"])
        
#         error, consignment, error_msg, _ = self.check_stage_po_consignment_exists(po_cust_ref_no, user, is_update, existing_consignment_id)
#         if error:
#             return StandardResponse(status=400, success=False, errors=[error_msg])
#         fields = [
#             "id",
#             "po_lines",
#             "package_name",
#             "package_type",
#             "description",
#             "is_stackable",
#             "measurement_method",
#             "length",
#             "height",
#             "width",
#             "dimension_unit",
#             "weight",
#             "weight_unit",
#             "package_id"
#         ]        

#         result = self.get_packaging_data(consignment, fields)
#         return StandardResponse(status=200, data=result, count=len(result["data"]))

#         # packages = (
#         #     ConsignmentPackagingStaging.objects
#         #     .filter(consignment=consignment)
#         #     .select_related("packaging_type")
#         #     .prefetch_related("allocations")
#         #     .annotate(
#         #         package_name=F("packaging_type__package_name"),
#         #         package_type=F("packaging_type__package_type"),
#         #         description=F("packaging_type__description"),
#         #         is_stackable=F("packaging_type__is_stackable"),
#         #         measurement_method=F("packaging_type__measurement_method"),
#         #         length=F("packaging_type__length"),
#         #         height=F("packaging_type__height"),
#         #         width=F("packaging_type__width"),
#         #         dimension_unit=F("packaging_type__dimension_unit"),
#         #         weight=F("packaging_type__weight"),
#         #         weight_unit=F("packaging_type__weight_unit"),
#         #         po_lines=Subquery(
#         #             PackagingAllocationStaging.objects.filter(
#         #                 consignment_packaging=OuterRef("id")
#         #             ).exclude(allocated_qty=0).select_related("po_line").values("consignment_packaging").annotate(
#         #                 json_agg=JSONBAgg(
#         #                     JSONObject(
#         #                         id="id",
#         #                         allocated_qty="allocated_qty",
#         #                         customer_reference_number="po_line__po_line_ref",
#         #                         sku="po_line__sku",
#         #                         description="po_line__purchase_order_line__description",
#         #                         qty_packed="po_line__qty_packed",
#         #                         is_dangerous_good="po_line__is_dangerous_good",
#         #                         manufacturing_country="po_line__manufacturing_country",
#         #                         eccn="po_line__eccn",
#         #                     )
#         #                 )
#         #             ).values("json_agg"),
#         #             output_field=JSONField()
#         #         )
#         #     ).values(*fields).order_by("package_id")
#         # )
        
#         # allocated_qty = PackagingAllocationStaging.objects.filter(
#         #     consignment_packaging__consignment=consignment
#         # ).aggregate(total=Sum('allocated_qty'))['total'] or 0
        
#         # total_allocated_qty = ConsignmentPOLineStaging.objects.filter(
#         #     consignment=consignment
#         # ).aggregate(total=Sum('qty_to_fulfill'))['total'] or 0
        
        
#         # remaining_qty = total_allocated_qty - allocated_qty
        
#         # return StandardResponse(status=200, data={"data":packages, "allocated_qty": allocated_qty, "remaining_qty": remaining_qty}, count=len(packages))
#         # return StandardResponse(status=200, data=packages, count=len(packages))
        
    
#     def post(self, request, *args, **kwargs):
#         po_cust_ref_no = request.data.get("purchase_order")
#         user = request.this_user.id
#         is_update = request.data.get("is_update")
#         existing = request.data.get("existing")
#         existing_consignment_id = None
#         if is_update:
#             existing_consignment_id = request.data.get("consignment_id")
#             if existing_consignment_id in ["", "undefined", "null", None]:
#                 return StandardResponse(status=400, success=False, errors=["consignment id is required"])
        
        
#         error, consignment, error_msg, po = self.check_stage_po_consignment_exists(po_cust_ref_no, user, is_update, existing_consignment_id)
#         if error:
#             return StandardResponse(status=400, success=False, errors=[error_msg])
        
#         qty = request.data.get("quantity")
#         # weight = Decimal(request.data.get("weight")) if request.data.get("weight") else 0
#         # weight_unit = request.data.get("weight_unit")
        
#         try:
#             qty = int(qty)
#             if qty <= 0:
#                 raise ValueError
#         except (ValueError, TypeError):
#             return StandardResponse(status=400, success=False, errors=["Invalid quantity"])

#         with transaction.atomic():
#             if existing:
#                 packaging_type_id = request.data.get("packaging_type")
#                 if not packaging_type_id:
#                     return StandardResponse(status=400, success=False, errors=["Packaging type is required"])
#             else:
#                 request.data["supplier"] = po.supplier.id
#                 serializer = PackagingTypeSerializer(data=request.data)
#                 if not serializer.is_valid():
#                     return StandardResponse(status=400, success=False, errors=serializer.errors)
#                 packaging_type_id = serializer.save().id
#             last_package = (
#                 ConsignmentPackagingStaging.objects
#                 .filter(consignment=consignment)
#                 .aggregate(max_id=Max("package_id"))
#             )

#             last_id_str = last_package["max_id"]
#             if last_id_str and last_id_str.startswith("PCKG"):
#                 last_num = int(last_id_str[4:])
#             else:
#                 last_num = 0
            
#             packages = []
#             for i in range(1, qty + 1):
#                 new_num = last_num + i
#                 package_id = f"PCKG{new_num:03d}"
#                 packages.append(
#                     ConsignmentPackagingStaging(
#                         consignment=consignment,
#                         packaging_type_id=packaging_type_id,
#                         package_id=package_id,
#                     )
#                 )
                
#             if packages:
#                 try:
#                     ConsignmentPackagingStaging.objects.bulk_create(packages)
#                 except Exception as e:
#                     transaction.set_rollback(True)
#                     return StandardResponse(status=400, success=False, errors=[str(e)])
                 
#         return StandardResponse(status=201, success=True, message="Consignment Packages created successfully", data={})
    
#     @transaction.atomic()
#     def put(self, request,id=None):
#         package_id = id
#         weight = request.data.get("weight")
#         weight_unit = request.data.get("weight_unit")

#         # Validation
#         if not package_id:
#             return StandardResponse(status=400, success=False, errors="Stage package ID required")

#         if weight is None or weight_unit is None:
#             return StandardResponse(status=400, success=False, errors="Weight and weight unit are required")

#         try:
#             weight = Decimal(weight)
#         except ValueError:
#             return StandardResponse(status=400, success=False, errors="Weight must be a valid number")

#         if weight <= 0:
#             return StandardResponse(status=400, success=False, errors="Weight must be greater than 0")

#         try:
#             package = ConsignmentPackagingStaging.objects.get(id=package_id)
#             package.weight = weight
#             package.weight_unit = weight_unit
#             package.save(update_fields=["weight", "weight_unit"])

#             return StandardResponse(status=200, success=True, message="Package weight updated successfully")

#         except ConsignmentPackagingStaging.DoesNotExist:
#             return StandardResponse(status=404, success=False, errors="Invalid ConsignmentPackagingStaging ID")

#         except Exception as e:
#             return StandardResponse(status=500, success=False, errors=str(e))

    
#     def delete(self, request, id=None, *args, **kwargs):
#         try:
#             obj = (ConsignmentPackagingStaging.objects
#                     .select_related('consignment')
#                     .prefetch_related('allocations')
#                     .get(id=id))
#         except ConsignmentPackagingStaging.DoesNotExist:
#             return StandardResponse(status=400, success=False, errors=["Object not found"])

#         allocations = obj.allocations.all()
#         updates = []
#         for alloc in allocations:

#             updates.append(ConsignmentPOLineStaging(
#                 id=alloc.po_line_id,
#                 qty_packed=F('qty_packed') - alloc.allocated_qty,
#                 qty_remaining=F('qty_remaining') + alloc.allocated_qty,
#             ))

#         with transaction.atomic():
#             if updates:
#                 ConsignmentPOLineStaging.objects.bulk_update(
#                     updates,
#                     ['qty_packed', 'qty_remaining']
#                 )

#             consignment = obj.consignment
#             obj.delete()
#             # self.update_stage_packed_quantities(obj.consignment)
#             self.update_stage_poline_packages(consignment)

#         return StandardResponse(status=200, message="Package deleted successfully.")
        
# class StagePOLineAllocationAPI(ConsignmentStagingMixin, APIView):

#     def put(self, request, id=None, *args, **kwargs):
#         try:
#             po_line = ConsignmentPOLineStaging.objects.select_related("consignment").get(id=id)
#         except ConsignmentPOLineStaging.DoesNotExist:
#             return StandardResponse(status=400, success=False, errors=["Po-line not found"])
        
#         packages = request.data.get("packages", [])

#         if not packages:
#             return StandardResponse(status=400, success=False, errors=["Packages data is required"])

#         allocations_to_create = []
#         to_update = []

#         with transaction.atomic():
#             try:
#                 package_ids = {pk.get("package_id"): pk.get("allocated_qty") for pk in packages if pk.get("package_id") and pk.get("allocated_qty")}

#                 if not package_ids:
#                     return StandardResponse(status=400, success=False, errors=["package_id and allocated_qty required for all entries"])
                
#                 existing_allocations = PackagingAllocationStaging.objects.filter(
#                     consignment_packaging_id__in=package_ids.keys(), po_line=po_line
#                 ).select_related("consignment_packaging")

#                 existing_allocations_map = {allocation.consignment_packaging_id: allocation for allocation in existing_allocations}

#                 for package_id, allocated_qty in package_ids.items():
#                     package_id = UUID(package_id)
#                     if package_id in existing_allocations_map:
#                         existing_allocations_map[package_id].allocated_qty = allocated_qty
#                         to_update.append(existing_allocations_map[package_id])
#                     else:
#                         try:
#                             package = ConsignmentPackagingStaging.objects.get(id=package_id)
#                             allocations_to_create.append(
#                                 PackagingAllocationStaging(
#                                     consignment_packaging_id=package.id, po_line=po_line, allocated_qty=allocated_qty
#                                 )
#                             )
#                         except ConsignmentPackagingStaging.DoesNotExist:
#                             return StandardResponse(status=400, success=False, errors=[f"Package {package_id} not found"])
                        
#                 if to_update:
#                     PackagingAllocationStaging.objects.bulk_update(to_update, ['allocated_qty'])

#                 if allocations_to_create:
#                     PackagingAllocationStaging.objects.bulk_create(allocations_to_create)

#                 self.update_stage_packed_quantities(po_line.consignment)
#                 self.update_stage_poline_packages(po_line.consignment)

#             except Exception as e:
#                 transaction.set_rollback(True)
#                 return StandardResponse(status=400, success=False, errors=[str(e)])

#         return StandardResponse(status=200, success=True, message="Po lines updated successfully", data={})



# class StagePackagingAllocationAPI(ConsignmentStagingMixin, APIView):
    
#     def get(self, request, id=None, *args, **kwargs):
#         try:
#             pacakage = ConsignmentPackagingStaging.objects.get(id=id)
#         except:
#             return StandardResponse(status=400, success=False, errors=["Package does not exists"])
        
#         allocations = PackagingAllocationStaging.objects.filter(consignment_packaging_id=id)
#         if allocations.exists():
#             po_lines = allocations.select_related("po_line").annotate(
#                 po_ref=F("po_line__po_ref"),
#                 po_line_ref=F("po_line__po_line_ref"),
#                 qty_to_fulfill=F("po_line__qty_to_fulfill"),
#                 qty_remaining=F("po_line__qty_remaining"),
#                 qty_packed=F("po_line__qty_packed"),
#                 sku=F("po_line__sku"),
#                 is_dangerous_good=F("po_line__is_dangerous_good"),
#                 manufacturing_country=F("po_line__manufacturing_country"),
#                 eccn=F("po_line__eccn"),
#                 poline_id=F("po_line__id"),
#             ).values().order_by("po_line__purchase_order_line__customer_reference_number")
#         else:
#             po_lines = ConsignmentPOLineStaging.objects.filter(consignment=pacakage.consignment).annotate(
#                 poline_id=F("id")
#                 ).values().order_by("purchase_order_line__customer_reference_number")
        
#         return StandardResponse(status=200, data=po_lines, count=po_lines.count())
                
#     def put(self, request, id=None, *args, **kwargs):
#         try:

#             package = ConsignmentPackagingStaging.objects.get(id=id)
#             if not package.weight or not package.weight_unit:
#                 return StandardResponse(status=400, success=False, errors=["Weight and UOM can't be empty"])

#         except ConsignmentPackagingStaging.DoesNotExist:
#             return StandardResponse(status=400, success=False, errors=["Package not found"])
        
#         po_lines = request.data.get("po_lines", [])
#         # allocations_to_create = [
#         #     PackagingAllocationStaging(consignment_packaging=package, po_line_id=po.get("po_line"), allocated_qty=po.get("allocated_qty"))
#         #     for po in po_lines
#         # ]
#         allocations_to_create = []
#         with transaction.atomic():
#             try:
#                 for po in po_lines:
#                     po_line_id = po.get("poline_id")
#                     existing_allocation = PackagingAllocationStaging.objects.filter(consignment_packaging=package, po_line__id=po_line_id).first()
#                     if existing_allocation:
#                         existing_allocation.allocated_qty = po.get("allocated_qty")
#                         existing_allocation.save()
#                     else:
#                         po_line = ConsignmentPOLineStaging.objects.filter(id=po_line_id,po_line_ref=po.get("po_line_ref")).first()

#                         # po_line = ConsignmentPOLineStaging.objects.filter(po_line_ref=po.get("po_line_ref")).first()                        
#                         if po_line:
#                             allocations_to_create.append(PackagingAllocationStaging(consignment_packaging=package, po_line=po_line, allocated_qty=po.get("allocated_qty")))
#                         else:
#                             return StandardResponse(status=400, success=False, errors=["PO Line not found"])

#                 if allocations_to_create:
#                     PackagingAllocationStaging.objects.bulk_create(allocations_to_create)
#                 self.update_stage_packed_quantities(package.consignment)
#                 self.update_stage_poline_packages(package.consignment)
                
#             except Exception as e:
#                 transaction.set_rollback(True)
#                 return StandardResponse(status=400, success=False, errors=[str(e)])
            
#         return StandardResponse(status=201, success=True, message="Consignment Packages created successfully", data={})


# class StageConsignmentAddressAPI(ConsignmentStagingMixin, APIView):
    
#     def post(self, request, *args, **kwargs):
#         po_cust_ref_no = request.data.get("po")
#         is_update = request.data.get("is_update") or False
#         user = request.this_user.id
#         existing_consignment_id = None
#         if is_update:
#             existing_consignment_id = request.data.get("consignment_id")
#             if existing_consignment_id in ["", "undefined", "null", None]:
#                 return StandardResponse(status=400, success=False, errors=["consignment id is required"])
        
                
#         error, consignment, error_msg, _ = self.check_stage_po_consignment_exists(po_cust_ref_no, user, is_update, existing_consignment_id)
#         if error:
#             return StandardResponse(status=400, success=False, errors=[error_msg])
        
#         serializer = ConsignmentStagingSerializer(consignment, data=request.data, partial=True)
#         if not serializer.is_valid():
#             return StandardResponse(status=400, success=False, errors=serializer.errors)
#         obj = serializer.save()
#         return StandardResponse(status=201, message="Address Added Successfully.", data={"id":obj.id})
    
        
# class StageConsignmentDocumentAPI(ConsignmentStagingMixin, APIView):
#     parser_classes = [MultiPartParser, FormParser]
    
#     def get(self, request, id=None, *args, **kwargs):
#         po_cust_ref_no = request.GET.get("po")
#         user = request.this_user.id
#         is_update = request.GET.get("is_update") == "true"
#         existing_consignment_id = None
#         if is_update:
#             existing_consignment_id = request.GET.get("consignment_id")
#             if existing_consignment_id in ["", "undefined", "null", None]:
#                 return StandardResponse(status=400, success=False, errors=["consignment id is required"])
        
#         error, consignment, error_msg, _ = self.check_stage_po_consignment_exists(po_cust_ref_no, user, is_update, existing_consignment_id)
#         if error:
#             return StandardResponse(status=400, success=False, errors=[error_msg])
                
#         documents = ConsignmentDocumentStaging.objects.filter(consignment=consignment).prefetch_related("attachments")

#         data = [
#             {
#                 "id": doc.id,
#                 "document_type": doc.document_type,
#                 "attachments": [
#                     {"id": att.id, "file": att.file.url}
#                     for att in doc.attachments.all()
#                 ],
#             }
#             for doc in documents
#         ]

#         return StandardResponse(status=200, data=data, count=documents.count()) 
#         # return StandardResponse(status=200, data=GetConsignmentDocumentStagingSerializer(documents, many=True).data, count=documents.count()) 

    
    
#     @transaction.atomic
#     def post(self, request, *args, **kwargs):
#         po_cust_ref_no = request.data.get("po")
#         is_update = request.data.get("is_update") == "true"
#         user = request.this_user.id 
#         existing_consignment_id = None
#         if is_update:
#             existing_consignment_id = request.data.get("consignment_id")
#             if existing_consignment_id in ["", "undefined", "null", None]:
#                 return StandardResponse(status=400, success=False, errors=["consignment id is required"])
        
#         error, consignment, error_msg, _ = self.check_stage_po_consignment_exists(po_cust_ref_no, user, is_update, existing_consignment_id)
#         if error:
#             return StandardResponse(status=400, success=False, errors=[error_msg])
        
#         documents = {}
#         files_dict = {}

#         documents_to_delete = request.data.get("deleted_doc_ids")
#         attachments_to_exclude = request.data.get("attachment_urls")
        
#         # Extract document strings
#         for key, value in request.data.items():
#             if key.startswith("document_type["):
#                 index = int(key.replace("document_type[", "").replace("]", ""))
#                 documents[index] = value 
#                 files_dict[value] = []
                

#         for key, file in request.FILES.items():
#             if key.startswith("attachments["):
#                 parts = key.replace("attachments[", "").replace("]", "").split("[")
#                 file_index = int(parts[0])
#                 doc_type = documents[file_index]

#                 files_dict[doc_type].append(file)
        
#         try:
#             if documents_to_delete:
#                 ConsignmentDocumentStaging.objects.filter(id__in=documents_to_delete.split(",")).delete()

#             if attachments_to_exclude:
#                 attachments_to_exclude = attachments_to_exclude.split(",")
#                 existing_attachments = ConsignmentDocumentAttachmentStaging.objects.filter(document__consignment=consignment)

#                 for attachment in existing_attachments:
#                     att_name = attachment.file.url

#                     if att_name not in attachments_to_exclude:
#                         if attachment.file:
#                             default_storage.delete(attachment.file.path)
#                         attachment.delete()
                
#             files_to_create = []
#             for doc_type, files in files_dict.items():
#                 document_obj, _ = ConsignmentDocumentStaging.objects.get_or_create(consignment=consignment, document_type=doc_type)
#                 files_to_create.extend([
#                     ConsignmentDocumentAttachmentStaging(document=document_obj, file=file)
#                     for file in files
#                 ])
                
#             ConsignmentDocumentAttachmentStaging.objects.bulk_create(files_to_create)
            
        
#         except Exception as e:
#             transaction.set_rollback(True)
#             return StandardResponse(status=400, success=False, errors=[str(e)])
        
#         return StandardResponse(status=201, message="Documents Added Successfully.")
    
    
#     @transaction.atomic
#     def delete(self, request, id=None, *args, **kwargs):
#         is_update = request.GET.get("is_update") == "true"
#         user = request.this_user.id 
#         existing_consignment_id = None
#         if is_update:
#             existing_consignment_id = request.GET.get("consignment_id")
#             if existing_consignment_id in ["", "undefined", "null", None]:
#                 return StandardResponse(status=400, success=False, errors=["consignment id is required"])
        
        
#         error, consignment, error_msg, _ = self.check_stage_po_consignment_exists(id, user, is_update, existing_consignment_id)
#         if error:
#             return StandardResponse(status=400, success=False, errors=[error_msg])
        
#         existing_attachments = ConsignmentDocumentAttachmentStaging.objects.filter(document__consignment=consignment)

#         try:
#             for attachment in existing_attachments:
#                 if attachment.file:
#                     default_storage.delete(attachment.file.path)
#                 attachment.delete()
                
#             ConsignmentDocumentStaging.objects.filter(consignment=consignment).delete()
            
#             return StandardResponse(status=201, message="Documents Deleted Successfully.")
        
#         except Exception as e:
#             transaction.set_rollback(True)
#             return StandardResponse(status=400, success=False, errors=[str(e)])
        
from django.db.models import Prefetch

class ConsignmentListAPI(FilterMixin, PurchaseOrderMixin,ConsignmentStatusSummary, SearchAndFilterMixin, PaginationMixin, PurchaseOrderLineQuantityMixin, ConsignmentMixin, APIView):

    
    transform_fields = {
        # "purchase_order": "purchase_order__customer_reference_number",
        # "adhoc": "adhoc__customer_reference_number",
        # "supplier": "supplier__name",
        # "client": "client__name",
        "console_id": "console",
        "consignor_address": "consignor_address__address_name",
        "delivery_address": "delivery_address__address_name",
        "freight_forwarder": "console__freight_forwarder",
        "packages": "packages_count",
        "ship_from_city" : "consignor_address",
        "ship_from_state" : "consignor_address",
        "ship_from_country" : "consignor_address",
        "ship_to_city" : "delivery_address",
        "ship_to_state" : "delivery_address",
        "ship_to_country" : "delivery_address",
    }
    
    fields = [
        "id", "consignment_id","supplier__name", "client__name","console__console_id","consignment_status", 
        "created_by", "created_by_name", "created_at","type","has_awb_files"
    ]
        
    search_fields = [
        "consignment_id", "supplier__name", "client__name", 
        "packages", "actual_pickup_datetime", "requested_pickup_datetime", "consignment_status","console__console_id",
        "created_at","is_completed", "type"
    ]
    
    def make_filters(self, request, qs=None):
        
        status = request.GET.get("status", "")
        po_line = request.GET.get("linked_to_po_line")
        po = request.GET.get("linked_to_po")
        search = request.GET.get("q", "").strip()

        queryset = qs
        filters = Q()
        filters = self.build_filter(request.this_user,filters=filters,model="Consignment")

        if status and status.lower() != "all":
            filters &= Q(consignment_status=status)
        
        if po_line:
            obj = PurchaseOrderLine.objects.filter(customer_reference_number = po_line,purchase_order__customer_reference_number = po).first()
            consignment_ids = self.line_related_consignments(obj)

            filters &= Q(id__in=consignment_ids)

        apply_filters = self.make_filters_list(request)

        if queryset:
            if apply_filters:
                for f in apply_filters:
                    queryset = self.filter_measured_annotations(queryset,f['column'])
                    f['column'] = ConsignmentListAPI.transform_fields.get(f['column'], f['column'])

                apply_filters = self.appy_dynamic_filter(apply_filters)  
                filters &= Q(apply_filters)
                queryset = queryset.filter(filters)

            if search:
                queryset = self.apply_search(ConsignmentListAPI.search_fields, queryset, search)
            
            queryset = queryset.filter(filters)
        return queryset if qs else filters


    def get_all_data(self, request):
        
         
        pg = request.GET.get("pg") or 0
        limit = request.GET.get("limit") or 25
       
        filters = Q()
        
        queryset = (
            Consignment.objects.filter(filters)
            .exclude(
                consignment_id__startswith = "DRAFT",
                consignment_status = ConsignmentStatusChoices.DRAFT
            )
            .select_related(
                "supplier", "client","console__console","created_by"
            )
            .prefetch_related("consignment_awb")
            .annotate(
                created_by_name=F("created_by__name"),
                has_awb_files=Count("consignment_awb", distinct=True),
            )
            .values(*ConsignmentListAPI.fields)
            .distinct()
            .order_by("-consignment_id")
        )
        
        queryset = ConsignmentListAPI.make_filters(self,request, queryset)

        count = queryset.count()
        paginate_result = self.paginate_results(queryset, pg, limit)
        for data in paginate_result:
            self._tranform_object(data)
        
        return StandardResponse(success=True, data=paginate_result, count=count, status=200)
    

    def check_supplier(self,con,user):
        con_supplier = con.supplier
        if user.role  == Role.SUPPLIER_USER:
            supplier = user.profile()

            if con_supplier == supplier.supplier:
                return True 

        return False


    @transaction.atomic
    def get(self, request, id=None, *args, **kwargs):
        
        try:
        
            ranges = [
                {"label": "<2", "max_days": 2, "min_days": None},  
                {"label": "2-7", "max_days": 7, "min_days": 2},    
                {"label": "7-15", "max_days": 15, "min_days": 7},  
                {"label": "15-30", "max_days": 30, "min_days": 15},
                {"label": "30-45", "max_days": 45, "min_days": 30},
                {"label": "45-60", "max_days": 60, "min_days": 45},
                {"label": "90-120", "max_days": 120, "min_days": 90},
                {"label": "120-150", "max_days": 150, "min_days": 120},
                {"label": "150-180", "max_days": 180, "min_days": 150},
                {"label": ">180", "max_days": None, "min_days": 180},
            ]

            days = request.GET.get("days")
            
            if days:
                
                items  = next((r for r in ranges if r.get("label") == days), None)
                
                if items:
                    paginated_result, count =  self.consignment_status_summary(request, min_days=items.get("min_days"), max_days=items.get("max_days"))
            
                    for data in paginated_result:
                        self._tranform_object(data)
                    
                    return StandardResponse(success=True, data=paginated_result, count=count, status=200)    

                return StandardResponse(status=400, data=[], message="Invalid days format. Use <, > or - to specify the range.")
            
            elif id == "list":
                return self.get_all_data(request)
            
            
            consignment = Consignment.objects.get(consignment_id=id)
            if not consignment:
                return StandardResponse(status=400, success=False, errors=["Consignment not found"])
            
            # is_view = request.GET.get("is_view") == "true"
            # generate_qr = True if request.GET.get("generate_qr") == "true" else False
            # show_package_status = True if request.GET.get("show_package_status") == "true" else False

            # if is_view or generate_qr or show_package_status:
            #     return self.get_single_data(consignment=consignment)
            

            # return self.get_single_data(consignment=consignment)

            return StandardResponse(success=True, data=[], status=200)
        
        except ObjectDoesNotExist:
            return StandardResponse(status=400, success=False, errors=["Purchase Order or Adhoc Order not found"])
        
        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=[str(e)])
        
    # @role_required(OperationUserRole.L1,OperationUserRole.L2,Role.SUPPLIER_USER)
    # @transaction.atomic
    # def post(self, request, *args, **kwargs):
    #     po_cust_ref_no = request.data.get("po")
    #     user = request.this_user.id
    #     is_update = request.data.get("is_update") or False
    #     submit = request.data.get("is_submit") or False
    #     existing_consignment_id = None
    #     if is_update:
    #         existing_consignment_id = request.data.get("consignment_id")
    #         if existing_consignment_id in ["", "undefined", "null", None]:
    #             return StandardResponse(status=400, success=False, errors=["consignment id is required"])
                
    #     error, stage_consignment, error_msg, po = self.check_stage_po_consignment_exists(po_cust_ref_no, user, is_update, existing_consignment_id)
    #     if error:
    #         return StandardResponse(status=400, success=False, errors=[error_msg])
        
    #     try:
    #         # Delete all the packages which do not contain any skus or have po lines with 0 quantity
    #         self.delete_unused_stage_packages(consignment=stage_consignment)
            
    #         if is_update:
    #             try:
    #                 consignment = Consignment.objects.get(consignment_id=stage_consignment.existing_consignment_id)
    #             except Consignment.DoesNotExist:
    #                 return StandardResponse(status=400, success=False, errors=["Consignment not found"])    
    #         else:
    #             consignment_data = {
    #                 "purchase_order": po,
    #                 "supplier": po.supplier,
    #                 "client": po.client,
    #                 "consignor_address": stage_consignment.consignor_address,
    #                 "delivery_address": stage_consignment.delivery_address,
    #                 "requested_pickup_datetime": stage_consignment.pickup_datetime,
    #                 "actual_pickup_datetime": None,
    #                 "pickup_timezone": stage_consignment.pickup_timezone,
    #                 "packages": [],
    #                 "created_by": stage_consignment.user
    #             }
    #             consignment = Consignment.objects.create(**consignment_data)

    #             notif = Notification.create_notification(
    #                 type = NotificationChoices.CONSIGNMENT,
    #                 model = "Consignment",
    #                 status = "Created",
    #                 ref=consignment.consignment_id,
    #                 attachments=""
    #             )

    #         if is_update:
    #             self.update_quantities_on_delete_packaging(consignment=consignment)
    #             ConsignmentPackaging.objects.filter(consignment=consignment).delete()
    #             ConsignmentDocument.objects.filter(consignment=consignment).delete()
                
                
    #         # Transfer ConsignmentPOLineStaging -> ConsignmentPackaging            
    #         for packaging in stage_consignment.stage_packagings.all():
    #             consignment_packaging = ConsignmentPackaging.objects.create(
    #                 consignment=consignment,
    #                 packaging_type=packaging.packaging_type,
    #                 package_id=packaging.package_id,
    #                 weight = packaging.weight,
    #                 weight_unit = packaging.weight_unit
    #             )
                
    #             for allocation in packaging.allocations.all():
    #                 if not allocation.allocated_qty <= 0 :
    #                     PackagingAllocation.objects.create(
    #                         consignment_packaging=consignment_packaging,
    #                         purchase_order_line=allocation.po_line.purchase_order_line,
    #                         allocated_qty=allocation.allocated_qty,
    #                         hs_code=allocation.po_line.hs_code,
    #                         is_dangerous_good=allocation.po_line.is_dangerous_good,
    #                         manufacturing_country=allocation.po_line.manufacturing_country,
    #                         eccn=allocation.po_line.eccn,
    #                     )
            
    #         self.update_packages(consignment)
            
    #         # Transfer ConsignmentDocumentStaging -> ConsignmentDocument     
    #         for staging_document in stage_consignment.documents.all():
    #             document = ConsignmentDocument.objects.create(
    #                 consignment=consignment,
    #                 document_type=staging_document.document_type
    #             )
                
    #             # Transfer ConsignmentDocumentAttachmentStaging -> ConsignmentDocumentAttachment
    #             ConsignmentDocumentAttachment.objects.bulk_create([
    #                 ConsignmentDocumentAttachment(
    #                     document=document,
    #                     file=attachment.file
    #                 ) for attachment in staging_document.attachments.all()
    #             ])
            
    #         # Update purchase order line open qty and fufilled qty
    #         self.when_create_consignment(consignment)
    #         if submit:
    #             stage_consignment.delete()
                
    #         if consignment.consignment_status == ConsignmentStatusChoices.REJECTED:
    #             consignment.consignment_status = ConsignmentStatusChoices.PENDING_FOR_APPROVAL
    #             consignment.save()
                
    #         return StandardResponse(status=201, data={"consignment_id":consignment.consignment_id},message="Consignment created Successfully.")
        
    #     except Exception as e:
    #         transaction.set_rollback(True)
    #         return StandardResponse(status=400, success=False, errors=[str(e)])
    
    
    
class ConsignmentStatusUpdateAPI(PurchaseOrderLineQuantityMixin, APIView):

    @role_required(OperationUserRole.L1)
    def has_approve_permission(self,request):
        ## this is to check if the user has permission to approve the consignment using decorator
        return True

    @transaction.atomic
    def post(self, request, id=None, *args, **kwargs):
        """
        We are not updating these status in this api
        --console assignment
        --freight forwarder assignment
        """
        
        try:

            status = request.data.get("status")
            consignment_ids = request.data.get("consignment_ids", [])
            actual_pickup_datetime = request.data.get("actual_pickup_datetime")
            cancellation_reason = request.data.get("cancellation_reason","")
            
            if not consignment_ids:
                return StandardResponse(status=400, errors=["Consignment IDs cannot be empty."])
            
            consignments = Consignment.objects.filter(consignment_id__in=consignment_ids)
            
            if not consignments.exists():
                return StandardResponse(status=404, errors=["Consignments not found."])
        
        
            if status == ConsignmentStatusChoices.PENDING_CONSOLE_ASSIGNMENT:
                self.has_approve_permission(request)
                ConsignmentStatusService.approved(consignments)


            elif status == ConsignmentStatusChoices.PICKUP_COMPLETED:
                ConsignmentStatusService.pickup_completed(consignments)


            elif status == ConsignmentStatusChoices.AT_CUSTOM:
                ConsignmentStatusService.at_customs(consignments)

            elif status == ConsignmentStatusChoices.CUSTOMS_CLEARED:
                ConsignmentStatusService.customs_cleared(consignments)

            elif status == ConsignmentStatusChoices.OUT_FOR_DELIVERY:
                ConsignmentStatusService.out_for_delivery(consignments)
            
            elif status == ConsignmentStatusChoices.REJECTED:
                ConsignmentStatusService.rejected(
                    consignments,
                    request.data.get("rejection_code"),
                    request.data.get("rejection_reason")
                )
            
            elif status == ConsignmentStatusChoices.DELIVERED:
                ConsignmentStatusService.delivered(consignments)
                
            elif status == ConsignmentStatusChoices.CANCELLED:
                ConsignmentStatusService.cancelled(consignments, cancellation_reason)
                
            elif status == ConsignmentStatusChoices.PENDING_BID:
                ConsignmentStatusService.pending_bid(consignments)
            
            else:
                ConsignmentStatusService.update_status_with_pickup_datetime(
                    consignments, status, actual_pickup_datetime
                )

            for con in consignments:
                con.update_console_status()
                ConsignmentServices.notify_consignment_update(request.this_user, instance=con)

            return StandardResponse(status=200, message="Status updated successfully.")
          
        except ServiceError as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=[e.error])
        
        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=[str(e) or "Something went wrong"])
        

class ConsignmentStatusCountAPI(FilterMixin,APIView):
    def get(self, request, id=None, *args, **kwargs):
        filters = Q()

        all_statuses = [choice[0] for choice in ConsignmentStatusChoices.choices]
        all_statuses.remove(ConsignmentStatusChoices.DRAFT)
        # all_statuses.remove(ConsignmentStatusChoices.PARTIALLY_RECEIVED)
        filters = self.build_filter(request.this_user,"Consignment",filters)
        filters = filters & Q(consignment_status__in = all_statuses)
        # filters["consignment_status__in"] = all_statuses
        
        status_counts = (
            Consignment.objects
            .filter(filters)
            .values('consignment_status')
            .annotate(count=Count('id', distinct=True))
        )
        
        status_dict = {status: 0 for status in all_statuses}
        for status in status_counts:
            status_dict[status['consignment_status']] = status['count']
        
        return StandardResponse(status=200, data=status_dict)
       
    
class ConsignmentAuditTrailGetAPI(APIView):
    
    def get(self, request, id=None, *args, **kwargs):
        try:
            obj = Consignment.objects.get(consignment_id=id)
        except:
            return StandardResponse(status=400, success=False, errors=["Consignment not found"])
        
        audit_trails = ConsignmentAuditTrail.objects.prefetch_related("fields").filter(consignment=obj).order_by("-updated_at")
        
        data = [
            {
                "audit_trail_id": trail.id,
                "updated_by": trail.updated_by.name,
                "created_at": trail.created_at,
                
                "fields": [
                    {
                        "field_name": field.field_name,
                        "title" : field.title,
                        "description": field.description,
                        "old_value": field.old_value,
                        "new_value": field.new_value,
                    }
                    for field in trail.fields.all()
                ],
            }
            for trail in audit_trails
        ]
        
        return StandardResponse(success=True, data=data, status=200)
    
    
class ConsignmentAssignFFAPI(APIView):
    ## this api is not needed now we changes the flow now we are assigning the ff to console no to consignment
    @role_required(OperationUserRole.L1,OperationUserRole.L2)
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        freight_forwarder = request.data.get("freight_forwarder")
        actual_pickup_datetime = request.data.get("actual_pickup_datetime")
        consignment_ids = request.data.get("consignment_ids", [])
        
        files = [file for key, file_list in request.FILES.lists() if key.startswith("ff_documents") for file in file_list]

        if not consignment_ids:
            return StandardResponse(status=400, errors=["Consignment IDs cannot be empty."])
        
        consignments = Consignment.objects.filter(consignment_id__in=consignment_ids.split(","))
        
        if not consignments.exists():
            return StandardResponse(status=404, errors=["Consignments not found."])
        try:
            if files:
                # Delete previous records and related files
                previous_documents = ConsignmentFFDocument.objects.filter(consignment__in=consignments)
                for ff_doc in previous_documents:
                    if ff_doc.file:
                        default_storage.delete(ff_doc.file.path)
                previous_documents.delete()

                # Create new records                
                ff_documents = [
                    ConsignmentFFDocument(consignment=consignment, file=file)
                    for consignment in consignments
                    for file in files
                ]
                ConsignmentFFDocument.objects.bulk_create(ff_documents)
            else:
                return StandardResponse(status=400, errors=["FF Documents cannot be empty."])
            
            actual_pickup_datetime = datetime.strptime(actual_pickup_datetime, "%Y-%m-%d %I:%M %p")

            formatted_date = actual_pickup_datetime.strftime("%Y-%m-%d %H:%M:%S")
            consignments.update(freight_forwarder_id=freight_forwarder, consignment_status=ConsignmentStatusChoices.FREIGHT_FORWARDER_ASSIGNED, actual_pickup_datetime=formatted_date)

        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=[str(e)])
                
        return StandardResponse(status=201, message="Freight Forwarder Assigned Successfully.")


class ConsignmentPackageListAPI(APIView):
    
    # def get(self, request, id=None, *args, **kwargs):
    #     fields = [
    #         "id",
    #         "po_lines",
    #         "package_name",
    #         "package_type",
    #         "description",
    #         "is_stackable",
    #         "measurement_method",
    #         "length",
    #         "height",
    #         "width",
    #         "dimension_unit",
    #         "weight",
    #         "weight_unit",
    #         "package_id"
    #     ] 
    #     packages = ConsignmentPackaging.objects.filter(consignment__consignment_id=id).select_related("packaging_type").annotate(
    #         package_name=F("packaging_type__package_name"),
    #         package_type=F("packaging_type__package_type"),
    #         description=F("packaging_type__description"),
    #         is_stackable=F("packaging_type__is_stackable"),
    #         measurement_method=F("packaging_type__measurement_method"),
    #         length=F("packaging_type__length"),
    #         height=F("packaging_type__height"),
    #         width=F("packaging_type__width"),
    #         dimension_unit=F("packaging_type__dimension_unit"),
    #         weight=F("packaging_type__weight"),
    #         weight_unit=F("packaging_type__weight_unit"),
    #         po_lines=Subquery(
    #             PackagingAllocation.objects.filter(
    #                 consignment_packaging=OuterRef("id")
    #             ).exclude(allocated_qty=0).select_related("purchase_order_line").values("consignment_packaging").annotate(
    #                 json_agg=JSONBAgg(
    #                     JSONObject(
    #                         id="id",
    #                         allocated_qty="allocated_qty",
    #                         reference_number="purchase_order_line__reference_number",
    #                         description="purchase_order_line__description",
    #                         customer_reference_number="purchase_order_line__customer_reference_number",
    #                         sku="purchase_order_line__sku",
    #                         is_dangerous_good="is_dangerous_good",
    #                         manufacturing_country="manufacturing_country",
    #                         eccn="eccn",
    #                     )
    #                 )
    #             ).values("json_agg"),
    #             output_field=JSONField()
    #         )
    #     ).values(*fields).order_by("package_id")
        
    #     return StandardResponse(status=200, data=packages, count=len(packages))

    def get(self, request, id=None, *args, **kwargs):
        # Step 1: Fetch all ConsignmentPackages with related PackagingType
        
        package_list, errors = ConsignmentWorkflowServices.get_consignment_packages(consignment_id=id)
        if errors:
            return StandardResponse(status=400, success=False, errors=errors)
        
        count = len(package_list)
        if not package_list:
            return StandardResponse(data = [], status=200)
        
        return StandardResponse(data = package_list, status=200,count=count)
    
class AdhocConsignmentCreateAPI(ConsignmentMixin, APIView):
    
    def get(self, request, id=None, *args, **kwargs):
        try:
            consignment = Consignment.objects.get(consignment_id=id)
        except:
            return StandardResponse(status=400, success=False, errors=["Consignment does not exists"])
        
        return StandardResponse(status=200, data=GetConsignmentSerializer(consignment).data)
    
    @role_required(OperationUserRole.L1,OperationUserRole.L2,Role.SUPPLIER_USER,Role.ADMIN)
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        po_cust_ref_no = request.data.get("purchase_order")
        user = request.this_user.id
        supplier = client = None
        
        if request.this_user.role == Role.ADMIN:
            supplier = Supplier.objects.filter(id=request.data.get("supplier")).select_related("client")
            client = Client.objects.filter(id=supplier.first().client.id)
            storerkey_ids = list(supplier.first().storerkeys.values_list("id", flat=True))
            if not storerkey_ids:
                return StandardResponse(status=400, success=False, errors=["Selected supplier don't have any storer key assigned"])
                
            
        elif request.this_user.role == Role.SUPPLIER_USER:
            profile = request.this_user.supplier_profile
            supplier = Supplier.objects.filter(id=profile.supplier.id).select_related("client")
            client = Client.objects.filter(id=supplier.first().client.id)
            storerkey_ids = list(profile.storerkeys.values_list("id", flat=True))
            
        else:
            return StandardResponse(status=400, success=False, errors=["Only Supplier user and Admin can create Consignment"])
                            
        adhoc_po, create = AdhocPurchaseOrder.objects.get_or_create(
            customer_reference_number=po_cust_ref_no,
            defaults={"client": client.first(), "supplier": supplier.first(), "storerkey_id": storerkey_ids[0]}
        )
        
        if not create:
            if adhoc_po.supplier != supplier.first():
                return StandardResponse(status=400, success=False, errors=[f"The PO is Already exists with {po_cust_ref_no}."])

            if adhoc_po.storerkey_id not in storerkey_ids:
                return StandardResponse(status=400, success=False, errors=[f"The PO is Already exists with {po_cust_ref_no}."])
                    
        consignment, created = Consignment.objects.get_or_create(
            adhoc=adhoc_po,
            is_completed=False,
            type=ConsignmentTypeChoices.ADHOC,
            defaults={"created_by_id": user, "supplier": supplier.first(), "client": client.first()}
        )
        
        supplier_data = supplier.values().first()
        client_data = client.values().first()
        
        if not created:
            if str(consignment.created_by_id) == str(user):
                return StandardResponse(status=200, data={"consignment_id": consignment.consignment_id, "client":client_data, "supplier": supplier_data})
            return StandardResponse(status=400, success=False, errors=["Purchase Order already picked for consignment by other supplier"])

        return StandardResponse(status=200, data={"consignment_id": consignment.consignment_id, "client":client_data, "supplier": supplier_data}, message="Saved Successfully")

    
    def put(self, request, id=None, *args, **kwargs):
        try:
            consignment = Consignment.objects.get(consignment_id=id)
        except:
            return StandardResponse(status=400, success=False, errors=["Consignment does not exists"])
        
        if request.data.get("pickup_datetime"):
            request.data["requested_pickup_datetime"] = request.data["pickup_datetime"]
        if request.data.get("is_submit"):
            request.data["is_completed"] = request.data["is_submit"]
            
        serializer = ConsignmentAdhocSerializer(consignment, data=request.data, partial=True)
        if not serializer.is_valid():
            return StandardResponse(status=400, success=False, errors=serializer.errors)
            
        obj = serializer.save()
        
        if obj.is_completed:
            self.update_packages(consignment)
        
        return StandardResponse(status=201, message="Updated Successfully.", data={"id":obj.id})
    
            
class AdhocConsignmentPackageView(AdhocPurchaseOrderLineMixin,APIView):
    
    def _transform_objects(self, data):
        return data.update({
            "customer_reference_number": data.pop("adhoc_line__customer_reference_number"),
            "reference_number": data.pop("adhoc_line__reference_number"),
            "sku": data.pop("adhoc_line__sku"),
        })
    
    # def get(self, request, id=None, *args, **kwargs):
        
    #     if id != "list":
    #         # try:
    #         #     package = ConsignmentPackaging.objects.get(id=id)
    #         # except:
    #         #     return StandardResponse(status=400, success=False, errors=["Package does not exists"])
    #         try:
    #             allocations = PackagingAllocation.objects.filter(consignment_packaging_id=id).values(
    #                 "id", "hs_code", "is_dangerous_good", "eccn", "manufacturing_country", "allocated_qty", "adhoc_line__reference_number", "adhoc_line__customer_reference_number", "adhoc_line__sku"
    #             )

    #             if allocations.exists():            
    #                 for all in allocations:
    #                     self._transform_objects(all)
    #                 return StandardResponse(status=200, data=allocations)
    #             else:
    #                 return StandardResponse(status=200, data=[])
    #         except ValidationError:
    #             return StandardResponse(status=400, success=False, errors=["Package does not exists"])            
        
    #     else:
    #         try:
    #             consignment = Consignment.objects.get(consignment_id=request.GET.get("consignment_id"))
    #         except:
    #             return StandardResponse(status=400, success=False, errors=["Consignment does not exists"])

    #         fields = [
    #             "id",
    #             "po_lines",
    #             "package_name",
    #             "package_type",
    #             "description",
    #             "is_stackable",
    #             "measurement_method",
    #             "length",
    #             "height",
    #             "width",
    #             "dimension_unit",
    #             "weight",
    #             "weight_unit",
    #             "package_id"
    #         ] 
    #         staging_qs = ConsignmentPackagingStaging.objects.filter(
    #             package_id=OuterRef("package_id"),
    #             consignment=consignment
    #         ).values("weight", "weight_unit")
            
    #         packages = ConsignmentPackaging.objects.filter(consignment=consignment).select_related("packaging_type").annotate(
    #             package_name=F("packaging_type__package_name"),
    #             package_type=F("packaging_type__package_type"),
    #             description=F("packaging_type__description"),
    #             is_stackable=F("packaging_type__is_stackable"),
    #             measurement_method=F("packaging_type__measurement_method"),
    #             length=F("packaging_type__length"),
    #             height=F("packaging_type__height"),
    #             width=F("packaging_type__width"),
    #             dimension_unit=F("packaging_type__dimension_unit"),
    #             # weight=F("packaging_type__weight"),
    #             # weight_unit=F("packaging_type__weight_unit"),
    #             weight=Subquery(staging_qs.values("weight")[:1]),
    #             weight_unit=Subquery(staging_qs.values("weight_unit")[:1]),
    #             po_lines=Subquery(
    #                 PackagingAllocation.objects.filter(
    #                     consignment_packaging=OuterRef("id")
    #                 ).exclude(allocated_qty=0).select_related("adhoc_line").values("consignment_packaging").annotate(
    #                     json_agg=JSONBAgg(
    #                         JSONObject(
    #                             id="id",
    #                             allocated_qty="allocated_qty",
    #                             reference_number="adhoc_line__reference_number",
    #                             customer_reference_number="adhoc_line__customer_reference_number",
    #                             sku="adhoc_line__sku",
    #                             is_dangerous_good="is_dangerous_good",
    #                             hs_code="hs_code",
    #                             eccn="eccn",
    #                             manufacturing_country="manufacturing_country",
    #                         )
    #                     )
    #                 ).values("json_agg"),
    #                 output_field=JSONField()
    #             )
    #         ).values(*fields).order_by("package_id")
                    
    #         return StandardResponse(status=200, data=packages)
     
    
    def post(self, request, id=None, *args, **kwargs):
        try:
            consignment = Consignment.objects.get(consignment_id=request.data.get("consignment_id"))
        except:
            return StandardResponse(status=400, success=False, errors=["Consignment does not exists"])
        
        qty = request.data.get("quantity")
        existing = request.data.get("existing")
        try:
            qty = int(qty)
            if qty <= 0:
                raise ValueError
        except (ValueError, TypeError):
            return StandardResponse(status=400, success=False, errors=["Invalid quantity"])


        with transaction.atomic():
            if existing:
                packaging_type_id = request.data.get("packaging_type")
                if not packaging_type_id:
                    return StandardResponse(status=400, success=False, errors=["Packaging type is required"])
            else:
                request.data["supplier"] = consignment.supplier_id
                serializer = PackagingTypeSerializer(data=request.data)
                if not serializer.is_valid():
                    return StandardResponse(status=400, success=False, errors=serializer.errors)
                packaging_type_id = serializer.save().id

            last_package = (
                ConsignmentPackaging.objects
                .filter(consignment=consignment)
                .aggregate(max_id=Max("package_id"))
            )

            last_id_str = last_package["max_id"]
            if last_id_str and last_id_str.startswith("PCKG"):
                last_num = int(last_id_str[4:])
            else:
                last_num = 0
    
            packages = []
            for i in range(1, qty + 1):
                new_num = last_num + i
                package_id = f"PCKG{new_num:03d}"
                packages.append(
                    ConsignmentPackaging(
                        consignment=consignment,
                        packaging_type_id=packaging_type_id,
                        package_id=package_id
                    )
                )
                
            if packages:
                try:
                    ConsignmentPackaging.objects.bulk_create(packages)
                except Exception as e:
                    transaction.set_rollback(True)
                    return StandardResponse(status=400, success=False, errors=[str(e)])
                 
        return StandardResponse(status=201, success=True, message="Consignment Packages created successfully", data={})
    
     
    def delete(self, request, id=None, *args, **kwargs):
        try:
            package = ConsignmentPackaging.objects.get(id=id)
            error = self.check_and_delete_related_adhoc_lines(package)
            if error:
                return StandardResponse(status=400, success=False, errors=[error])
            package.delete()
        except (ConsignmentPackaging.DoesNotExist, ValidationError):
            return StandardResponse(status=400, success=False, errors=["Object not found"])
        
        return StandardResponse(status=200, message="Package deleted successfully.")
        

class AdhocConsignmentAlloationAPI(AdhocPurchaseOrderLineMixin,APIView):
    
    @transaction.atomic
    def post(self, request, package_id=None, *args, **kwargs):
        try:
            package = ConsignmentPackaging.objects.get(id=package_id)
        except:
            return StandardResponse(status=400, success=False, errors=["Package does not exists"])
        po = package.consignment.adhoc
        allocations = []
        documents = []
        
        try:
            packagings = PackagingAllocation.objects.filter(consignment_packaging=package)

            error = self.check_and_delete_related_adhoc_lines(package)
            if error:
                return StandardResponse(status=400, success=False, errors=[error])
            
            packagings.delete()
            ConsignmentDocument.objects.filter(consignment=package.consignment).delete()
            
            for allocation in request.data.get("allocations"):
                customer_ref = allocation.get("customer_reference_number")
                conflict = AdhocPurchaseOrderLine.objects.filter(customer_reference_number=customer_ref).exclude(purchase_order=po).exists()

                if conflict:
                    return StandardResponse(status=400, success=False, errors=[f"PO Line already exists with the {customer_ref} for other Purchase Order"])
                
                po_line, created = AdhocPurchaseOrderLine.objects.get_or_create(
                    purchase_order=po,
                    customer_reference_number=customer_ref,
                    defaults={"reference_number": allocation.get("reference_number"), "sku": allocation.get("sku")}
                )
                if not created:
                    if po_line.sku != allocation.get("sku"):
                        return StandardResponse(status=400, success=False, errors=[f"PO Line already exists with {po_line.sku} SKU."])
                is_dangerous_good = allocation.get("is_dangerous_good")
                allocations.append(PackagingAllocation(
                    consignment_packaging=package,
                    adhoc_line=po_line,
                    hs_code=allocation.get("hs_code"),
                    manufacturing_country=allocation.get("manufacturing_country"),
                    eccn=allocation.get("eccn", False),
                    allocated_qty=allocation.get("allocated_qty"),
                    is_dangerous_good=is_dangerous_good,
                ))
                if is_dangerous_good:
                    documents.append(ConsignmentDocument(
                        consignment=package.consignment,
                        adhoc_line=po_line,
                        document_type=ConsignmentDocumentTypeChoices.DANGEROUS_GOOD
                    ))                 
            PackagingAllocation.objects.bulk_create(allocations)
            if documents:
                ConsignmentDocument.objects.bulk_create(documents)
            
            return StandardResponse(status=200, message="Allocation completed.")
        
        except Exception as e:
            return StandardResponse(status=400, errors=[str(e) or "something went wrong"])
        
                    
class AdhocConsignmentDocumentAPI(APIView):
    
    def get(self, request, id=None, *args, **kwargs):
        
        if id != "list":
            try:
                document = ConsignmentDocument.objects.get(id=id)
            except:
                return StandardResponse(status=400, success=False, errors=["Document does not exists"])
            
            return StandardResponse(status=200, data=AdhocConsignmentDocumentSerializer(document).data)
        else:
            try:
                consignment = Consignment.objects.get(consignment_id=request.GET.get("consignment_id"))
            except:
                return StandardResponse(status=400, success=False, errors=["consignment does not exists"])
            
            documents = ConsignmentDocument.objects.select_related("adhoc_line").filter(consignment=consignment)
            
            return StandardResponse(status=200, data=AdhocConsignmentDocumentSerializer(documents, many=True).data)
    
    
    @transaction.atomic
    def post(self, request, *args, **kwargs):
        try:
            consignment = Consignment.objects.get(consignment_id=request.data.get("consignment_id"))
        except:
            return StandardResponse(status=400, success=False, errors=["consignment does not exists"])

        adhoc_line_id = request.data.get("adhoc_line")        
        document_type = request.data.get("document_type")
        attachments = [file for key, file_list in request.FILES.lists() if key.startswith("attachments") for file in file_list]
         
        try:
            document_obj, _ = ConsignmentDocument.objects.get_or_create(consignment=consignment, document_type=document_type, adhoc_line_id=adhoc_line_id)
            files_to_create = [
                ConsignmentDocumentAttachment(document=document_obj, file=file)
                for file in attachments
            ]                
            ConsignmentDocumentAttachment.objects.bulk_create(files_to_create)
            
        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=[str(e)])
        
        return StandardResponse(status=201, message="Documents Added Successfully.")
    
    
    @transaction.atomic
    def put(self, request, id=None, *args, **kwargs):
        try:
            document = ConsignmentDocument.objects.get(id=id)
        
            attachments_to_exclude = request.data.get("attachment_urls")
            
            files = [file for key, file_list in request.FILES.lists() if key.startswith("attachments") for file in file_list]

            if attachments_to_exclude:
                existing_attachments = ConsignmentDocumentAttachment.objects.filter(document=document)

                for attachment in existing_attachments:
                    if "/media/"+attachment.file.name not in attachments_to_exclude.split(","):
                        if attachment.file:
                            default_storage.delete(attachment.file.path)
                        attachment.delete()
                        
            files_to_create = [
                ConsignmentDocumentAttachment(document=document, file=file)
                for file in files
            ]
            
            ConsignmentDocumentAttachment.objects.bulk_create(files_to_create)
            return StandardResponse(status=201, message="Documents Added Successfully.")
       
        except (ConsignmentDocument.DoesNotExist, ValidationError):
            return StandardResponse(status=400, success=False, errors=["Object not found"])

        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=[str(e)])
        
        
    @transaction.atomic
    def delete(self, request, id=None, *args, **kwargs):
        try:
            obj = ConsignmentDocument.objects.get(id=id)
        
            existing_attachments = ConsignmentDocumentAttachment.objects.filter(document=obj)
            for attachment in existing_attachments:
                if attachment.file:
                    default_storage.delete(attachment.file.path)
                attachment.delete()
                    
            obj.delete()
            return StandardResponse(status=200, message="Package deleted successfully.")
        except (ConsignmentDocument.DoesNotExist, ValidationError):
            return StandardResponse(status=400, success=False, errors=["Object not found"])
        
        except Exception as e:
            transaction.set_rollback(True)
            return StandardResponse(status=400, success=False, errors=[str(e)])
        
        
     
class ConsignmentUpload(APIView):
    parser_classes = [MultiPartParser, FormParser]
    
    box_declaration_fields = [
        "package_id","package_name", "package_type", "length", "width", "height","dimension_unit","weight","weight_unit", "description", "stackable"
    ]
    
    box_packing_fields = [
        "package_id", "po_line", "quantity", "hs_code", "eccn", "manufacturing_country", "is_dangerous_good"
    ]
    
    @role_required(Role.SUPPLIER_USER,OperationUserRole.L1,OperationUserRole.L2)
    def get(self, request, *args, **kwargs):
        po_ref_no = request.GET.get("po")
        try:
            po = PurchaseOrder.objects.select_related("storerkey").get(customer_reference_number=po_ref_no)
        except (PurchaseOrder.DoesNotExist, ValidationError):
            return StandardResponse(status=400, success=False, errors=["Purchase Order does not exists"])
        
        package_type = [choice[0] for choice in PackagingTypeChoices.choices]
        
        if po.storerkey.measurement_method == MeasurementTypeChoices.METRIC_SYSTEM:
            weight_units = METRIC_SYSTEM_WEIGHT_UNIT
            dimension_unit = METRIC_SYSTEM_DIMENSION_UNIT
        else:
            weight_units = IMPERIAL_SYSTEM_WEIGHT_UNIT
            dimension_unit = IMPERIAL_SYSTEM_DIMENSION_UNIT
            
        sheet_data = {
            "Box Declaration": {
                "fields": {
                    "PACKAGE ID": {"type": "text"},
                    "PACKAGE NAME" : {"type": "text"},
                    "PACKAGE TYPE": {"type": "list", "choices": package_type},
                    "LENGTH": {"type": "number"},
                    "WIDTH": {"type": "number"},
                    "HEIGHT": {"type": "number"},
                    "DIMENSION UNIT": {"type": "list", "choices": dimension_unit},
                    "WEIGHT": {"type": "number"},
                    "WEIGHT UNIT": {"type": "list", "choices": weight_units},
                    "DESCRIPTION": {"type": "text"},
                    "STACKABLE": {"type": "list", "choices": ["YES", "NO"]}
                }
            },
            "Box Packing": {
                "fields": {
                    "PACKAGE ID": {"type": "text"},
                    "PO LINE": {"type": "list", "choices": list(PurchaseOrderLine.objects.filter(purchase_order_id=po.id).exclude(open_quantity=0).values_list("customer_reference_number", flat=True))},
                    "QUANTITY": {"type": "number"},
                    "HS CODE": {"type": "list"},
                    "ECCN": {"type": "list", "choices": ["YES", "NO"]},
                    "MANUFACTURING COUNTRY": {"type": "list", "choices": list(DropDownValues.objects.filter(dropdown_name="ISO2").values_list("label", flat=True))},
                    "IS DANGEROUS GOOD": {"type": "list", "choices": ["YES", "NO"]},   
                }
            }
        }

        # sample_dir = "media/consignments/sample_file"
        # empty_directory(sample_dir)
        # filename = f"{sample_dir}/Consignment_for_{po_ref_no}.xlsx"
                
        sample_dir = os.path.join(settings.MEDIA_ROOT, "consignments/sample_file")
        empty_directory(sample_dir)
        filename = os.path.join(sample_dir, f"Consignment_for_{po_ref_no}.xlsx")
        excel_service = ExcelService()
        filename = excel_service.download_formatted_file(filename, fields={}, sheets_data=sheet_data, meta_data={"PO": po_ref_no})
        
        relative_path = os.path.relpath(filename, settings.MEDIA_ROOT)
        file_url = f"{settings.MEDIA_URL}{relative_path.replace(os.sep, '/')}"  # ensure forward slashes

        return StandardResponse(status=200, data=file_url)
        # return JsonResponse(data=filename, safe=False, status=200)
           
    
    def check_validations(self, wb):
        box_declaration_sheet = wb["Box Declaration"]
        bd_headers = [str(cell.value).strip().lower().replace(" ", "_") for cell in next(box_declaration_sheet.iter_rows(min_row=1, max_row=1))]
        missing_bd_headers = [field for field in self.box_declaration_fields if field not in bd_headers]

        box_packing_sheet = wb["Box Packing"]
        bp_headers = [str(cell.value).strip().lower().replace(" ", "_") for cell in next(box_packing_sheet.iter_rows(min_row=1, max_row=1))]
        missing_bp_headers = [field for field in self.box_packing_fields if field not in bp_headers]

        header_errors = []
        if missing_bd_headers:
            header_errors.append(f"Missing in 'Box Declaration': {', '.join(missing_bd_headers)}")
        if missing_bp_headers:
            header_errors.append(f"Missing in 'Box Packing': {', '.join(missing_bp_headers)}")

        if header_errors:
            return True, header_errors
        
        return False, None, box_declaration_sheet, bd_headers, box_packing_sheet, bp_headers
    
    
    def file_has_error(self, box_declaration_sheet, bd_headers, box_packing_sheet, bp_headers, po_lines):
        is_error = False
        box_declaration_data = []
        box_packing_data = []
        if "ERROR" not in bd_headers:
            error_col_index = len(bd_headers) + 1
            error_cell = box_declaration_sheet.cell(row=1, column=error_col_index, value="ERROR")
            error_cell.font = Font(bold=True)
        else:
            error_col_index = bd_headers.index("ERROR") + 1
        
        excel_package_ids = []
        excel_package_names = []    
        for i, row in enumerate(box_declaration_sheet.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in row[:len(bd_headers)]]  # exclude Error column
            if not any(values):
                break
            
            error_msg = ""
            if values[0] in excel_package_ids:
                error_msg = f"Duplicate package id found {values[0]}, "
            else:
                excel_package_ids.append(values[0])    
            
            if values[1] in excel_package_names:
                error_msg += f"Duplicate package name found {values[1]}, "
            else:
                excel_package_names.append(values[1])    
                
            has_empty = any(v is None or (isinstance(v, str) and not v.strip()) for v in values)

            if has_empty:
                error_msg = error_msg + "Missing value(s) in row, all columns are required."
                
            if error_msg:
                box_declaration_sheet.cell(row=i, column=error_col_index, value=error_msg)
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for col_index in range(1, len(bd_headers) + 1):
                    cell = box_declaration_sheet.cell(row=i, column=col_index)
                    cell.fill = red_fill
                is_error = True
                
            else:
                box_declaration_data.append(dict(zip(bd_headers, values)))
                
                
        if "ERROR" not in bp_headers:
            error_col_index = len(bp_headers) + 1
            error_cell = box_packing_sheet.cell(row=1, column=error_col_index, value="ERROR")
            error_cell.font = Font(bold=True)
        else:
            error_col_index = bp_headers.index("ERROR") + 1
            
        for i, row in enumerate(box_packing_sheet.iter_rows(min_row=2), start=2):
            values = [cell.value for cell in row[:len(bp_headers)]]  # exclude Error column
            if not any(values):
                break 
            
            error_msg = ""
            
            if values[0] not in excel_package_ids:
                error_msg = f"Box {values[0]} not configured in Box Decalartion sheet,"
                
            if values[1] not in po_lines:
                error_msg = error_msg + f"Invalid Po Line {values[1]}, "
                
            if values[2] > po_lines.get(values[1], {}).get("open_quantity", 0):
                error_msg = error_msg + f"{values[2]} must be less than open quantity "+ str(po_lines.get(values[1], {}).get("open_quantity")) + ", "
            
            has_empty = any(v is None or (isinstance(v, str) and not v.strip()) for v in values)

            if has_empty:
                error_msg = error_msg + "Missing value(s) in row, all columns are required."
            
            if error_msg:
                box_packing_sheet.cell(row=i, column=error_col_index, value=error_msg)
                red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                for col_index in range(1, len(bp_headers) + 1):
                    cell = box_packing_sheet.cell(row=i, column=col_index)
                    cell.fill = red_fill
                is_error = True
                
            else:
                box_packing_data.append(dict(zip(bp_headers, values)))
                
        return is_error, box_declaration_data, box_packing_data
    
    # @role_required(Role.SUPPLIER_USER,OperationUserRole.L1,OperationUserRole.L2)
    # @transaction.atomic()
    # def post(self, request, *args, **kwargs):
    #     file = request.FILES.get("file")
    #     if not file:
    #         return StandardResponse(errors=["File not found"], status=400, success=False)
    #     if file.size == 0:
    #         return StandardResponse(errors=["Excel file is empty"], status=400, success=False)

    #     try:
    #         wb = load_workbook(file, data_only=True)
    #     except InvalidFileException:
    #         return StandardResponse(errors=["Uploaded file is not a valid Excel file"], status=400, success=False)
    #     except Exception:
    #         return StandardResponse(errors=["Could not read Excel file"], status=400, success=False)

    #     if not wb.sheetnames:
    #         return StandardResponse(errors=["Excel file has no sheets"], status=400, success=False)
                
    #     try:
    #         po_ref_no = wb["Meta"]["B1"].value
    #     except:
    #         return StandardResponse(success=False, status=400, errors=["Please use latest downloaded excel file."])
        
    #     try:
    #         po = PurchaseOrder.objects.get(customer_reference_number=po_ref_no)
    #     except (PurchaseOrder.DoesNotExist, ValidationError):
    #         return StandardResponse(status=400, success=False, errors=["Purchase Order does not exists"])
        
    #     user_id = request.this_user.id
        
    #     consignment, created = ConsignmentStaging.objects.get_or_create(
    #         purchase_order=po,
    #         is_update=False,
    #         defaults={"user_id": user_id}
    #     )       
        
    #     client_data = Client.objects.filter(id=po.client.id).values().first()
    #     supplier_data = Supplier.objects.filter(id=po.supplier.id).values().first()
        
    #     if not created:
    #         if str(consignment.user_id) != str(user_id):
    #             return StandardResponse(status=400, success=False, errors=["Purchase Order already picked for consignment by other user"])
                    
    #     if "Box Declaration" not in wb.sheetnames or "Box Packing" not in wb.sheetnames:
    #         return StandardResponse(status=400, success=False, errors=["Please use latest downloaded excel file."])
        
    #     validation_error, validation_msg, box_declaration_sheet, bd_headers, box_packing_sheet, bp_headers = self.check_validations(wb)
        
    #     if validation_error:
    #         return StandardResponse(success=False, status=400, errors=[validation_msg])
        
    #     po_line_mapping = {
    #         i.get("customer_reference_number"): i
    #         for i in PurchaseOrderLine.objects.filter(purchase_order_id=po.id).values()
    #     }
        
    #     is_error, box_declaration_data, box_packing_data = self.file_has_error(box_declaration_sheet, bd_headers, box_packing_sheet, bp_headers, po_line_mapping)
                
    #     if is_error:   
    #         upload_dir = os.path.join(settings.MEDIA_ROOT, "consignments/upload/errors/")
    #         media_root = os.path.join(settings.MEDIA_URL, "consignments/upload/errors/")
    #         os.makedirs(upload_dir, exist_ok=True)  # Create if it doesn't exist

    #         filename = f"Consignment_for_{po_ref_no}.xlsx"

    #         file_path = os.path.join(upload_dir, filename)
            
    #         wb.save(file_path)
    #         return StandardResponse(status=400, success=False, data=media_root+filename)

    #     packages_to_create = []
    #     consignment_packages_to_create = []
    #     for box in box_declaration_data:

    #         existing_package = PackagingType.objects.filter(supplier = po.supplier, package_name = box.get("package_name")).first()

    #         if not existing_package:
    #             box_id = uuid4()
    #             package_data = {
    #                 "id": box_id,
    #                 "supplier": po.supplier,
    #                 "package_name": box.get("package_name"),
    #                 "package_type": box.get("package_type"),
    #                 "measurement_method": client_data.get("measurement_method"),
    #                 "description": box.get("description"),
    #                 # "weight": box.get("weight"),
    #                 # "weight_unit": box.get("weight_unit"),
    #                 "length": box.get("length"),
    #                 "width": box.get("width"),
    #                 "height": box.get("height"),
    #                 "dimension_unit": box.get("dimension_unit"),
    #                 "is_stackable": True if box.get("stackable") == "YES" else False, 
    #             }
    #             packages_to_create.append(PackagingType(**package_data))
            
    #         consignment_pckg_data = {
    #             "package_id": box.get("package_id"),
    #             "consignment": consignment,
    #             "packaging_type_id": box_id if not existing_package else existing_package.id,
    #             "weight": box.get("weight"),
    #             "weight_unit": box.get("weight_unit"),
    #         }
    #         consignment_packages_to_create.append(ConsignmentPackagingStaging(**consignment_pckg_data))
        
    #     PackagingType.objects.bulk_create(packages_to_create)
    #     consignment_package_ids = ConsignmentPackagingStaging.objects.bulk_create(consignment_packages_to_create)
    #     consignment_package_ids = [pkg.id for pkg in consignment_package_ids]
    #     consignment_package_mapping = {
    #         i.get("package_id"): i.get("id") 
    #         for i in ConsignmentPackagingStaging.objects.filter(id__in=consignment_package_ids).values("package_id", "id")
    #     }
                
    #     allocations_to_create = []
    #     for packing in box_packing_data:
    #         po_line = po_line_mapping.get(packing.get("po_line"))
    #         con_po_line, _ = ConsignmentPOLineStaging.objects.get_or_create(
    #             consignment=consignment,
    #             purchase_order_line_id=po_line.get("id"),
    #             defaults={
    #                "sku": po_line.get("sku"),
    #                 "po_ref": po_ref_no,
    #                 "po_line_ref": po_line.get("customer_reference_number"),
    #                 "hs_code": packing.get("hs_code"),
    #                 "eccn": True if packing.get("eccn") == "YES" else False,
    #                 "manufacturing_country": packing.get("manufacturing_country"),
    #                 "is_dangerous_good": True if packing.get("is_dangerous_good") == "YES" else False,
    #                 "qty_to_fulfill": 0,
    #                 "qty_packed": 0,
    #                 "qty_remaining": 0
    #             }
    #         )
            
    #         allocation_data = {
    #             "consignment_packaging_id": consignment_package_mapping.get(str(packing.get("package_id"))),
    #             "po_line": con_po_line,
    #             "allocated_qty": packing.get("quantity")
    #         }
    #         allocations_to_create.append(PackagingAllocationStaging(**allocation_data))
            
        
    #     PackagingAllocationStaging.objects.bulk_create(allocations_to_create)
            
    #     self.update_stage_packed_quantities(consignment, updating_all=True)
    #     self.update_stage_poline_packages(consignment)
        
    #     po = {
    #         "id": po.id,
    #         "reference_number": po.reference_number,
    #         "customer_reference_number": po.customer_reference_number,
    #     }
    #     return StandardResponse(status=200, data={"purchase_order": po, "client":client_data, "supplier": supplier_data})




class UserGridPreferencesAPI(APIView):

    def get(self,request):
        user = request.this_user
        grid_name = request.query_params.get("grid_name")

        if not grid_name:
            return StandardResponse(
                status=400,
                success=False,
                message="Missing required query parameter: grid_name"
            )

        try:
            obj = UserGridPreferences.objects.get(user=user, grid_name=grid_name)
        except UserGridPreferences.DoesNotExist:
            return StandardResponse(
                status=400,
                success=False,
                message="Grid preferences not found for the given grid name."
            )

        return StandardResponse(
            status=200,
            success=True,
            message="User's Grid Preferences retrieved successfully.",
            data={
            "id": obj.id,
            "grid_name": obj.grid_name,
            "order_list": obj.order_list 
        }
        )
        



    def post(self,request):
        serializer = UserGridPreferecesSerializer(data=request.data)
        
        if not serializer.is_valid():
            return StandardResponse(status=400, success=False, errors=serializer.errors)
        
        user = request.this_user
        grid_name = request.data.get("grid_name")
        order_list = request.data.get("order_list")
        obj,created = UserGridPreferences.objects.update_or_create(user = user,grid_name = grid_name,defaults={"order_list":order_list})

        message = "created" if created else "updated"
        return StandardResponse(
            status=201 if created else 200,
            message=f"User's Grid Preferences {message} successfully.",
            data={"id": obj.id}
        )