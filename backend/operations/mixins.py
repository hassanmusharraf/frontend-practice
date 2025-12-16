from django.db.models import Sum, F, Q, Count
from django.db import transaction
from .models import (
    # PackagingAllocationStaging,
    # ConsignmentPOLineStaging,
    # ConsignmentPackagingStaging,
    ConsignmentPackaging,
    PurchaseOrder,
    # ConsignmentStaging,
    PackagingAllocation,
    PurchaseOrderLine,
    Consignment
)
from portal.choices import PurchaseOrderStatusChoices, Role, POUploadStatusChoices, PackageStatusChoices
from django.utils import timezone
from datetime import timedelta
from core.response import StandardResponse
from portal.mixins import PaginationMixin
from django.db.models import Prefetch, Sum
from collections import defaultdict
from entities.models import Supplier,StorerKey, Client, MaterialMaster
from uuid import uuid4
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from django.db import transaction
from django.core.files import File
import os
from django.core.exceptions import ValidationError
from decimal import Decimal
from .services import PurchaseOrderService

# class ConsignmentStagingMixin:

    # def get_packaging_data(self, consignment, fields, po_line_id=None):

    #     base_query = PackagingAllocationStaging.objects.filter(
    #         consignment_packaging=OuterRef("id")
    #     )
    #     print('****',consignment)
    #     if po_line_id:
    #         base_query = base_query.filter(po_line_id=po_line_id)

    #     base_query = (
    #         base_query.exclude(allocated_qty=0)
    #         .select_related("po_line")
    #         .values("consignment_packaging")
    #         .annotate(
    #             json_agg=JSONBAgg(
    #                 JSONObject(
    #                     id="id",
    #                     allocated_qty="allocated_qty",
    #                     customer_reference_number="po_line__po_line_ref",
    #                     sku="po_line__sku",
    #                     description="po_line__purchase_order_line__description",
    #                     qty_packed="po_line__qty_packed",
    #                     is_dangerous_good="po_line__is_dangerous_good",
    #                     manufacturing_country="po_line__manufacturing_country",
    #                     eccn="po_line__eccn",
    #                 )
    #             )
    #         )
    #         .values("json_agg")
    #     )

    #     packages = (
    #         ConsignmentPackagingStaging.objects
    #         .filter(consignment=consignment)
    #         .select_related("packaging_type")
    #         .prefetch_related("allocations")
    #         .annotate(
    #             package_name=F("packaging_type__package_name"),
    #             package_type=F("packaging_type__package_type"),
    #             description=F("packaging_type__description"),
    #             is_stackable=F("packaging_type__is_stackable"),
    #             measurement_method=F("packaging_type__measurement_method"),
    #             length=F("packaging_type__length"),
    #             height=F("packaging_type__height"),
    #             width=F("packaging_type__width"),
    #             dimension_unit=F("packaging_type__dimension_unit"),
    #             weight=F("packaging_type__weight"),
    #             weight_unit=F("packaging_type__weight_unit"),
    #             po_lines=Subquery(base_query, output_field=JSONField())
    #         )
    #         .values(*fields)
    #         .order_by("package_id")
    #     )

    #     if po_line_id:

    #         qty_remaining = ConsignmentPOLineStaging.objects.filter(
    #             consignment=consignment,
    #             id=po_line_id
    #         ).values_list('qty_remaining', flat=True).first()
        

    #         qty_packed = ConsignmentPOLineStaging.objects.filter(
    #             consignment=consignment,
    #             id=po_line_id
    #         ).values_list('qty_packed', flat=True).first()

    #         return {
    #             "data": packages,
    #             "allocated_qty": qty_packed,
    #             "remaining_qty": qty_remaining
    #         }

    #     else:

    #         allocated_qty = PackagingAllocationStaging.objects.filter(
    #             consignment_packaging__consignment=consignment
    #         ).aggregate(total=Sum('allocated_qty'))['total'] or 0

    #         total_allocated_qty = ConsignmentPOLineStaging.objects.filter(
    #             consignment=consignment
    #         ).aggregate(total=Sum('qty_to_fulfill'))['total'] or 0

    #         remaining_qty = total_allocated_qty - allocated_qty

    #         return {
    #             "data": packages,
    #             "allocated_qty": allocated_qty,
    #             "remaining_qty": remaining_qty
    #         }
        

    # def get_packaging_data(self, consignment, fields, po_line_id=None):
    #     # Step 1: Get all packages for the consignment
    #     packages_qs = (
    #         ConsignmentPackagingStaging.objects
    #         .filter(consignment=consignment)
    #         .select_related("packaging_type")
    #         .prefetch_related(
    #             Prefetch(
    #                 "allocations",
    #                 queryset=PackagingAllocationStaging.objects
    #                 .select_related("po_line", "po_line__purchase_order_line")
    #                 .exclude(allocated_qty=0)
    #                 .filter(po_line_id=po_line_id) if po_line_id else
    #                 PackagingAllocationStaging.objects
    #                 .select_related("po_line", "po_line__purchase_order_line")
    #                 .exclude(allocated_qty=0)
    #             )
    #         )
    #         .order_by("package_id")
    #     )

    #     result_packages = []

    #     # Step 2: Build data in Python
    #     for pkg in packages_qs:
    #         po_lines = []

    #         for alloc in pkg.allocations.all():
    #             po_line = alloc.po_line
    #             base = po_line.purchase_order_line if hasattr(po_line, "purchase_order_line") else None
    #             po_lines.append({
    #                 "id": alloc.id,
    #                 "allocated_qty": alloc.allocated_qty,
    #                 "customer_reference_number": po_line.po_line_ref if po_line else None,
    #                 "sku": po_line.sku if po_line else None,
    #                 "description": base.description if base else None,
    #                 "qty_packed": po_line.qty_packed if po_line else None,
    #                 "is_dangerous_good": po_line.is_dangerous_good if po_line else None,
    #                 "manufacturing_country": po_line.manufacturing_country if po_line else None,
    #                 "eccn": po_line.eccn if po_line else None,
    #             })

    #         package_data = {
    #             "id": pkg.id,
    #             "po_lines": po_lines,
    #             "package_name": pkg.packaging_type.package_name if pkg.packaging_type else None,
    #             "package_type": pkg.packaging_type.package_type if pkg.packaging_type else None,
    #             "description": pkg.packaging_type.description if pkg.packaging_type else None,
    #             "is_stackable": pkg.packaging_type.is_stackable if pkg.packaging_type else None,
    #             "measurement_method": pkg.packaging_type.measurement_method if pkg.packaging_type else None,
    #             "length": pkg.packaging_type.length if pkg.packaging_type else None,
    #             "height": pkg.packaging_type.height if pkg.packaging_type else None,
    #             "width": pkg.packaging_type.width if pkg.packaging_type else None,
    #             "dimension_unit": pkg.packaging_type.dimension_unit if pkg.packaging_type else None,
    #             "weight": pkg.weight if pkg else None,
    #             "weight_unit": pkg.weight_unit if pkg else None,
    #             "package_id": pkg.package_id
    #         }

    #         result_packages.append({field: package_data.get(field) for field in fields})

    #     # Step 3: Calculate quantities
    #     if po_line_id:
    #         qty_remaining = (
    #             ConsignmentPOLineStaging.objects
    #             .filter(consignment=consignment, id=po_line_id)
    #             .values_list("qty_remaining", flat=True)
    #             .first()
    #         ) or 0

    #         qty_packed = (
    #             ConsignmentPOLineStaging.objects
    #             .filter(consignment=consignment, id=po_line_id)
    #             .values_list("qty_packed", flat=True)
    #             .first()
    #         ) or 0

    #         return {
    #             "data": result_packages,
    #             "allocated_qty": qty_packed,
    #             "remaining_qty": qty_remaining
    #         }

    #     else:
    #         allocated_qty = (
    #             PackagingAllocationStaging.objects
    #             .filter(consignment_packaging__consignment=consignment)
    #             .aggregate(total=Sum("allocated_qty"))["total"] or 0
    #         )

    #         total_allocated_qty = (
    #             ConsignmentPOLineStaging.objects
    #             .filter(consignment=consignment)
    #             .aggregate(total=Sum("qty_to_fulfill"))["total"] or 0
    #         )

    #         remaining_qty = total_allocated_qty - allocated_qty

    #         return {
    #             "data": result_packages,
    #             "allocated_qty": allocated_qty,
    #             "remaining_qty": remaining_qty
    #         }

    # def check_stage_po_consignment_exists(self, po, user, is_update=False, existing_consignment_id=None):
    #     try:
    #         po = PurchaseOrder.objects.get(customer_reference_number=po)
    #     except PurchaseOrder.DoesNotExist:
    #         return True, None, "Purchase order does not exist", None
            
    #     try:
    #         consignment = ConsignmentStaging.objects.get(purchase_order=po, user_id=user, is_update=is_update, existing_consignment_id=existing_consignment_id)
    #     except ConsignmentStaging.DoesNotExist:
    #         return True, None, "Staging Consignment does not exist", None

    #     return False, consignment, "", po

    # def update_stage_packed_quantities(self, consignment, updating_all=False):
    #     allocations = PackagingAllocationStaging.objects.filter(
    #         consignment_packaging__consignment=consignment
    #     ).values("po_line").annotate(total_allocated=Sum("allocated_qty"))
        
    #     for allocation in allocations:
    #         if allocation["po_line"]:
    #             po_line_id = allocation["po_line"]
    #             total_allocated = allocation["total_allocated"] or 0
            
    #             if not updating_all:
    #                 ConsignmentPOLineStaging.objects.filter(id=po_line_id, consignment=consignment).update(
    #                     qty_packed=total_allocated,
    #                     qty_remaining=F("qty_to_fulfill") - total_allocated
    #                 )
    #             else:
    #                 ConsignmentPOLineStaging.objects.filter(id=po_line_id, consignment=consignment).update(
    #                     qty_packed=total_allocated,
    #                     qty_to_fulfill=total_allocated,
    #                     qty_remaining=0
    #                 )

    # def update_stage_poline_packages(self, consignment):
    #     qs = PackagingAllocationStaging.objects.filter(
    #         consignment_packaging__consignment=consignment
    #     ).exclude(allocated_qty=0)

    #     # allocs = qs.values('po_line').annotate(
    #     #     packages=Count('consignment_packaging_id', distinct=True)
    #     # )

    #     raw_allocs = qs.values_list('po_line', 'consignment_packaging_id')
        
    #     mappings = defaultdict(set)
    #     for po_line_id, pkg_id in raw_allocs:
    #         mappings[po_line_id].add(str(pkg_id))

    #     updates = []
    #     for po_line_id, pkg_set in mappings.items():
    #         updates.append(
    #             ConsignmentPOLineStaging(id=po_line_id, packages=list(pkg_set))
    #         )

    #     with transaction.atomic():
    #         if updates:
    #             ConsignmentPOLineStaging.objects.bulk_update(
    #                 updates, ['packages']
    #             )
    #         else:
    #             # clear all if no allocations exist
    #             ConsignmentPOLineStaging.objects.filter(consignment=consignment).update(packages=[])

    # def delete_unused_stage_packages(self, consignment):
    #     # Delete allocations with allocated_qty <= 0 or po_line is None
    #     allocations_to_delete = PackagingAllocationStaging.objects.filter(
    #         consignment_packaging__consignment=consignment
    #     ).filter(Q(allocated_qty__lte=0) | Q(po_line__isnull=True))
    #     allocations_to_delete.delete()

    #     # Delete packages with no allocations left
    #     ConsignmentPackagingStaging.objects.filter(
    #         consignment=consignment
    #     ).annotate(
    #         allocation_count=Count('allocations')
    #     ).filter(allocation_count=0).delete()
    #     # ConsignmentPackagingStaging.objects.filter(
    #     #     consignment=consignment
    #     # ).annotate(
    #     #     total_allocated=Sum("allocations__allocated_qty")
    #     # ).filter(
    #     #     Q(total_allocated=0) | Q(total_allocated__isnull=True)
    #     # ).delete()


class ConsignmentMixin:
    
    def update_packages(self, consignment):
    #     consignment.packages = list(consignment.packagings.values_list("id", flat=True))     
    #     consignment.save()       

        consignment.packages = [
            str(id) for id in consignment.packagings.values_list("id", flat=True)
        ]
        consignment.save()

    def _tranform_object(self, data):

        mapping = {
        "purchase_order__customer_reference_number": "purchase_order",
        "adhoc__customer_reference_number": "adhoc",
        "supplier__name": "supplier",
        "client__name": "client",
        "consignor_address__address_name": "consignor_address",
        "delivery_address__address_name": "delivery_address",
        "console__console_id": "console_id",
        "freight_forwarder__name": "freight_forwarder"
        }

        for old_key, new_key in mapping.items():
            if old_key in data:  # Check existence
                data[new_key] = data.pop(old_key)  


class PurchaseOrderLineQuantityMixin:

    def update_quantities_on_delete_packaging(self, consignment):
        #adjust the quantities while deleting existing packaging

        allocations = PackagingAllocation.objects.filter(consignment_packaging__consignment=consignment).values("purchase_order_line_id").annotate(
            total_allocated=Sum("allocated_qty")
        )
        
        for allocation in allocations:
            if allocation["purchase_order_line_id"]:
                po_line_id = allocation["purchase_order_line_id"]
                total_allocated = allocation["total_allocated"] or 0
                PurchaseOrderLine.objects.filter(id=po_line_id).update(
                    processed_quantity=F("processed_quantity") - total_allocated,
                    open_quantity=F("open_quantity") + total_allocated
                )
         
    def when_create_consignment(self, consignment):
        allocations = PackagingAllocation.objects.filter(consignment_packaging__consignment=consignment).values("purchase_order_line_id").annotate(
            total_allocated=Sum("allocated_qty")
        )
        
        for allocation in allocations:
            po_line_id = allocation["purchase_order_line_id"]
            total_allocated = allocation["total_allocated"] or 0
                        
            PurchaseOrderLine.objects.filter(id=po_line_id).update(
                processed_quantity=F("processed_quantity") + total_allocated,
                open_quantity=F("open_quantity") - total_allocated
            )  
    
    def when_cancel_consignment(self, consignment):
        # allocations = PackagingAllocation.objects.filter(consignment_packaging__consignment=consignment).values("purchase_order_line_id").annotate(
        #     total_allocated=Sum("allocated_qty")
        # )
        
        # for allocation in allocations:
        #     # if allocation["purchase_order_line_id"] == None:
        #     po_line_id = allocation["purchase_order_line_id"]
        #     total_allocated = allocation["total_allocated"] or 0
                        
        #     PurchaseOrderLine.objects.filter(id=po_line_id).update(
        #         processed_quantity=F("processed_quantity") - total_allocated,
        #         open_quantity=F("open_quantity") + total_allocated
        #     )   
        po_lines = PurchaseOrderLine.objects.filter(
           packaging_allocations__consignment_packaging__consignment=consignment
        ).annotate(
            total_allocated=Sum('packaging_allocations__allocated_qty')
        ).distinct()
 
        # Bulk update all affected PO lines in a single query
        PurchaseOrderLine.objects.bulk_update(
            [
                PurchaseOrderLine(
                    id=po_line.id,
                    processed_quantity=po_line.processed_quantity - (po_line.total_allocated or 0),
                    open_quantity=po_line.open_quantity + (po_line.total_allocated or 0)
                )
                for po_line in po_lines
            ],
            fields=['processed_quantity', 'open_quantity']
        )
    
    def when_delivered_consignment(self, consignment):
        
        po_lines = consignment.purchase_order_lines.all()

        allocations = PackagingAllocation.objects.filter(
            consignment_packaging__consignment=consignment
        ).values("purchase_order_line_id").annotate(
            total_allocated=Sum("allocated_qty")
        )


        # Index by ID for fast access
        po_line_map = {line.id: line for line in po_lines}

        lines_to_update = []

        for allocation in allocations:
            po_line_id = allocation["purchase_order_line_id"]
            total_allocated = allocation["total_allocated"] or Decimal("0")

            if not po_line_id or po_line_id not in po_line_map:
                continue

            po_line = po_line_map[po_line_id]

            # Calculate new values
            new_processed_qty = po_line.processed_quantity - total_allocated
            new_fulfilled_qty = po_line.fulfilled_quantity + total_allocated
            new_open_qty = po_line.quantity - new_fulfilled_qty - new_processed_qty

            # Validate all rules
            errors = []

            if new_processed_qty < 0:
                errors.append(f"Processed quantity for PO Line {po_line_id} would go negative.")

            if new_fulfilled_qty > po_line.quantity:
                errors.append(f"Fulfilled quantity for PO Line {po_line_id} exceeds line quantity.")

            if new_open_qty < 0:
                errors.append(f"Open quantity for PO Line {po_line_id} would go negative.")

            if errors:
                raise ValidationError(errors)

            # Apply updated values to in-memory instance
            po_line.processed_quantity = new_processed_qty
            po_line.fulfilled_quantity = new_fulfilled_qty
            po_line.open_quantity = new_open_qty

            lines_to_update.append(po_line)

        if lines_to_update:
            # Use safe bulk update
            PurchaseOrderLine.objects.bulk_update(
                lines_to_update,
                ["processed_quantity", "fulfilled_quantity", "open_quantity"]
            )
        
        ConsignmentPackaging.objects.filter(
            allocations__id__in=allocations.values_list("id",flat=True)).update(
            status=PackageStatusChoices.DELIVERED
        )
                    
    def change_status(self, purchase_orders):
        purchase_order_ids = [po.id for po in purchase_orders]

        # Update PARTIALLY_FULFILLED lines in one query
        PurchaseOrderLine.objects.filter(
            purchase_order_id__in=purchase_order_ids
        ).exclude(fulfilled_quantity=0).update(
            status=PurchaseOrderStatusChoices.PARTIALLY_FULFILLED
        )

        # Update CLOSED lines in one query
        PurchaseOrderLine.objects.filter(
            purchase_order_id__in=purchase_order_ids,
            open_quantity=0,
            processed_quantity=0
        ).update(
            status=PurchaseOrderStatusChoices.CLOSED
        )
             

class AdhocPurchaseOrderLineMixin:
    def check_and_delete_related_adhoc_lines(self, package):
        allocations = PackagingAllocation.objects.filter(consignment_packaging=package)

        adhoc_lines = [allocation.adhoc_line for allocation in allocations]
        if adhoc_lines:
            with transaction.atomic():
                try:
                    for adhoc_line in adhoc_lines:
                            if PackagingAllocation.objects.filter(adhoc_line=adhoc_line ).count() > 1:
                                continue
                            adhoc_line.delete() 
                except Exception as e:
                    transaction.set_rollback(True)
                    return  str(e)
            return  None
        else:
            return  None
        
class ConsignmentStatusSummary(PaginationMixin):

    def consignment_status_summary(self,request, min_days=None, max_days=None):
        now = timezone.now()

        fields = ["id", "consignment_id","packages", "actual_pickup_datetime", "requested_pickup_datetime",
                "consignment_status", "created_at", "is_completed", "type", "purchase_order",
                "adhoc", "supplier__name", "client__name", "consignor_address__address_name", "delivery_address__address_name",
                "console__console_id", "freight_forwarder__name"]
        
        pg = request.GET.get("pg") or 0
        limit = request.GET.get("limit") or 25
        status = request.GET.get("status", "")
        
        filters = {}
        if max_days is not None:
            filters["created_at__gte"] = now - timedelta(days=max_days)
        if min_days is not None:
            filters["created_at__lt"] = now - timedelta(days=min_days)
        if status and status != "all":
            filters["consignment_status"] = status

        queryset =  Consignment.objects.filter(**filters).select_related("freight_forwarder","console","supplier", "client", "purchase_order", "adhoc","consignor_address","delivery_address").values(*fields).order_by("-consignment_id")

        if not queryset.exists():
            return StandardResponse(success=True, data=[], count=0, status=200)
        
        count = queryset.count()
        paginate_result = self.paginate_results(queryset, pg, limit)

        return paginate_result, count
        

class PurchaseOrderMixin:

    def parse_excel_to_pos(self,wb):
        header_sheet = wb["PO Header"]
        line_sheet = wb["PO Lines"]

        # Convert headers to keys
        headers = [str(c.value).strip().lower().replace(" ", "_") for c in next(header_sheet.iter_rows(min_row=1, max_row=1))]
        lines_headers = [str(c.value).strip().lower().replace(" ", "_") for c in next(line_sheet.iter_rows(min_row=1, max_row=1))]

        lines_by_po = {}
        for row in line_sheet.iter_rows(min_row=2, values_only=True):
            line = dict(zip(lines_headers, row))
            crn = line.get("purchase_order_crn")
            if crn:
                lines_by_po.setdefault(crn, []).append(line)

        po_entries = []
        for row in header_sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue  # Skip blank row
            record = dict(zip(headers, row))
            crn = record.get("customer_reference_number")
            record["pieces_detail"] = lines_by_po.get(crn, [])
            record["seller_details"] = {"seller_code": record.get("supplier_code")}
            record["buyer_details"] = {"buyer_code": record.get("buyer_code")}
            po_entries.append(record)

        return po_entries
    
    def line_related_consignments(self, line, get_count = False):

        consignments = set()
        package_list = line.packaging_allocations.all()
        for package in package_list:
            consignment_packaging = package.consignment_packaging
            consignments.add(consignment_packaging.consignment.id)

        if get_count:
            return len(consignments)

        return consignments

    def enrich_party_details(self,supplier_code, buyer_code):
        try:
            supplier = Supplier.objects.select_related("client").get(supplier_code=supplier_code)
        except Supplier.DoesNotExist:
            return None, None, f"Supplier with code '{supplier_code}' not found"

        client = supplier.client
        if client.client_code != buyer_code:
            return None, None, "Seller Code and Buyer Code do not match"
        
        return supplier, client, ""

    def validate_po_entries(self,po_entries):

        validation_errors = []
        for po_data in po_entries:
            seller = po_data.get("seller_details") or {}
            buyer = po_data.get("buyer_details") or {}
            po_lines = po_data.get("pieces_detail", [])
            po_crn = po_data.get("customer_reference_number")
            error_base = f"For PO '{po_crn}' "

            required_po_fields = ["plant_id", "center_code", "storer_key", "customer_reference_number","reference_number"]
            required_seller_fields = ["seller_code"]
            required_buyer_fields = ["buyer_code"]
            required_po_line_fields = ["customer_reference_number","reference_number","quantity"]

            for field in required_po_fields:
                if not po_data.get(field):
                    validation_errors.append(f"{error_base}Field '{field}' is required")

            for field in required_seller_fields:
                if not seller.get(field):
                    validation_errors.append(f"{error_base}Field '{field}' is required")

            for field in required_buyer_fields:
                if not buyer.get(field):
                    validation_errors.append(f"{error_base}Field '{field}' is required under buyer details")

            for field in required_po_line_fields:
                for line in po_lines:
                    if not line.get(field):
                        validation_errors.append(f"{error_base}Field '{field}' is required under PO Line")

            if not po_lines:
                validation_errors.append(f"At least one PO line is required")
            
            po_crns = [po.get("customer_reference_number") for po in po_entries]
            for crn in set(po_crns):
                if po_crns.count(crn) > 1:
                    validation_errors.append(f"Duplicate customer reference number '{crn}'")

            po_line_crns = [line.get("customer_reference_number") for line in po_lines]
            for crn in set(po_line_crns):
                if po_line_crns.count(crn) > 1:
                    validation_errors.append(f"{error_base}Duplicate Po line customer reference number '{crn}'")

        return validation_errors
    
    @transaction.atomic()
    def create_purchase_orders(self,data,is_for_sheet = False):
        po_to_create = []
        po_lines_to_create = []
        error_for_sheet = {'PO': {}, 'PO-Line': {}}

        ## This errors is used for both File Upload and for Import API
        ## That's the prefix of the error is different

        def add_error(po_crn, message, is_line=False):
            category = 'PO-Line' if is_line else 'PO'
            # intro = f"For {category} {po_crn}': " if not is_for_sheet else ""
            # message = intro + message

            container = error_for_sheet[category]
            if po_crn in container:
                container[po_crn] += f" | {message}"
            else:
                container[po_crn] = message




        for po_data in data:
            supplier_data = po_data.get("seller_details", {})
            client_data = po_data.get("buyer_details", {})
            po_lines = po_data.get("pieces_detail", [])
            buyer_code = client_data.get("buyer_code")
            po_crn = po_data.get("customer_reference_number")

            try:
                storerkey_object = StorerKey.objects.get(storerkey_code=po_data.get("storer_key"))
                supplier_object = Supplier.objects.select_related("client").get(supplier_code=supplier_data.get("seller_code"))
                provided_client = Client.objects.get(client_code = buyer_code)
                client_object = supplier_object.client
                # supplier_object, client_object = self.enrich_party_details(supplier_data, buyer_code)

            except ValueError as e:
                add_error(po_crn, str(e))
                continue
            except StorerKey.DoesNotExist as e:
                add_error(po_crn,"Storerkey not found.")
                continue
            except Supplier.DoesNotExist:
                add_error(po_crn, "Supplier not found.")
                continue
            
            if client_object != provided_client:
                add_error(po_crn, "Seller code and buyer code are not linked.")

            if storerkey_object.client.id != provided_client.id:
                add_error(po_crn,"Storer Key is not linked to the client.")

            if not supplier_object.storerkeys.filter(id=storerkey_object.id).exists():
                add_error(po_crn,"Storer Key is not linked to the supplier.")

            if PurchaseOrder.objects.filter(customer_reference_number=po_crn).exists():
                add_error(po_crn,f"Purchase Order already exists.")

            po_line_cust_ref_numbers = [i.get("customer_reference_number") for i in po_lines]
            duplicates = [crm for crm in set(po_line_cust_ref_numbers) if po_line_cust_ref_numbers.count(crm) > 1]
            if duplicates:
                add_error(po_crn, [f"Duplicate customer reference number '{crm}' in the '{po_crn}'." for crm in duplicates])

            # if error_for_sheet['PO'] or error_for_sheet['PO-Line']:
            #     continue

            po_id = uuid4()
            po_data.update({
                "id": po_id,
                "client": client_object,
                "supplier": supplier_object,
                "storerkey": storerkey_object
            })
            po_data.update(self.seller_details(supplier_data))

            del po_data["storer_key"]
            del po_data["seller_details"]
            del po_data["pieces_detail"]

            if "supplier_code" in po_data:
                del po_data["supplier_code"]
            if "buyer_code" in po_data:
                del po_data["buyer_code"]
            if "quantity" in po_data:
                po_data["open_quantity"] = po_data["quantity"]
                del po_data["quantity"]
            
            if not error_for_sheet['PO'] or not error_for_sheet['PO-Line']:
                po_to_create.append(PurchaseOrder(**po_data))

            
            for po_line in po_lines:
                
                if not self.check_material_exists(po_line.get("product_code"), storerkey_object.storerkey_code, po_data.get("group_code")):
                    error_message = f"Material {po_line.get('product_code')} does not exist for PO : {po_crn} | PO-Line : {po_line['customer_reference_number']}"
                    add_error(po_crn, error_message, is_line=True)

                try:
                    if "purchase_order_crn" in po_line:
                        del po_line["purchase_order_crn"]

                    if "chemical" in po_line:
                        po_line["is_chemical"] = po_line["chemical"] if po_line["chemical"] else False
                        del po_line["chemical"]

                    if "dangerous_good" in po_line:
                        po_line["is_dangerous_good"] = po_line["dangerous_good"] if po_line["dangerous_good"] else False
                        del po_line["dangerous_good"]
                    
                    po_line["purchase_order_id"] = po_id
                    po_line["open_quantity"] = po_line.get("quantity")
                    po_lines_to_create.append(PurchaseOrderLine(**po_line))
                
                except Exception as e:
                    add_error(po_line["customer_reference_number"], str(e), is_line=True)

        if error_for_sheet['PO'] or error_for_sheet['PO-Line']:
            return None, error_for_sheet
        

        try:
            pos = PurchaseOrder.objects.bulk_create(po_to_create)
        except Exception as e:
            add_error(None, str(e))
            return None, error_for_sheet

        try:
            PurchaseOrderLine.objects.bulk_create(po_lines_to_create)
        except Exception as e:
            add_error(None, str(e), is_line=True)
            return None, error_for_sheet

        PurchaseOrderService.update_open_quantity(pos)
        # for po in pos:
        #     po.update_quantity()

        return pos, None
    
    def process_excel_file(self,fileupload_obj):
        try:
            file = fileupload_obj.uploaded_file
            wb = load_workbook(file, data_only=True)
            po_entries = self.parse_excel_to_pos(wb)

            po_header_sheet = wb["PO Header"]
            poh_headers = [str(cell.value).strip().lower().replace(" ", "_") for cell in next(po_header_sheet.iter_rows(min_row=1, max_row=1))]
            
            po_lines_sheet = wb["PO Lines"]
            pol_headers = [str(cell.value).strip().lower().replace(" ", "_") for cell in next(po_lines_sheet.iter_rows(min_row=1, max_row=1))]
                
            error_col_index = len(poh_headers) + 1
            error_cell = po_header_sheet.cell(row=1, column=error_col_index, value="ERROR")

            error_col_index_pol_headers = len(pol_headers) + 1
            error_cell = po_lines_sheet.cell(row=1, column=error_col_index_pol_headers, value="ERROR")
            

            error_cell.font = Font(bold=True)

            def find_row_index(sheet, search_value, column_ind):
                if search_value:
                    for row in sheet.iter_rows(min_row=2, min_col=1, max_col=column_ind):  # skip header
                        for cell in row:
                            if cell.value == search_value:
                                return cell.row
                return 2  # Not found

            def update_error_in_file(errors,sheet,error_col_index):
                
                for key,value in errors.items():
                    row = find_row_index(sheet,key,22)
                    sheet.cell(row = row, column=error_col_index, value=value)
                    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    for col_index in range(1, len(poh_headers) + 1):
                        cell = sheet.cell(row=row, column=col_index)
                        cell.fill = red_fill

                self.save_error_file(fileupload_obj,wb)
                return False
            
            pos, errors = self.create_purchase_orders(po_entries,True)

            if errors:
                
                if errors["PO"] :
                    update_error_in_file(errors["PO"],po_header_sheet,error_col_index)
                
                if errors["PO-Line"]:
                    update_error_in_file(errors["PO-Line"],po_lines_sheet,error_col_index_pol_headers)

                return False
            
            fileupload_obj.status = POUploadStatusChoices.SUCCESS
            fileupload_obj.save()
                
            return True
        
        except Exception as e:
            unexpected_error_cell = po_header_sheet.cell(row=1, column=error_col_index+1, value="UNEXPECTED ERROR")
            unexpected_error_cell.font = Font(bold=True)
            po_header_sheet.cell(row = 2, column=error_col_index+1, value=str(e))
            self.save_error_file(fileupload_obj,wb)

    def save_error_file(self,fileupload_obj,wb):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"Purchase_Order_Errors_{timestamp}.xlsx"  
        error_dir = os.path.join("media", "po", "upload", "errors")
        if not os.path.exists(error_dir):
            os.makedirs(error_dir, exist_ok=True)
        filename = os.path.join(error_dir, file_name)
        wb.save(filename)
        
        with open(filename, "rb") as f:
            fileupload_obj.error_file.save(file_name, File(f), save=False)

        fileupload_obj.status = POUploadStatusChoices.ERROR
        fileupload_obj.save()

    def seller_details(self,seller_data):
        data_to_sent = {}
        fields = ["address_line_1", "address_line_2", "city", "state", "country", "postal_code", "phone_number", "tax_number", "email"]
        for f in fields:
            data_to_sent["seller_" + f] = seller_data.get(f, "") 
        return data_to_sent

    def check_material_exists(self, product_code, storerkey, hub):
        return MaterialMaster.objects.filter(product_code=product_code, storerkey__storerkey_code=storerkey, hub__hub_code=hub).exists()

class FilterMixin:

    def build_filter(self,user,model,filters=None):

        data = {}
        role = user.role
        profile = user.profile()
        storerkey_ids = self.get_storerkeys(profile) if profile else []

        # Base data for logging or future use
        # data = self.get_base_data(user, profile)

        if not storerkey_ids:
            return filters
        
        if isinstance(filters, Q):
            filters &= self.get_model_q_filters(model, storerkey_ids)

            ## Also add the same filters below for dict filters
            if role == Role.SUPPLIER_USER and model == "PO":
                filters &= Q(supplier = profile.supplier)

            if role == Role.SUPPLIER_USER and model == "Consignment":
                filters &= Q(supplier = profile.supplier)

            if role == Role.SUPPLIER_USER and model == "PO-Line":
                filters &= Q(purchase_order__supplier = profile.supplier)

            if role == Role.CLIENT_USER and model == "PO":
                filters &= Q(client = profile.client)

            if role == Role.CLIENT_USER and model == "Consignment":
                filters &= Q(client = profile.client)

        elif isinstance(filters, dict):
            filters.update(self.get_model_dict_filters(model, storerkey_ids))

            if role == Role.SUPPLIER_USER and model == "PO":
                filters["supplier"] = profile.supplier

            if role == Role.SUPPLIER_USER and model == "Consignment":
                filters["supplier"] = profile.supplier

            if role == Role.SUPPLIER_USER and model == "PO-Line":
                filters["purchase_order__supplier"] = profile.supplier

            if role == Role.CLIENT_USER and model == "PO":
                filters["client"] = profile.client

            if role == Role.CLIENT_USER and model == "Consignment":
                filters["client"] = profile.client
            
            
        return filters
    

    def apply_filter(self, qs, filters):
        """Applies the filter to a queryset."""
        if filters and qs:
            if isinstance(filters, Q):
                return qs.filter(filters)
            elif isinstance(filters, dict):
                return qs.filter(**filters)
        return qs

    def get_model_q_filters(self, model, storerkey_ids):
        """Returns Q filters based on model and storerkeys."""
        if model == "PO":
            return Q(storerkey__in=storerkey_ids)
        elif model == "Consignment":
            # return Q(purchase_order_lines__purchase_order__storerkey__in=storerkey_ids) | Q(adhoc__storerkey__in=storerkey_ids)
            return Q(purchase_order_lines__purchase_order__storerkey__in=storerkey_ids)

        # Add more model conditions here
        return Q()
    
    def get_model_dict_filters(self, model, storerkey_ids):
        """Returns dict filters based on model and storerkeys."""
        filter_dict = {}
        if model == "PO":
            filter_dict["storerkey__in"] = storerkey_ids
        elif model == "Consignment":
            filter_dict["purchase_order__storerkey__in"] = storerkey_ids
            # filter_dict["adhoc__storerkey__in"] = storerkey_ids
        # Add more model conditions here
        return filter_dict
    
    def get_storerkeys(self, profile):
        """Extracts storerkey IDs from the profile."""
        return list(profile.storerkeys.values_list("id", flat=True)) if profile else []
    
    def get_base_data(self, user, profile):
        """Optional helper to extract IDs based on user role for auditing/debugging."""
        if not profile:
            return {}

        if user.role == Role.SUPPLIER_USER:
            return {
                "supplier_id": profile.id,
                "supplier_user_id": profile.supplier.id
            }

        elif user.role == Role.CLIENT_USER:
            return {
                "client_id": profile.id,
                "client_user_id": profile.client.id
            }

        elif user.role == Role.OPERATIONS:
            return {
                "operations_id": profile.id,
                "access_level": profile.access_level
            }

        return {}
    
