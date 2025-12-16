from core.response import StandardResponse, ServiceError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from portal.choices import POImportFormatsChoices, POUploadStatusChoices, PurchaseOrderStatusChoices, OrderTypeChoices
import os
import csv
from io import StringIO, BytesIO
from django.db import transaction
from datetime import datetime
from operations.models import PurchaseOrder, PurchaseOrderLine, ConsignmentPOLine
from operations.utils import parse_any_date
from operations.notifications import NotificationService
from entities.models import Supplier,StorerKey, MaterialMaster
from uuid import uuid4
from django.core.files import File
from decimal import Decimal
from ..services import POLineService, PurchaseOrderService
from collections import defaultdict, Counter
from portal.models import CostCenterCode
class POImportService:

    @classmethod
    def standard_json_format(cls):

        return {
            "reference_number": "",
            "customer_reference_number": "",
            "supplier_code": "",
            "storer_key": "",
            "inco_terms": "",
            "quantity": "",
            "buyer_code": "",
            "plant_id": "",
            "center_code": "",
            "description": "",
            "group_code": "",
            "type": "",
            "notes": "",
            "payment_terms": "",
            "destination_country": "",
            "origin_country": "",
            "order_date": "",
            "expected_delivery_date": "",
            "order_due_date": "",

            "pieces_detail": [
                {
                    "purchase_order_crn": "",
                    "reference_number": "",
                    "customer_reference_number": "",
                    "sku": "",
                    "quantity": "",
                    "description": "",
                    "product_code": "",
                    "hs_code": "",
                    "alternate_unit": "",
                    "stock_number": "",
                    "chemical": "",
                    "dangerous_good": "",
                    "unit_price": "",
                    "unit_cost": "",
                    "weight": "",
                    "volume": "",
                    "length": "",
                    "width": "",
                    "height": ""
                }
            ],

            "seller_details": {
                "seller_code": ""
            },

            "buyer_details": {
                "buyer_code": ""
            }
        }

    
    @classmethod
    def parse_excel_to_pos_standard_format(cls,wb):

        header_sheet = wb["PO Header"]
        line_sheet = wb["PO Lines"]

        # Convert headers to keys
        headers = [str(c.value).strip().lower().replace(" ", "_") for c in next(header_sheet.iter_rows(min_row=1, max_row=1))]
        lines_headers = [str(c.value).strip().lower().replace(" ", "_") for c in next(line_sheet.iter_rows(min_row=1, max_row=1))]

        lines_by_po = {}
        for row in line_sheet.iter_rows(min_row=2, values_only=True):
            line = dict(zip(lines_headers, row))
            crn = line.get("purchase_order_crn")
            if crn:
                lines_by_po.setdefault(crn, []).append(line)

        po_entries = []
        for row in header_sheet.iter_rows(min_row=2, values_only=True):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue  # Skip blank row
            record = dict(zip(headers, row))
            crn = record.get("customer_reference_number")
            record["pieces_detail"] = lines_by_po.get(crn, [])
            record["seller_details"] = {"seller_code": record.get("supplier_code")}
            record["buyer_details"] = {"buyer_code": record.get("buyer_code")}
            po_entries.append(record)

        return po_entries


class POImportValidationService:

    @classmethod
    def validate_slb_pos(cls,po_entries):

        validation_errors = []
        for po_data in po_entries:
            seller = po_data.get("seller_details") or {}
            buyer = po_data.get("buyer_details") or {}
            po_lines = po_data.get("pieces_detail", [])
            po_crn = po_data.get("customer_reference_number")
            error_base = f"For PO '{po_crn}' "

            required_po_fields = ["plant_id", "center_code", "customer_reference_number"]
            required_seller_fields = ["seller_code"]
            required_buyer_fields = ["buyer_code"]
            required_po_line_fields = ["customer_reference_number","quantity"]

            for field in required_po_fields:
                if not po_data.get(field):
                    validation_errors.append(f"{error_base}Field '{field}' is required")

            for field in required_seller_fields:
                if not seller.get(field):
                    validation_errors.append(f"{error_base}Field '{field}' is required")

            for field in required_buyer_fields:
                if not buyer.get(field):
                    validation_errors.append(f"{error_base}Field '{field}' is required under buyer details")

            for field in required_po_line_fields:
                for line in po_lines:
                    if not line.get(field):
                        validation_errors.append(f"{error_base}Field '{field}' is required under PO Line")

            if not po_lines:
                validation_errors.append(f"At least one PO line is required")
            
            po_crns = [po.get("customer_reference_number") for po in po_entries]
            for crn in set(po_crns):
                if po_crns.count(crn) > 1:
                    validation_errors.append(f"Duplicate customer reference number '{crn}'")

            po_line_crns = [line.get("customer_reference_number") for line in po_lines]
            for crn in set(po_line_crns):
                if po_line_crns.count(crn) > 1:
                    validation_errors.append(f"{error_base}Duplicate Po line customer reference number '{crn}'")

        return validation_errors
    
    
    @staticmethod
    def validate_headers_for_slb(headers):

        required_headers = [
            "purchase_order_number", "po_line_number", "buyer_code", "supplier_number", "sap_plant_code"
        ]
        
        for header_name in required_headers:
            if header_name not in headers:
                return StandardResponse(status=400, success=False, errors=[f"Required headers : {required_headers} "])
        

    @staticmethod
    def standard_po_import_validations(wb):

        if "PO Header" not in wb.sheetnames or "PO Lines" not in wb.sheetnames:
            return StandardResponse(status=400, success=False, errors=["Please use latest downloaded excel file."])
    

    @classmethod
    def file_validations(cls, file, format=POImportFormatsChoices.PO):
    
        if not file:
            raise ServiceError(error="File not found")
            # return StandardResponse(errors=["File not found"], status=400, success=False)
        if file.size == 0:
            return StandardResponse(errors=["Excel file is empty"], status=400, success=False)

        try:
            
            ext = os.path.splitext(file.name)[1].lower()

            if ext == ".xlsx":
                wb = load_workbook(file, data_only=True)
                if not wb.sheetnames:
                    return StandardResponse(errors=["Excel file has no sheets"], status=400, success=False)
                
                header_sheet = wb.worksheets[0]   
                headers = [str(c.value).strip().lower().replace(" ", "_") for c in next(header_sheet.iter_rows(min_row=1, max_row=1))]
                
                
            elif ext == ".csv":
                file.seek(0)
                reader = csv.reader(file.read().decode("utf-8-sig").splitlines())
                headers = [h.strip().lower().replace(" ", "_") for h in next(reader)]


            else:
                raise ValueError("Unsupported file type. Only .xlsx or .csv allowed.")


            if format == POImportFormatsChoices.PO:
                POImportValidationService.standard_po_import_validations(headers)

            if format == POImportFormatsChoices.SLB:
                POImportValidationService.validate_headers_for_slb(headers)



        except InvalidFileException:
            return StandardResponse(errors=["Uploaded file is not a valid Excel file"], status=400, success=False)
        except Exception as e:
            return StandardResponse(errors=[f"Could not read Excel file : {str(e)}"], status=400, success=False)


class SLBPOImportService:

    @classmethod
    def parse_excel_to_pos(cls, file):
        ext = os.path.splitext(file.name)[1].lower()
        headers = []
        data = []

        # --- Read Excel file ---
        if ext == ".xlsx":
            wb = load_workbook(file, data_only=True)
            if not wb.sheetnames:
                return StandardResponse(errors=["Excel file has no sheets"], status=400, success=False)
            
            header_sheet = wb.worksheets[0]

            # Headers from first row
            headers = [
                str(c).strip().lower().replace(" ", "_") 
                for c in next(header_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            ]

            # Read remaining rows
            for row in header_sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                data.append(dict(zip(headers, row)))

        # --- Read CSV file ---
        elif ext == ".csv":
            file.seek(0)
            decoded = file.read().decode("utf-8-sig")
            reader = csv.reader(StringIO(decoded))

            # Headers
            headers = [h.strip().lower().replace(" ", "_") for h in next(reader)]

            # Rows
            for row in reader:
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                data.append(dict(zip(headers, row)))

        else:
            raise ValueError("Unsupported file type. Only .xlsx or .csv allowed.")

        # --- Process POs ---
        po_dict = {} 
        pol_dict = [] 

        for record in data:
            crn = record.get("purchase_order_number")
            if not crn:
                continue

            # Get or create PO

            if record.get("purchase_order_date", None):
                order_date = parse_any_date(record.get("purchase_order_date"))
                # order_date = datetime.strptime(record.get("purchase_order_date"), "%d-%m-%Y").strftime("%Y-%m-%d")

            if record.get("purchase_need/ship_date", None):
                expected_delivery_date = parse_any_date(record.get("purchase_need/ship_date"))
                # expected_delivery_date = datetime.strptime(record.get("purchase_need/ship_date"), "%d-%m-%Y").strftime("%Y-%m-%d")
            
            if record.get("purchase_due_date", None):
                order_due_date = parse_any_date(record.get("purchase_due_date"))
                # order_due_date = datetime.strptime(record.get("purchase_due_date"), "%d-%m-%Y").strftime("%Y-%m-%d")

            po = {
                "customer_reference_number": crn,
                "reference_number": crn,
                "supplier_code": record.get("supplier_number", ""),
                "order_date": order_date or "",
                "expected_delivery_date": expected_delivery_date or "",
                "order_due_date": order_due_date or "",
                "storerkey": record.get("buyer_code", ""),
                "inco_terms": record.get("header_inco_terms", "").split("|") if record.get("header_inco_terms") else [],
                "quantity": Decimal(str(record.get("quantity", 0.0))),
                "plant_id": record.get("sap_plant_code", ""),
                "center_code": record.get("ems_unit_level_1_name", ""),
                "sloc" : record.get("inventory_location", ""),
                "description": record.get("description", ""),
                "group_code": record.get("group_code", ""),
                "type": record.get("type", ""),
                "notes": record.get("notes", ""),
                "payment_terms": record.get("payment_terms", ""),
                "destination_country": record.get("destination_country", ""),
                "origin_country": record.get("origin_country", ""),
                "order_date": order_date or "",
                "expected_delivery_date": expected_delivery_date or "",
                "order_due_date": order_due_date or "",
                "pieces_detail": [],
                "order_type" : record.get("po_type", ""),    ## Order Type BTS, BTO or BOTH
                "seller_details": {"seller_code": record.get("supplier_number", "")},
                # "buyer_details": {"buyer_code": record.get("buyer_code", "")}
            }


            # Add piece line
            line = {
                "po": po,
                "reference_number": record.get("reference_number", ""),
                "customer_reference_number": record.get("po_line_number", ""),
                "product_code": record.get("part_number", ""),
                "sku": record.get("uom", ""),
                "quantity": Decimal(str(record.get("order_quantity", 0.0) or 0.0)),
                "open_quantity": Decimal(str(record.get("open_quantity", 0.0) or 0.0)),
                "description": record.get("part_description", ""),
                "inco_terms": record.get("line_incoterms","").split("|") if record.get("line_incoterms") else [],
                "hs_code": record.get("hs_code", ""),
                "alternate_unit": record.get("alternate_unit", ""),
                "stock_number": record.get("stock_number", ""),
                "is_chemical": record.get("chemical" "") or False ,
                "is_dangerous_good": record.get("dangerous_good", "") or False,
                "unit_price": Decimal(str(record.get("unit_price", 0.0) or 0.0)),
                "unit_cost": Decimal(str(record.get("item_unit_price", 0.0) or 0.0)),
                "weight": Decimal(str(record.get("weight", 0.0) or 0.0)),
                "volume": Decimal(str(record.get("volume", 0.0) or 0.0)),
                "length": Decimal(str(record.get("length", 0.0) or 0.0)),
                "width": Decimal(str(record.get("width", 0.0) or 0.0)),
                "height":Decimal(str(record.get("height", 0.0) or 0.0)),
            }

            pol_dict.append(line)

        # Return as a list
        return list(pol_dict)
    

    @classmethod
    def parse_excel_to_pos_standard_format(cls, file):
        ext = os.path.splitext(file.name)[1].lower()
        headers = []
        data = []

        # --- Read Excel file ---
        if ext == ".xlsx":
            wb = load_workbook(file, data_only=True)
            if not wb.sheetnames:
                return StandardResponse(errors=["Excel file has no sheets"], status=400, success=False)
            
            header_sheet = wb.worksheets[0]

            # Headers from first row
            headers = [
                str(c).strip().lower().replace(" ", "_") 
                for c in next(header_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            ]

            # Read remaining rows
            for row in header_sheet.iter_rows(min_row=2, values_only=True):
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                data.append(dict(zip(headers, row)))

        # --- Read CSV file ---
        elif ext == ".csv":
            file.seek(0)
            decoded = file.read().decode("utf-8-sig")
            reader = csv.reader(StringIO(decoded))

            # Headers
            headers = [h.strip().lower().replace(" ", "_") for h in next(reader)]

            # Rows
            for row in reader:
                if all(cell is None or str(cell).strip() == "" for cell in row):
                    continue
                data.append(dict(zip(headers, row)))

        else:
            raise ValueError("Unsupported file type. Only .xlsx or .csv allowed.")

        # --- Process POs ---
        po_dict = {}  

        for record in data:
            crn = record.get("purchase_order_number")
            if not crn:
                continue

            # Get or create PO
            if crn not in po_dict:
                if record.get("purchase_order_date", None):
                    order_date = datetime.strptime(record.get("purchase_order_date"), "%d-%m-%Y").strftime("%Y-%m-%d")

                if record.get("purchase_need/ship_date", None):
                    expected_delivery_date = datetime.strptime(record.get("purchase_need/ship_date"), "%d-%m-%Y").strftime("%Y-%m-%d")
                
                if record.get("purchase_due_date", None):
                    order_due_date = datetime.strptime(record.get("purchase_due_date"), "%d-%m-%Y").strftime("%Y-%m-%d")

                po_dict[crn] = {
                    "reference_number": record.get("reference_number", ""),
                    "customer_reference_number": crn,
                    "supplier_code": record.get("supplier_number", ""),
                    "storer_key": record.get("storer_key", ""),
                    "inco_terms": record.get("header_inco_terms", "").split("|") if record.get("header_inco_terms") else [],
                    "quantity": Decimal(str(record.get("quantity", 0.0))),
                    "buyer_code": record.get("buyer_code", ""),
                    "plant_id": record.get("sap_plant_code", ""),
                    "center_code": record.get("ems_unit_level_1_name", ""),
                    "description": record.get("description", ""),
                    "group_code": record.get("group_code", ""),
                    "type": record.get("type", ""),
                    "notes": record.get("notes", ""),
                    "payment_terms": record.get("payment_terms", ""),
                    "destination_country": record.get("destination_country", ""),
                    "origin_country": record.get("origin_country", ""),
                    "order_date": order_date or "",
                    "expected_delivery_date": expected_delivery_date or "",
                    "order_due_date": order_due_date or "",
                    "pieces_detail": [],
                    "seller_details": {"seller_code": record.get("supplier_number", "")},
                    "buyer_details": {"buyer_code": record.get("buyer_code", "")}
                }

            # Add piece line
            pieces_detail = {
                "purchase_order_crn": crn,
                "reference_number": record.get("reference_number", ""),
                "customer_reference_number": record.get("po_line_number", ""),
                "sku": record.get("uom", ""),
                "quantity": Decimal(str(record.get("order_quantity", 0.0) or 0.0)),
                "open_quantity": Decimal(str(record.get("open_quantity", 0.0) or 0.0)),
                "description": record.get("part_description", ""),
                "product_code": record.get("part_number", ""),
                "inco_terms": record.get("line_incoterms","").split("|") if record.get("line_incoterms") else [],
                "hs_code": record.get("hs_code", ""),
                "alternate_unit": record.get("alternate_unit", ""),
                "stock_number": record.get("stock_number", ""),
                "is_chemical": record.get("chemical" "") or False ,
                "is_dangerous_good": record.get("dangerous_good", "") or False,
                "unit_price": Decimal(str(record.get("unit_price", 0.0) or 0.0)),
                "unit_cost": Decimal(str(record.get("item_unit_price", 0.0) or 0.0)),
                "weight": Decimal(str(record.get("weight", 0.0) or 0.0)),
                "volume": Decimal(str(record.get("volume", 0.0) or 0.0)),
                "length": Decimal(str(record.get("length", 0.0) or 0.0)),
                "width": Decimal(str(record.get("width", 0.0) or 0.0)),
                "height":Decimal(str(record.get("height", 0.0) or 0.0)),
            }

            po_dict[crn]["pieces_detail"].append(pieces_detail)

        # Return as a list
        return list(po_dict.values())
        

    @staticmethod
    def check_material_exists(product_code, storerkey=[], hub=[]):
        return MaterialMaster.objects.filter(product_code=product_code, storerkey__storerkey_code=storerkey, hub__hub_code=hub).exists()
    

    @staticmethod
    def save_error_file(self,fileupload_obj,wb):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"Purchase_Order_Errors_{timestamp}.xlsx"  
        error_dir = os.path.join("media", "po", "upload", "errors")
        if not os.path.exists(error_dir):
            os.makedirs(error_dir, exist_ok=True)
        filename = os.path.join(error_dir, file_name)
        wb.save(filename)
        
        with open(filename, "rb") as f:
            fileupload_obj.error_file.save(file_name, File(f), save=False)

        fileupload_obj.status = POUploadStatusChoices.ERROR
        fileupload_obj.save()


    @staticmethod
    def seller_details(seller_data):
        data_to_sent = {}
        fields = ["address_line_1", "address_line_2", "city", "state", "country", "postal_code", "phone_number", "tax_number", "email"]
        for f in fields:
            data_to_sent["seller_" + f] = seller_data.get(f, "") 
        return data_to_sent


    @staticmethod
    def update_error_in_csv(errors, file_path, fileupload_obj):
        """
        Update CSV file rows with error messages and save a copy.
        """
        temp_rows = []
        headers = []

        # Read the CSV file
        with open(file_path, mode="r", encoding="utf-8-sig") as csvfile:
            reader = csv.reader(csvfile)
            headers = next(reader)
            if "error_message" not in headers:
                headers.append("error_message")  # Add error column if not present

            for row in reader:
                # Assuming key (used in errors dict) is in first column
                key = row[0]
                error_message = errors.get(key, "")
                if len(row) < len(headers):
                    row.extend([""] * (len(headers) - len(row)))  # ensure same length
                row[-1] = error_message  # put error in last column
                temp_rows.append(row)

        # Save updated file
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"Purchase_Order_Errors_{timestamp}.csv"
        error_dir = os.path.join("media", "po", "upload", "errors")
        os.makedirs(error_dir, exist_ok=True)
        file_with_path = os.path.join(error_dir, file_name)

        with open(file_with_path, mode="w", newline="", encoding="utf-8-sig") as outfile:
            writer = csv.writer(outfile)
            writer.writerow(headers)
            writer.writerows(temp_rows)

        # Save to fileupload model
        with open(file_with_path, "rb") as f:
            fileupload_obj.error_file.save(file_name, File(f), save=False)

        fileupload_obj.status = POUploadStatusChoices.ERROR
        fileupload_obj.save()
        return False
    

    @classmethod
    @transaction.atomic
    def create_purchase_orders(cls,lines_data):
        
        """
        Optimized & maintainable approach to create/update Purchase Orders and Lines.
        """

        required_po_fields = ["customer_reference_number", "storerkey", "supplier_code","plant_id","center_code","sloc"]
        required_line_fields = ["customer_reference_number", "quantity"]

        error_for_sheet = []                      # list of dicts {(po_crn, pol_crn): message}
        error_in_pos = []                # po_crn that have any error (skip them entirely)
        valid_po_groups = defaultdict(list)             # {po_crn: [line_dicts]}


        for idx, line in enumerate(lines_data, start=1):
            po = line.get("po") or {}
            po_crn = po.get("customer_reference_number")
            pol_crn = line.get("customer_reference_number")

            # helper for adding errors
            def add_error_local(po_key, pol_key, msg):
                error_for_sheet.append({(po_key or "UNKNOWN", pol_key or "UNKNOWN"): msg})
                error_in_pos.append(po_key)

            if not po_crn:
                add_error_local("UNKNOWN", pol_crn, "PO number missing.")
                continue

            # required PO-level fields (validate once per line; duplicates are fine)
            # (add_error_local(po_crn, pol_crn, f"Field '{field}' is required for PO") for field in required_po_fields if not po.get(field))
            for field in required_po_fields:
                if not po.get(field):
                    add_error_local(po_crn, pol_crn, f"Field '{field}' is required for PO")

            # required line-level fields
            # (add_error_local(po_crn, pol_crn, f"Field '{field}' is required in PO Line") for field in required_line_fields if not line.get(field) and line.get(field) != 0)
            for field in required_line_fields:
                if not line.get(field) and line.get(field) != 0:
                    add_error_local(po_crn, pol_crn, f"Field '{field}' is required in PO Line")

            valid_po_groups[po_crn].append(line)


        for po_crn, lines in valid_po_groups.items():
            if po_crn in error_in_pos:
                continue
            line_crns = [l.get("customer_reference_number") for l in lines]
            c = Counter(line_crns)
            for pol_crn, cnt in c.items():
                if cnt > 1:
                    error_for_sheet.append({(po_crn, pol_crn): f"PO '{po_crn}' has duplicate PO Line number '{pol_crn}'"})
                    error_in_pos.append(po_crn)


        valid_po_crns = [crn for crn in valid_po_groups.keys() if crn not in error_in_pos]
        if not valid_po_crns:
            return [], error_for_sheet
        

        all_po_crns = list(valid_po_groups.keys())
        
        existing_pos = {
            po.customer_reference_number: po
            for po in PurchaseOrder.objects.filter(customer_reference_number__in=all_po_crns)
            .select_related("client", "storerkey", "supplier")
        }

        existing_lines = {
            (l.purchase_order.customer_reference_number, l.customer_reference_number): l
            for l in PurchaseOrderLine.objects.filter(
                purchase_order__customer_reference_number__in=all_po_crns
            )
        }
        
        all_supplier_codes = {
            l["po"]["seller_details"]["seller_code"]
            for lines in valid_po_groups.values()
            for l in lines
            if l["po"].get("seller_details")
        }
        
        suppliers = {
            s.supplier_code: s
            for s in Supplier.objects.filter(supplier_code__in=all_supplier_codes)
        }

        all_storer_codes = {
            l["po"].get("storerkey")
            for lines in valid_po_groups.values()
            for l in lines
        }
        
        storerkeys = {
            s.storerkey_code: s
            for s in StorerKey.objects.filter(storerkey_code__in=all_storer_codes)
            .select_related("hub", "client")
        }

        cc_codes = {
            (
                l["po"].get("plant_id"),
                l["po"].get("center_code"), 
                l["po"].get("sloc") 
            ): CostCenterCode.objects.filter(
                plant_id = l["po"].get("plant_id"),
                center_code = l["po"].get("center_code"),
                sloc = l["po"].get("sloc")
            )
            for lines in valid_po_groups.values()
            for l in lines
        }

        po_to_create = []
        lines_to_create, lines_to_update, lines_to_cancel = [], [], []
        material_to_create = []
        storerkeys_to_update, cc_codes_to_create = [], []
        unique_cc_codes = {}

        ## helper to add errors
        def add_error(po_crn,pol_crn, message):
            error_for_sheet.append({(po_crn,pol_crn): message})
            error_in_pos.append(po_crn)

        try:
            
            for pol in lines_data:

                po = pol.get("po")
                po_crn = po.get("customer_reference_number")
                if po_crn in error_in_pos:
                    continue

                pol_crn = pol.get("customer_reference_number")           
                purchase_order = existing_pos.get(po_crn)
                purchase_order_line = None
                if purchase_order:
                    purchase_order_line = existing_lines.get((po_crn, pol["customer_reference_number"]))
                
                pol.pop("po")

                supplier_data = po.get("seller_details", {})
                storerkey_code = po.get("storerkey",None)

                supplier = suppliers.get(supplier_data.get("seller_code"))
                if not supplier:
                    add_error(po_crn,pol_crn,"Supplier not found.")
                    continue
                
                storerkey = storerkeys.get(storerkey_code)
                if not storerkey:
                    add_error(po_crn,pol_crn,"StorerKey not found.")
                    continue
                
                if po.get("order_type") in [OrderTypeChoices.BTS, OrderTypeChoices.BTO]:
                    if storerkey.order_type != OrderTypeChoices.BOTH:
                        if storerkey.order_type != po.get("order_type"):
                            add_error(po_crn,pol_crn,f"{po.get("order_type")} type POs not allowed for this storerkey")
                            continue
                
                else:
                    add_error(po_crn,pol_crn,"Invalid Storerkey po type.")
                    continue
                
                plant_id = po.get("plant_id","")
                center_code = po.get("center_code","") 
                sloc = po.get("sloc","") 
                
                key = (plant_id,center_code,sloc)
                cc_code = cc_codes.get(key)

                if not storerkey.cc_code:
                    add_error(po_crn, pol_crn, "Storerkey dont have Cost Center Code.")
                    continue

                if not cc_code:
                    add_error(po_crn, pol_crn, "Cost Center Code not found.")
                    continue

                cc_value = cc_code.first()
                if storerkey.cc_code != cc_value:
                    add_error(po_crn, pol_crn, "Storerkey not linked with Cost Center Code.")
                    continue
                
                # if not cc_code:

                #     existing_cc_code = unique_cc_codes.get(key)
                #     if existing_cc_code:
                #         storerkey.cc_code = existing_cc_code
                #         storerkeys_to_update.append(storerkey)
                    
                #     else:

                #         new_code = CostCenterCode(
                #             plant_id=plant_id,
                #             center_code=center_code,
                #             sloc=sloc,
                #             cc_code = plant_id + center_code + sloc)
                    
                #         unique_cc_codes[key] = new_code 
                #         cc_codes_to_create.append(new_code)
                #         storerkey.cc_code = new_code
                #         storerkeys_to_update.append(storerkey)
                # else:
                #     if storerkey.cc_code != cc_code.first():
                #         add_error(po_crn, pol_crn, "Storerkey not linked with CC Centre.")
                #         continue

                supplier_storerkeys = list(storerkey.suppliers.all())
                if supplier not in supplier_storerkeys:
                    add_error(po_crn, pol_crn, "Storerkey not matched with supplier.")
                    continue
                

                client = storerkey.client
                if not client:
                    add_error(po_crn,pol_crn,"Client not found.")
                    continue

                hub = storerkey.hub

                if client != supplier.client:
                    add_error(po_crn,pol_crn,"Seller code and buyer code are not linked.")
                    continue
                
                po_lines = valid_po_groups[po_crn]
                line_crns = [l["customer_reference_number"] for l in po_lines]
                duplicates = [crn for crn in set(line_crns) if line_crns.count(crn) > 1]
                if duplicates:
                    for crn in duplicates:
                        add_error(po_crn,pol_crn,f"Duplicate PO line '{crn}' in PO '{po_crn}'.")
                    continue

                po.update({
                    "id" : uuid4() if not purchase_order else purchase_order.id,
                    "client": client,
                    "supplier": supplier,
                    "storerkey": storerkey
                })
                po.update(SLBPOImportService.seller_details(supplier_data))
                
                
                # Clean unnecessary fields
                for key in ["storer_key", "seller_details", "pieces_detail", "supplier_code", "buyer_code"]:
                    po.pop(key, None)
                
                if purchase_order:
                    existing_lines_crn = list(purchase_order.lines.values_list("customer_reference_number", flat=True))
                    missing_crns = set(existing_lines_crn) - set(line_crns)
                    
                    lines_to_cancel.extend(
                        PurchaseOrderLine.objects.filter(
                            purchase_order=purchase_order, customer_reference_number__in=missing_crns
                        )
                    )

                else:
                    if "quantity" in po:
                        po["open_quantity"] = po["quantity"]
                        # del po["quantity"]


                    keys_to_remove = ("sloc","quantity")
                    for key in keys_to_remove:
                        po.pop(key, None)
                    
                    new_po = PurchaseOrder(**po)
                    if po_crn in error_in_pos and new_po in po_to_create:
                        po_to_create.remove(new_po)
                        continue
                    
                    existing_po = next((p for p in po_to_create if p.customer_reference_number == po_crn), None)
                    if not existing_po:
                        po_to_create.append(
                            PurchaseOrder(**po)
                        )
                    else:
                        po["id"] = existing_po.id
                        
                
                if "purchase_order_crn" in pol:
                    del pol["purchase_order_crn"]

                product_code = po_crn + pol.get("customer_reference_number", "") + "NOSKU"
                pol["product_code"] = pol.get("product_code") or product_code     


                material = None
                if not MaterialMaster.objects.filter(
                    product_code=product_code,
                    storerkey=storerkey,
                    hub=hub).exists():
                    
                    material = MaterialMaster(
                        storerkey=storerkey,
                        hub=hub,
                        product_code=product_code,
                        is_chemical = pol.get("is_chemical"),
                        is_dangerous_good = pol.get("is_dangerous_good")
                    )

                    material_to_create.append(material)

                if purchase_order_line:
                    data = pol
                    data, error = POLineService.po_line_quantity_validations(data,purchase_order_line)
                    
                    if error:
                        add_error(po_crn,pol_crn,error)
                        if material and material in material_to_create:
                            material_to_create.remove(material)
                        continue

                    purchase_order_line.quantity = data["quantity"]
                    purchase_order_line.status = data["status"]
                    purchase_order_line.open_quantity = data["open_quantity"]
                    purchase_order_line.purchase_order = purchase_order
                    lines_to_update.append(purchase_order_line)
                    continue
                    # lines_to_update.append(PurchaseOrderLine(**pol))
        
                pol["purchase_order_id"] = po.get("id")
                pol["open_quantity"] = pol.get("quantity")
                
                if not po_crn in error_in_pos:
                    lines_to_create.append(PurchaseOrderLine(**pol))
            pos = None
            ## Create all the records
            if material_to_create:
                MaterialMaster.objects.bulk_create(material_to_create)
            if po_to_create:
                pos = PurchaseOrder.objects.bulk_create(po_to_create)
            if lines_to_create:
                PurchaseOrderLine.objects.bulk_create(lines_to_create)
            if cc_codes_to_create:
                CostCenterCode.objects.bulk_create(cc_codes_to_create)
            
            ## update all the records
            if lines_to_update:
                PurchaseOrderLine.objects.bulk_update(lines_to_update,fields=["status","quantity","open_quantity"])
            if storerkeys_to_update:
                StorerKey.objects.bulk_update(storerkeys_to_update,fields=["cc_code"])

            # if po_to_update:
            #     PurchaseOrder.objects.bulk_update(po_to_update,fields = ["storerkey","supplier","client"])
            
            if lines_to_cancel:
                for line in lines_to_cancel:
                    if not ConsignmentPOLine.objects.filter(purchase_order_line_id = line.id).exists():
                        line.status = PurchaseOrderStatusChoices.CANCELLED
                PurchaseOrderLine.objects.bulk_update(lines_to_cancel, fields=["status"])
            
            if pos:
                created_ids = [po.id for po in pos]
                qs = PurchaseOrder.objects.filter(id__in=created_ids)
                PurchaseOrderService.update_open_quantity(qs)
            

            return po_to_create , error_for_sheet
            
        except Exception as e:
            transaction.set_rollback(True)
            error_for_sheet.append({("UNKNOWN", "UNKNOWN") : str(e)})
            return None, error_for_sheet
    
    
    @staticmethod
    def update_errors_in_file(file_obj, errors):
        file = file_obj.uploaded_file
        file.seek(0)
        file_content = file.read().decode("utf-8-sig")

        reader = csv.reader(file_content.splitlines())
        rows = list(reader)
        if not rows:
            raise ServiceError(error="Uploaded CSV is empty")

        # Normalize headers
        headers = [h.strip().lower().replace(" ", "_") for h in rows[0]]

        # Ensure columns exist
        if "error" not in headers:
            headers.append("error")
        if "unexpected_errors" not in headers:
            headers.append("unexpected_errors")

        error_col_index = headers.index("error")
        unexpected_col_index = headers.index("unexpected_errors")

        # Find PO + POL column indexes
        try:
            po_col_index = headers.index("purchase_order_number")
        except ValueError:
            raise ServiceError(error="purchase_order_number column not found")

        pol_col_index = None
        if "po_line_number" in headers:
            pol_col_index = headers.index("po_line_number")

        po_error_map = {}
        ## Converting from list to flat dict
        for e in errors:
            po_error_map.update(e)

        # Prepare final rows
        updated_rows = [headers]

        unexpected_key = ("UNKNOWN", "UNKNOWN")
        unexpected_msg = po_error_map.get(unexpected_key, None)

        for row in rows[1:]:
            po_number = row[po_col_index] if len(row) > po_col_index else None
            pol_number = row[pol_col_index] if pol_col_index is not None and len(row) > pol_col_index else None
            
            # Normal error for that row
            error_msg = po_error_map.get((po_number, pol_number), "")

            # Append the errors (in correct columns)
            row += [""] * (len(headers) - len(row))  # ensure row length

            row[error_col_index] = error_msg

            updated_rows.append(row)

        if unexpected_msg:
            if len(updated_rows) > 1:
                # place into the first actual data row (preserves existing data)
                updated_rows[1][unexpected_col_index] = unexpected_msg
            else:
                # no data rows existed originally: create one blank data row and set unexpected
                blank_row = [""] * len(headers)
                blank_row[unexpected_col_index] = unexpected_msg
                updated_rows.append(blank_row)


        # Write updated CSV
        output_str = StringIO()
        writer = csv.writer(output_str, lineterminator="\n")
        writer.writerows(updated_rows)

        updated_bytes = BytesIO(output_str.getvalue().encode("utf-8"))
        file_obj.error_file.save("updated_with_errors.csv", updated_bytes)
        file_obj.status = POUploadStatusChoices.ERROR
        file_obj.save()

        return False

        
    @classmethod
    def process_slb_po_file(cls, obj):
        
        file = obj.uploaded_file
        
        # Step 3: Run your business logic
        pol_entries = SLBPOImportService.parse_excel_to_pos(file)
        _ , errors = SLBPOImportService.create_purchase_orders(pol_entries)

        # Step 4: If there are errors, create updated CSV
        if errors:
            resp = SLBPOImportService.update_errors_in_file(file_obj=obj,errors=errors)
            
            NotificationService.notify_po_file_upload(
                user=obj.uploaded_by,
                header="Bulk PO Upload",
                message="File has errors",
                po_upload=obj
            )
            return resp
            
        # Step 7: If successful
        obj.status = POUploadStatusChoices.SUCCESS
        obj.save()

        NotificationService.notify_po_file_upload(
            user=obj.uploaded_by,
            header="Bulk PO Upload",
            message="File processed successfully.",
            po_upload=obj
        )
        
        return True

        

        
    

