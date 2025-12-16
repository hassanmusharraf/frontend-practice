from celery import shared_task
from .models import PurchaseOrder, PurchaseOrderLine,Consignment, PurchaseOrderUpload, ComprehensiveReport
from entities.models import Client, Supplier, StorerKey
from django.db import transaction, IntegrityError
from uuid import uuid4
from datetime import datetime
from portal.utils import convert_to_decimal
from django.utils import timezone
from datetime import timedelta
from portal.choices import POUploadStatusChoices
from openpyxl.styles import Font, PatternFill
from openpyxl import load_workbook
from django.core.files import File
import os
from .mixins import PurchaseOrderMixin
from operations.services import ComprehensiveReportService 
# import datetime

def _is_valid_date(date_val, field_name):

    # print(f"name value and type {field_name} ---- {date_val} ---- {type(date_val)}")

    if isinstance(date_val, str):
        try:
            datetime.strptime(date_val, "%Y-%m-%d")
        except ValueError:
            return False, f"Invalid date format for {field_name}, expected YYYY-MM-DD."

    return True , ""

def _handle_unexpected_errors(temp_obj,wb,e):
    print("unexpected error",str(e))
    po_header_sheet = wb["PO Header"]
    poh_headers = [str(cell.value).strip().lower().replace(" ", "_") for cell in next(po_header_sheet.iter_rows(min_row=1, max_row=1))]
    
    error_col_index = len(poh_headers) + 2
    error_cell = po_header_sheet.cell(row=1, column=error_col_index, value="UNEXPECTED ERROR")
    error_cell.font = Font(bold=True)
    
    for i, row in enumerate(po_header_sheet.iter_rows(min_row=2), start=2):
        po_header_sheet.cell(row=i, column=error_col_index, value=str(e))
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        for col_index in range(1, len(poh_headers) + 1):
            cell = po_header_sheet.cell(row=i, column=col_index)
            cell.fill = red_fill
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"Purchase_Order_Errors_{timestamp}.xlsx"  
        error_dir = os.path.join("media", "po", "upload", "errors")
        if not os.path.exists(error_dir):
            os.makedirs(error_dir, exist_ok=True)
        filename = os.path.join(error_dir, file_name)
        wb.save(filename)
        
        with open(filename, "rb") as f:
            temp_obj.error_file.save(file_name, File(f), save=False)

        # temp_obj.status = POUploadStatusChoices.ERROR
        temp_obj.save()
        return

def parse_date_field(val):
    if isinstance(val, datetime):
        return val
    if isinstance(val, str):
        # Try common formats
        try:
            return datetime.strptime(val, "%Y-%m-%d")
        except ValueError:
            pass
        # If none matched, raise or return None
        return None
    return None

def file_has_error(wb, suppliers, storerkeys, existing_po, existing_po_lines):
    po_header_sheet = wb["PO Header"]
    poh_headers = [str(cell.value).strip().lower().replace(" ", "_") for cell in next(po_header_sheet.iter_rows(min_row=1, max_row=1))]
    
    po_lines_sheet = wb["PO Lines"]
    pol_headers = [str(cell.value).strip().lower().replace(" ", "_") for cell in next(po_lines_sheet.iter_rows(min_row=1, max_row=1))]
        
    is_error = False
    po_header_data = []
    po_lines_data = []
    error_col_index = len(poh_headers) + 1
    error_cell = po_header_sheet.cell(row=1, column=error_col_index, value="ERROR")
    error_cell.font = Font(bold=True)
    
    excel_poh_crn = []    
    for i, row in enumerate(po_header_sheet.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in row[:len(poh_headers)]]  # exclude Error column
        if not any(values):
            break

        poh_crn = str(values[1])
        supplier = str(values[2])
        storerkey = str(values[3])
        order_due_date = values[18]
        expected_delivery_date = values[17]
        order_date = values[16]

        error_msg = ""
        if poh_crn in excel_poh_crn:
            error_msg = f"Duplicate Customer Reference found {values[0]}, "
        else:
            excel_poh_crn.append(poh_crn) 
            
        if poh_crn in existing_po:
            error_msg = error_msg + f", {poh_crn} already exists"
            
        if supplier not in suppliers:
            error_msg = error_msg + ", Supplier not found"
            
        if storerkey not in storerkeys:
            error_msg = error_msg + ", Storer Key not found"
            
        if order_due_date:    # order due date
            error,msg = _is_valid_date(order_due_date, "Order Due Date")
            if error:
                error_msg += msg
            # try:
            #     datetime.strptime(order_due_date, "%Y-%m-%d")
            # except:
            #     error_msg = error_msg + f"Invalid date format for Order Due Date, Please add date in YYYY-MM-DD"
        
        if expected_delivery_date:    # expected delivery date
            error,msg = _is_valid_date(expected_delivery_date, "Expected Delivery Date")
            if error:
                error_msg += msg
            # try:
            #     datetime.strptime(expected_delivery_date, "%Y-%m-%d")
            # except:
            #     error_msg = error_msg + f"Invalid date format for Expected Delivery Date, Please add date in YYYY-MM-DD"
                
        if order_date:    # order date
            error,msg = _is_valid_date(order_date, "Order Date")
            if error:
                error_msg += msg
            # try:
            #     datetime.strptime(order_date, "%Y-%m-%d")
            # except:
            #     error_msg = error_msg + f"Invalid date format for Order Date, Please add date in YYYY-MM-DD"

        # has_empty = any(v is None or (isinstance(v, str) and not v.strip()) for v in values)

        # if has_empty:
        #     error_msg = error_msg + "Missing value(s) in row, all columns are required."
            
        if error_msg:
            po_header_sheet.cell(row=i, column=error_col_index, value=error_msg)
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for col_index in range(1, len(poh_headers) + 1):
                cell = po_header_sheet.cell(row=i, column=col_index)
                cell.fill = red_fill
            is_error = True
            
        else:
            po_header_data.append(dict(zip(poh_headers, values)))
            
            
    error_col_index = len(pol_headers) + 1
    error_cell = po_lines_sheet.cell(row=1, column=error_col_index, value="ERROR")
    error_cell.font = Font(bold=True)
        
    for i, row in enumerate(po_lines_sheet.iter_rows(min_row=2), start=2):
        values = [cell.value for cell in row[:len(pol_headers)]]  # exclude Error column
        if not any(values):
            break 
        
        error_msg = ""
        
        poh_crn = str(values[0])
        order_due_date = values[22]
        expected_delivery_date = values[21]

        if poh_crn not in excel_poh_crn:
            error_msg = f"Purchase Order {poh_crn} not configured in PO Header sheet,"
         
        key = (poh_crn, str(values[2]))   
        if key in existing_po_lines:
            error_msg = error_msg + f", customer reference number already exists with respective PO"
        
        
        if order_due_date:    # order due date
            error,msg = _is_valid_date(order_due_date, "Order Due Date")
            if error:
                error_msg += msg
            # try:
            #     datetime.strptime(order_due_date, "%Y-%m-%d")
            # except:
            #     error_msg = error_msg + f"Invalid date format for Order Due Date, Please add date in YYYY-MM-DD"
        
        if expected_delivery_date:
            error,msg = _is_valid_date(expected_delivery_date, "Expected Delivery Date")
            if error:
                error_msg += msg

        # has_empty = any(v is None or (isinstance(v, str) and not v.strip()) for v in values)

        # if has_empty:
        #     error_msg = error_msg + "Missing value(s) in row, all columns are required."
        
        if error_msg:
            po_lines_sheet.cell(row=i, column=error_col_index, value=error_msg)
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for col_index in range(1, len(pol_headers) + 1):
                cell = po_lines_sheet.cell(row=i, column=col_index)
                cell.fill = red_fill
            is_error = True
            
        else:
            po_lines_data.append(dict(zip(pol_headers, values)))
    return is_error, po_header_data, po_lines_data

def process_purchase_orders_v2(temp_order_id):
    temp_obj = PurchaseOrderUpload.objects.filter(id=temp_order_id).first()
    if not temp_obj:
        print("No temporary purchase orders found")
        return f"No data to process {temp_order_id}"
    
    handler = PurchaseOrderMixin()
    return handler.process_excel_file(temp_obj)

@shared_task
def process_purchase_orders(temp_order_id):
    temp_obj = PurchaseOrderUpload.objects.filter(id=temp_order_id).first()
    if not temp_obj:
        print("No temporary purchase orders found")
        return f"No data to process {temp_order_id}"
    
    file = temp_obj.uploaded_file
    try:

        wb = load_workbook(file, data_only=True)
        suppliers = {supplier.supplier_code: supplier for supplier in Supplier.objects.filter()}
        client = {client.client_code: client for client in Client.objects.filter()}
        storerkeys = {storerkey.storerkey_code: storerkey for storerkey in StorerKey.objects.filter().select_related("client")}
        existing_po = {po.customer_reference_number: po for po in PurchaseOrder.objects.filter()}
        existing_po_lines = {(po_line.purchase_order.customer_reference_number, po_line.customer_reference_number): po_line for po_line in PurchaseOrderLine.objects.filter().select_related("purchase_order")}
        is_error, po_header_data, po_lines_data = file_has_error(wb, suppliers, storerkeys, existing_po, existing_po_lines)
        
        if is_error: 
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            file_name = f"Purchase_Order_Errors_{timestamp}.xlsx"  
            error_dir = os.path.join("media", "po", "upload", "errors")
            if not os.path.exists(error_dir):
                os.makedirs(error_dir, exist_ok=True)
            filename = os.path.join(error_dir, file_name)
            wb.save(filename)
            
            with open(filename, "rb") as f:
                temp_obj.error_file.save(file_name, File(f), save=False)

            temp_obj.status = POUploadStatusChoices.ERROR
            temp_obj.save()
            return
            
            
        po_to_create = []
        po_lines_to_create = []


        for item in po_header_data:
            order_quantity = convert_to_decimal(item.get("quantity"), 2)
            data = {
                "reference_number": item.get("reference_number"),
                "customer_reference_number": item.get("customer_reference_number"),
                "inco_terms": item.get("inco_terms"),
                "supplier": suppliers[str(item.get("supplier_code"))],
                "storerkey": storerkeys[str(item.get("storer_key"))],
                "client": storerkeys[str(item.get("storer_key"))].client,
                "open_quantity": order_quantity,
                "center_code": item.get("center_code"),
                "plant_id": item.get("plant_id"),
                "description": item.get("description"),
                "expected_delivery_date": parse_date_field(item.get("expected_delivery_date")),
                "order_due_date": parse_date_field(item.get("order_due_date")),
                "client" : client[str(item.get("buyer_code"))],
                "group_code" : item.get("group_code"),
                "type" : item.get("type"),
                "notes" : item.get("notes"),
                "payment_terms" : item.get("payment_terms"),
                "destination_country" : item.get("destination_country"),
                "origin_country" : item.get("origin_country"),
                "order_date" : parse_date_field(item.get("order_date")),
            }
            po_to_create.append(PurchaseOrder(**data))
    except Exception as e:
        _handle_unexpected_errors(temp_obj,wb,e)

    with transaction.atomic():
        try:
            print(len(po_to_create))
            
            po_header_objects = PurchaseOrder.objects.bulk_create(po_to_create)
            purchase_order_mapping = {
                i.customer_reference_number: i.id
                for i in po_header_objects
            }
            # purchase_order_mapping = {
            #     i.get("customer_reference_number"): i.get("id") 
            #     for i in PurchaseOrder.objects.filter(id__in=po_header_ids).values("customer_reference_number", "id")
            # }
            
            for item in po_lines_data:
                quantity = convert_to_decimal(item.get("quantity"), 2)
                unit_price = convert_to_decimal(item.get("unit_price"), 2)
                unit_cost = convert_to_decimal(item.get("unit_cost"), 2)
                weight = convert_to_decimal(item.get("weight"), 2)
                volume = convert_to_decimal(item.get("volume"), 2)
                length = convert_to_decimal(item.get("length"), 2)
                width = convert_to_decimal(item.get("width"), 2)
                height = convert_to_decimal(item.get("height"), 2)
                is_chemical = True if item.get("chemical") else False
                is_dangerous_good = True if item.get("dangerous_good") else False
                data = {
                    "purchase_order_id": purchase_order_mapping[item.get("purchase_order_crn")],
                    "reference_number": item.get("reference_number"),
                    "customer_reference_number": item.get("customer_reference_number"),
                    "sku": item.get("sku"),
                    "quantity": quantity,
                    "open_quantity": quantity,
                    "description": item.get("description"),
                    "product_code": item.get("product_code"),
                    "hs_code": item.get("hs_code"),
                    "expected_delivery_date": parse_date_field(item.get("expected_delivery_date")),
                    "order_due_date": parse_date_field(item.get("order_due_date")),
                    "alternate_unit" : item.get("alternate_unit"),
                    "stock_number" : item.get("stock_number"),
                    "is_chemical" : is_chemical,
                    "is_dangerous_good" : is_dangerous_good,
                    "unit_price" : unit_price,
                    "unit_cost" : unit_cost,
                    "weight" : weight,
                    "volume" : volume,
                    "length" : length,
                    "width" : width,
                    "height" : height,
                    "source_location" : item.get("source_location"),
                    "notes" : item.get("notes"),                    
                }
                po_lines_to_create.append(PurchaseOrderLine(**data))
                
                
            print(len(po_lines_to_create))
            PurchaseOrderLine.objects.bulk_create(po_lines_to_create, batch_size=500)
            temp_obj.status = POUploadStatusChoices.SUCCESS
            temp_obj.save()
            return {
                "purchase_orders_created": len(po_to_create),
                "purchase_order_lines_created": len(po_lines_to_create),
            }
        except IntegrityError as e:
            transaction.set_rollback(True)
            _handle_unexpected_errors(temp_obj,wb,e)
            return str(e)
            # Handle error (e.g., log it)
        except Exception as e:
            transaction.set_rollback(True)
            _handle_unexpected_errors(temp_obj,wb,e)
            return str(e)
        # Handle error (e.g., log it)

# @shared_task
# def delete_old_staging_data():
#     cutoff_time = timezone.now() - timedelta(minutes=15)

#     deleted_count, _ = ConsignmentStaging.objects.filter(created_at__lt=cutoff_time).delete()
    
#     return f"{deleted_count} old staging records deleted"


# @shared_task(bind=True)
def generate_comperhensive_report(consignments_qs):
    
    data, errors = ComprehensiveReportService.json_data(consignments_qs)
    
    if errors:
        return errors
    
    if not data:
        return "Not data found for the given date range and status. Please try again."
    
    
    file_path, errors = ComprehensiveReportService.build_report(data)
    if errors:
        return errors
    
    return f"File generation successfull : {file_path} "